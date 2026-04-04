"""Import items from CONTROL_CENTER.xlsx into ERPNext Item master."""
import frappe
from frappe.utils import flt
import openpyxl

def run():
    wb = openpyxl.load_workbook(
        "/Users/sayanthns/frappe-bench/sites/mysite.local/public/files/CONTROL_CENTER.xlsx",
        data_only=True,
    )
    ws = wb["12_CUSTOMER_ITEM_MASTER"]

    # Ensure Item Group exists
    if not frappe.db.exists("Item Group", "Telecom Services"):
        frappe.get_doc({
            "doctype": "Item Group",
            "item_group_name": "Telecom Services",
            "parent_item_group": "All Item Groups",
        }).insert(ignore_permissions=True)
        print("Created Item Group: Telecom Services")

    uom_map = {
        "SITE": "Nos", "PCS": "Nos", "PAIR": "Nos", "SET": "Nos",
        "METER": "Meter", "KM": "Nos", "TRIP": "Nos", "CBM": "Nos", "LOT": "Nos",
    }

    created = 0
    skipped = 0
    errors = 0
    seen = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        item_code = row[1]
        if not item_code:
            continue
        item_code = str(item_code).strip()
        if item_code in seen:
            continue
        seen.add(item_code)

        if frappe.db.exists("Item", item_code):
            skipped += 1
            continue

        desc = str(row[5] or item_code)[:140]
        unit = str(row[6] or "").upper().strip()
        uom = uom_map.get(unit, "Nos")

        try:
            doc = frappe.get_doc({
                "doctype": "Item",
                "item_code": item_code,
                "item_name": desc,
                "item_group": "Telecom Services",
                "stock_uom": uom,
                "is_stock_item": 0,
                "description": str(row[5] or ""),
            })
            doc.insert(ignore_permissions=True)
            created += 1
        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"  Error {item_code}: {str(e)[:100]}")

    frappe.db.commit()
    print(f"Items: created={created}, skipped={skipped}, errors={errors}")


    # Also import into Customer Item Master doctype
    cim_created = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        item_code = row[1]
        if not item_code:
            continue
        item_code = str(item_code).strip()
        customer = str(row[0] or "").strip()

        try:
            frappe.get_doc({
                "doctype": "Customer Item Master",
                "customer": customer,
                "item_code": item_code,
                "customer_activity_type": str(row[2] or ""),
                "domain": str(row[4] or ""),
                "item_description": str(row[5] or ""),
                "unit_type": str(row[6] or ""),
                "standard_rate_sar": flt(row[7] or 0),
                "hard_rate_sar": flt(row[8] or 0),
                "active_flag": 1 if str(row[11] or "").strip().lower() == "yes" else 0,
            }).insert(ignore_permissions=True)
            cim_created += 1
        except Exception:
            pass

    frappe.db.commit()
    print(f"Customer Item Master: created={cim_created}")
    print(f"Total Items in system: {frappe.db.count('Item')}")
    print(f"Total CIM records: {frappe.db.count('Customer Item Master')}")
