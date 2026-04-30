"""
Command Center API — Pipeline Operations & Dashboard Aggregation
for inet_app (Frappe 15)
"""

import calendar
import csv
import json
import os
import re

import frappe
from inet_app.inet_app.doctype.po_intake.po_intake import normalize_po_intake_status as _normalize_po_intake_status
from inet_app.region_type import is_hard_region, region_type_from_center_area
from frappe.utils import (
    add_days,
    cint,
    flt,
    get_datetime,
    get_first_day,
    get_last_day,
    getdate,
    now_datetime,
    nowdate,
    time_diff_in_seconds,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _dashboard_etag(*discriminators):
    """Cheap version tag for dashboard payloads.

    Reads ``MAX(modified)`` from the three tables that any dashboard pulls
    from (PO Dispatch, Daily Execution, Work Done), plus the caller's
    discriminators (user, im, date range). Hashed to a short opaque token.

    Used for an If-None-Match-style short-circuit: if the FE passes the
    same token it had last time, the endpoint returns a tiny ``unchanged``
    payload instead of regenerating 5–15kB of JSON.
    """
    import hashlib
    try:
        rows = frappe.db.sql(
            """
            SELECT
              (SELECT UNIX_TIMESTAMP(MAX(modified)) FROM `tabPO Dispatch`)    AS pd,
              (SELECT UNIX_TIMESTAMP(MAX(modified)) FROM `tabDaily Execution`) AS de,
              (SELECT UNIX_TIMESTAMP(MAX(modified)) FROM `tabWork Done`)       AS wd
            """,
            as_dict=True,
        )
        row = rows[0] if rows else {}
    except Exception:
        row = {}
    parts = [
        str(row.get("pd") or 0),
        str(row.get("de") or 0),
        str(row.get("wd") or 0),
    ]
    for d in discriminators:
        parts.append("" if d is None else str(d))
    h = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()
    return h[:16]


def _iso_now():
    """Return current time as ISO-8601 with the site timezone offset.

    `frappe.utils.now()` returns a naive string in the site timezone with no
    offset, so a browser in a different timezone parses it as its own local
    time and the displayed "last updated" drifts by the offset. Emitting an
    offset-aware string makes the FE's `new Date(s)` parse to the correct
    instant regardless of browser timezone.
    """
    tz_name = frappe.utils.get_system_timezone() or "UTC"
    try:
        from zoneinfo import ZoneInfo
        tz = ZoneInfo(tz_name)
    except Exception:
        try:
            import pytz
            tz = pytz.timezone(tz_name)
        except Exception:
            import datetime as _dt
            return _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds")
    import datetime as _dt
    return _dt.datetime.now(tz).isoformat(timespec="seconds")


def _parse_rows(rows):
    """Accept a JSON string or a Python list/dict and always return a list."""
    if isinstance(rows, str):
        rows = frappe.parse_json(rows)
    if isinstance(rows, dict):
        rows = [rows]
    return rows or []


def _portal_row_limit(limit, default=500):
    """
    Normalize portal list size.

    - ``0`` = unlimited (omit ``LIMIT`` in raw SQL; use ``limit_page_length=0`` / ``page_length=0``
      in Frappe list APIs so no row cap is applied).
    - Positive values are clamped to 1..10000.
    - Missing/invalid uses ``default`` (must be >= 1).
    """
    try:
        lim = cint(limit)
    except Exception:
        lim = -1
    if lim == 0:
        return 0
    if lim < 1:
        lim = cint(default) or 500
    return int(min(max(lim, 1), 10000))


def _sql_limit_suffix(limit_val):
    """Append to SQL; empty string when ``limit_val`` is 0 (unlimited)."""
    if not limit_val:
        return ""
    return f" LIMIT {int(limit_val)}"


def _portal_filters_dict(portal_filters):
    """Parse optional portal_filters JSON from the SPA into a dict."""
    if isinstance(portal_filters, str):
        portal_filters = frappe.parse_json(portal_filters) if portal_filters else {}
    return portal_filters if isinstance(portal_filters, dict) else {}


def _ensure_list(raw):
    """Coerce a filter value (string / JSON array / list) into a deduped list
    of non-empty stripped strings. Enables multi-select on any filter that
    previously only accepted a single value."""
    if raw is None:
        return []
    if isinstance(raw, (list, tuple, set)):
        items = list(raw)
    elif isinstance(raw, str):
        s = raw.strip()
        if not s:
            return []
        # JSON array?
        if s[0] == "[":
            try:
                parsed = json.loads(s)
                if isinstance(parsed, list):
                    items = parsed
                else:
                    items = [s]
            except Exception:
                items = [s]
        else:
            items = [s]
    else:
        items = [raw]
    seen = set()
    out = []
    for it in items:
        v = str(it or "").strip()
        if not v or v in seen:
            continue
        seen.add(v)
        out.append(v)
    return out


def _sql_in_or_eq(expr, raw):
    """Build ``expr = %s`` or ``expr IN (%s, %s, …)`` with params for a single
    or multi-value filter. Returns (clause_or_None, params)."""
    vals = _ensure_list(raw)
    if not vals:
        return None, []
    if len(vals) == 1:
        return f"{expr} = %s", vals
    ph = ", ".join(["%s"] * len(vals))
    return f"{expr} IN ({ph})", vals


def _sql_like_pattern(term):
    """Build a LIKE pattern with % wildcards; escape % and _ in user input."""
    t = (term or "").strip()
    if not t:
        return None
    t = t.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    return f"%{t}%"


def _sql_like_tokens(term, max_tokens=50):
    """Split a free-text search on whitespace / commas / semicolons / pipes
    into individual LIKE patterns. Any row matching any token is returned
    (OR'd). Multi-word phrases are not preserved — users who need phrase
    matching should paste a distinctive substring instead."""
    if not term:
        return []
    raw = re.split(r"[\s,;|]+", str(term))
    seen = set()
    out = []
    for piece in raw:
        s = (piece or "").strip()
        if not s or s in seen:
            continue
        seen.add(s)
        esc = s.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        out.append(f"%{esc}%")
        if len(out) >= max_tokens:
            break
    return out


def _sql_search_clause(concat_expr, term):
    """Build an OR'd LIKE clause across pasted tokens for a concat expression.
    Returns (clause_sql_or_None, params_list). Empty term → (None, [])."""
    patterns = _sql_like_tokens(term)
    if not patterns:
        return None, []
    ors = " OR ".join([f"{concat_expr} LIKE %s"] * len(patterns))
    return f"({ors})", patterns


def _po_dispatch_col_expr(col, alias=None, prefix="pd"):
    """Return ``"<prefix>.<col> AS <alias>"`` if the column exists on PO Dispatch,
    otherwise ``"NULL AS <alias>"``. Lets queries that touch optional columns
    survive on environments where a migration hasn't run yet.
    """
    alias = alias or col
    if frappe.db.has_column("PO Dispatch", col):
        return f"{prefix}.{col} AS {alias}"
    return f"NULL AS {alias}"


def _remark_select(prefix="pd"):
    """SELECT-clause fragment for the role-scoped remark trio. Falls back to
    NULLs on environments that haven't migrated the new columns yet."""
    return ", ".join([
        _po_dispatch_col_expr("general_remark", prefix=prefix),
        _po_dispatch_col_expr("manager_remark", prefix=prefix),
        _po_dispatch_col_expr("team_lead_remark", prefix=prefix),
    ])


_PIC_STATUS_TO_BILLING = {
    # PIC scale (11 values) → Work Done billing bucket (Pending / Invoiced / Closed).
    "Commercial Invoice Closed":    "Closed",
    "PO Line Canceled":             "Closed",
    "Commercial Invoice Submitted": "Invoiced",
    "Ready for Invoice":            "Invoiced",
    "Under I-BUY":                  "Invoiced",
    "Under ISDP":                   "Invoiced",
    # Anything else (Work Not Done, Under Process to Apply, both Rejected
    # variants, PO Need to Cancel) maps to Pending.
}


def _billing_status_from_pic(pic_status, fallback=None):
    """Roll the 11-value PIC scale up to the 3-bucket billing_status used by
    PM / IM Work Done views. Falls back to the existing Work Done value when
    no PIC status has been set yet."""
    if pic_status:
        return _PIC_STATUS_TO_BILLING.get(pic_status, "Pending")
    return fallback or "Pending"


def _normalize_int_token(value):
    """Render a numeric value as a clean integer string when it has no fractional part.

    Excel binary parsers hand back ``1.0`` for what should be an integer; if we
    drop that into a POID we end up with ugly suffixes like ``-1-1.0``.
    """
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    s = str(value).strip()
    # Strip a trailing ".0" left over from float-to-string conversions upstream.
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        return s[:-2]
    return s


def _make_poid(po_no, po_line_no, shipment_number):
    """Build POID: PO No - PO Line No - Shipment No."""
    parts = [str(po_no or "").strip(), str(cint(po_line_no) if po_line_no else 0)]
    ship = _normalize_int_token(shipment_number)
    if ship and ship != "0":
        parts.append(ship)
    return "-".join(parts)


def _resolve_line_poid(po_no, po_line_no, shipment_number, fallback=None):
    """Prefer explicit POID from intake line; otherwise rebuild from line metadata."""
    poid = (fallback or "").strip()
    return poid or _make_poid(po_no, po_line_no, shipment_number)


def _resolve_customer_link_name(customer_input):
    """Resolve Excel/UI label to Customer.name for Link fields."""
    c = (customer_input or "").strip()
    if not c:
        return None
    if frappe.db.exists("Customer", c):
        return c
    return frappe.db.get_value("Customer", {"customer_name": c}, "name")


def ensure_item_master(item_code_input, item_name=None):
    """
    Create a minimal Item if missing (for PO import). Requires ERPNext Item / Item Group.
    """
    code = (item_code_input or "").strip()
    if not code:
        return False, "item_code is required"
    if frappe.db.exists("Item", code):
        return True, None
    if not frappe.db.exists("DocType", "Item"):
        return False, "Item master is not available (install ERPNext Stock or create Item manually)"
    ig = frappe.db.get_value("Item Group", {"is_group": 0}, "name")
    if not ig:
        ig = frappe.db.sql_list("SELECT name FROM `tabItem Group` LIMIT 1")
        ig = ig[0] if ig else None
    if not ig:
        return False, "No Item Group found — create one under Stock"
    uom = frappe.db.get_single_value("Stock Settings", "stock_uom") or "Nos"
    try:
        it = frappe.new_doc("Item")
        it.item_code = code
        it.item_name = (item_name or code)[:140]
        it.item_group = ig
        it.stock_uom = uom
        if hasattr(it, "is_stock_item"):
            it.is_stock_item = 0
        it.insert(ignore_permissions=True)
    except Exception as e:
        if frappe.db.exists("Item", code):
            return True, None
        return False, str(e)
    return True, None


def ensure_project_control_center(project_code, customer_input, project_name=None, center_area=None):
    """Create Project Control Center when missing (PO import)."""
    pc = (project_code or "").strip()
    if not pc:
        return False, "project_code is required"
    if frappe.db.exists("Project Control Center", pc):
        return True, None
    cust = _resolve_customer_link_name(customer_input)
    if not cust:
        return False, "Customer is required to auto-create project"
    try:
        doc = frappe.new_doc("Project Control Center")
        doc.project_code = pc
        doc.project_name = ((project_name or pc) or "").strip()[:140] or pc
        doc.customer = cust
        if center_area:
            doc.center_area = str(center_area).strip()
        if hasattr(doc, "region_type"):
            doc.region_type = region_type_from_center_area(doc.center_area)
        doc.active_flag = "Yes"
        doc.project_status = "Active"
        doc.insert(ignore_permissions=True)
    except Exception as e:
        return False, str(e)
    return True, None


def ensure_customer_item_master(customer_input, item_code_input):
    """
    Ensure an active Customer Item Master exists for this customer and Item.
    Creates one with zero rates if missing; reactivates if an inactive row exists.

    Returns
    -------
    (ok: bool, err: str|None)
    """
    customer = _resolve_customer_link_name(customer_input)
    item_code = (item_code_input or "").strip()
    if not customer:
        return False, "Customer not found in Customer master"
    if not item_code:
        return False, "item_code is required"
    ok_item, err_item = ensure_item_master(item_code)
    if not ok_item:
        return False, err_item

    existing = frappe.db.get_value(
        "Customer Item Master",
        {"customer": customer, "item_code": item_code},
        ["name", "active_flag"],
        as_dict=True,
    )
    if existing:
        if not cint(existing.active_flag):
            frappe.db.set_value(
                "Customer Item Master",
                existing.name,
                "active_flag",
                1,
                update_modified=True,
            )
        return True, None

    try:
        doc = frappe.new_doc("Customer Item Master")
        doc.customer = customer
        doc.item_code = item_code
        doc.active_flag = 1
        doc.standard_rate_sar = 0
        doc.hard_rate_sar = 0
        doc.insert(ignore_permissions=True)
    except Exception as e:
        if frappe.db.exists(
            "Customer Item Master", {"customer": customer, "item_code": item_code}
        ):
            return True, None
        return False, str(e)

    return True, None


def ensure_duid_master(duid_input, site_name=None, center_area=None):
    """Ensure DUID Master exists for site_code; DUID and site_code are same."""
    duid = str(duid_input or "").strip()
    if not duid:
        return True, None
    if not frappe.db.exists("DocType", "DUID Master"):
        return True, None

    existing = frappe.db.get_value(
        "DUID Master",
        duid,
        ["name", "site_name", "center_area"],
        as_dict=True,
    )
    if existing:
        updates = {}
        if site_name and not existing.get("site_name"):
            updates["site_name"] = str(site_name).strip()
        if center_area and not existing.get("center_area"):
            updates["center_area"] = str(center_area).strip()
        if updates:
            frappe.db.set_value("DUID Master", duid, updates, update_modified=False)
        return True, None

    try:
        doc = frappe.new_doc("DUID Master")
        doc.duid = duid
        doc.site_name = (site_name or "")[:140]
        doc.center_area = (center_area or "")[:140]
        doc.insert(ignore_permissions=True)
    except Exception as e:
        if frappe.db.exists("DUID Master", duid):
            return True, None
        return False, str(e)
    return True, None


def _sanitize_table_pref_config(config):
    """Sanitize table personalization payload."""
    if isinstance(config, str):
        config = frappe.parse_json(config)
    config = config or {}
    if not isinstance(config, dict):
        return {}

    out = {}

    order = config.get("order")
    if isinstance(order, list):
        out["order"] = [str(x) for x in order if str(x).strip()][:200]

    hidden = config.get("hidden")
    if isinstance(hidden, list):
        out["hidden"] = [str(x) for x in hidden if str(x).strip()][:200]

    widths = config.get("widths")
    if isinstance(widths, dict):
        clean_widths = {}
        for k, v in widths.items():
            key = str(k).strip()
            if not key:
                continue
            px = cint(v)
            # Guardrails
            px = max(60, min(px, 1200))
            clean_widths[key] = px
        out["widths"] = clean_widths

    filters = config.get("filters")
    if isinstance(filters, dict):
        clean_filters = {}
        for k, v in filters.items():
            key = str(k).strip()
            if not key:
                continue
            clean_filters[key] = str(v or "")[:200]
        out["filters"] = clean_filters

    out["show_filters"] = 1 if cint(config.get("show_filters")) else 0

    dyn = config.get("dynamic_fields")
    if isinstance(dyn, list):
        clean_dyn = []
        for item in dyn[:40]:
            if not isinstance(item, dict):
                continue
            key = str(item.get("key") or "").strip()[:120]
            if not key or not key.startswith("dyn_"):
                continue
            dt = str(item.get("doctype") or "").strip()[:120]
            fn = str(item.get("fieldname") or "").strip()[:120]
            sk = str(item.get("source_key") or "").strip()[:120]
            if not dt or not fn or not sk:
                continue
            clean_dyn.append(
                {
                    "key": key,
                    "doctype": dt,
                    "fieldname": fn,
                    "source_key": sk,
                    "label": str(item.get("label") or fn)[:200],
                }
            )
        out["dynamic_fields"] = clean_dyn

    return out


# Frappe stores list-view prefs in `__UserSettings` (user, doctype, data) — not on `tabUser`.
# Use a dedicated synthetic doctype key so portal table layouts persist reliably.
_PORTAL_TABLE_PREFS_DOCTYPE = "INET PMS Table Prefs"


def _get_all_table_prefs_for_user(user):
    if not user or user == "Guest":
        return {}
    try:
        rows = frappe.db.sql(
            "select `data` from `__UserSettings` where `user`=%s and `doctype`=%s",
            (user, _PORTAL_TABLE_PREFS_DOCTYPE),
        )
    except Exception:
        rows = None
    if rows and rows[0] and rows[0][0]:
        try:
            blob = rows[0][0]
            parsed = frappe.parse_json(blob) if isinstance(blob, str) else blob
        except Exception:
            parsed = {}
        if isinstance(parsed, dict):
            inner = parsed.get("inet_table_preferences")
            if isinstance(inner, dict):
                return inner
    # Legacy: some sites may still have JSON on User (if column exists)
    if frappe.db.has_column("User", "user_settings"):
        try:
            raw = frappe.db.get_value("User", user, "user_settings")
        except Exception:
            raw = None
        if raw:
            try:
                parsed = frappe.parse_json(raw) if isinstance(raw, str) else raw
            except Exception:
                parsed = {}
            if isinstance(parsed, dict):
                inner = parsed.get("inet_table_preferences") or {}
                if isinstance(inner, dict):
                    return inner
    return {}


def _save_all_table_prefs_for_user(user, prefs):
    """Persist portal table prefs for the session user (same mechanism as Desk __UserSettings)."""
    if not user or user == "Guest":
        frappe.throw("Not permitted", frappe.PermissionError)
    if user != frappe.session.user:
        frappe.throw("Not permitted", frappe.PermissionError)
    payload = json.dumps({"inet_table_preferences": prefs})
    frappe.db.multisql(
        {
            "mariadb": (
                "INSERT INTO `__UserSettings`(`user`, `doctype`, `data`) "
                "VALUES (%s, %s, %s) ON DUPLICATE KEY UPDATE `data`=%s"
            ),
            "postgres": (
                'INSERT INTO "__UserSettings" ("user", "doctype", "data") '
                'VALUES (%s, %s, %s) ON CONFLICT ("user", "doctype") DO UPDATE SET "data"=%s'
            ),
        },
        (user, _PORTAL_TABLE_PREFS_DOCTYPE, payload, payload),
    )
    try:
        frappe.cache.hset("_user_settings", f"{_PORTAL_TABLE_PREFS_DOCTYPE}::{user}", None)
    except Exception:
        pass


@frappe.whitelist()
def get_table_preferences(table_id):
    """Read per-user table preferences for a page/table id."""
    table_id = str(table_id or "").strip()
    if not table_id:
        frappe.throw("table_id is required")
    user = frappe.session.user
    prefs = _get_all_table_prefs_for_user(user)
    config = prefs.get(table_id) or {}
    return _sanitize_table_pref_config(config)


@frappe.whitelist()
def get_distinct_field_values(doctype, fields):
    """Return {fieldname: [distinct values]} for the given fields on a doctype.

    Powers filter dropdowns that must show ALL possible values regardless of
    the current row-limit. SQL-injection safe: doctype must exist, fields
    must appear in the doctype meta.
    """
    if not doctype or not isinstance(doctype, str):
        return {}
    if not frappe.db.exists("DocType", doctype):
        return {}

    if isinstance(fields, str):
        try:
            fields_list = frappe.parse_json(fields)
        except Exception:
            fields_list = []
    else:
        fields_list = fields
    if not isinstance(fields_list, list):
        return {}

    meta = frappe.get_meta(doctype)
    allowed = {f.fieldname for f in meta.fields}
    allowed.add("name")  # always safe

    out = {}
    table_name = "tab" + doctype
    for f in fields_list:
        if not isinstance(f, str) or f not in allowed:
            continue
        try:
            rows = frappe.db.sql(
                f"SELECT DISTINCT `{f}` FROM `{table_name}` "
                f"WHERE `{f}` IS NOT NULL AND `{f}` != '' "
                f"ORDER BY `{f}` LIMIT 5000",
                as_list=True,
            )
            out[f] = [r[0] for r in rows if r[0]]
        except Exception:
            out[f] = []
    return out


@frappe.whitelist()
def get_all_table_preferences():
    """Return every saved table preference for the current user in one call.

    Prefetched at app boot so per-table `load(tableId)` lookups are synchronous —
    removes a blocking round-trip that used to precede every table render.
    """
    user = frappe.session.user
    prefs = _get_all_table_prefs_for_user(user) or {}
    out = {}
    for k, v in prefs.items():
        if not isinstance(k, str):
            continue
        out[k] = _sanitize_table_pref_config(v or {})
    return out


@frappe.whitelist()
def save_table_preferences(table_id, config=None):
    """Save per-user table preferences for a page/table id."""
    table_id = str(table_id or "").strip()
    if not table_id:
        frappe.throw("table_id is required")
    if len(table_id) > 400:
        frappe.throw("table_id is too long")

    if isinstance(config, str):
        try:
            config = frappe.parse_json(config) if (config or "").strip() else {}
        except Exception:
            config = {}

    user = frappe.session.user
    prefs = _get_all_table_prefs_for_user(user)
    prefs[table_id] = _sanitize_table_pref_config(config)
    _save_all_table_prefs_for_user(user, prefs)
    return {"ok": True}


@frappe.whitelist()
def get_table_field_values(doctype, fieldname, names):
    """
    Fetch one field for a list of document names.
    Used by table personalization to add doctype-backed columns dynamically.
    """
    doctype = str(doctype or "").strip()
    fieldname = str(fieldname or "").strip()
    if isinstance(names, str):
        names = frappe.parse_json(names)
    names = names or []

    if not doctype or not fieldname:
        frappe.throw("doctype and fieldname are required")
    if not isinstance(names, list):
        frappe.throw("names must be a list")

    meta = frappe.get_meta(doctype)
    if not meta.has_field(fieldname) and fieldname != "name":
        frappe.throw(f"Field '{fieldname}' not found in {doctype}")

    rows = frappe.get_all(
        doctype,
        filters={"name": ["in", [str(n) for n in names if str(n).strip()]]},
        fields=["name", fieldname],
        limit_page_length=max(1, len(names)),
        ignore_permissions=True,
    )
    out = {r.get("name"): r.get(fieldname) for r in rows}
    return {"values": out}


@frappe.whitelist()
def get_doctype_fields(doctype):
    """Return selectable fields for a doctype as {label, fieldname}."""
    doctype = str(doctype or "").strip()
    if not doctype:
        frappe.throw("doctype is required")
    meta = frappe.get_meta(doctype)
    fields = [{"label": "Name", "fieldname": "name"}]
    for df in meta.fields:
        if not getattr(df, "fieldname", None):
            continue
        if df.fieldtype in ("Section Break", "Column Break", "Tab Break", "Button", "Fold", "Heading", "HTML"):
            continue
        fields.append({
            "label": (df.label or df.fieldname),
            "fieldname": df.fieldname,
        })
    return {"fields": fields}


def _insert_po_dispatch_with_poid(dispatch_doc, poid):
    """
    Insert PO Dispatch. `name` stays as the SYS-YYYY-##### series ID.
    `poid` field holds the business POID; `system_id` mirrors `name` for legacy compatibility.
    """
    poid = (poid or "").strip()
    if poid:
        dispatch_doc.poid = poid
    dispatch_doc.insert(ignore_permissions=True)
    frappe.db.set_value("PO Dispatch", dispatch_doc.name, "system_id", dispatch_doc.name, update_modified=False)
    return dispatch_doc.name


def _upsert_po_dispatch_for_line(
    po_intake_name,
    po_no,
    line_dict,
    *,
    customer=None,
    im=None,
    target_month=None,
    planning_mode="Plan",
    dispatch_status="Pending",
    dispatch_mode="Manual",
):
    """Create or update PO Dispatch for a line while preserving immutable system_id."""
    po_line_no = cint(line_dict.get("po_line_no") or 0)
    project_code = line_dict.get("project_code")
    center_area = line_dict.get("center_area") or line_dict.get("area")
    site_code = (line_dict.get("site_code") or "").strip()
    site_name = line_dict.get("site_name")
    ok_duid, err_duid = ensure_duid_master(site_code, site_name, center_area)
    if not ok_duid:
        frappe.throw(err_duid or f"Could not create DUID Master for {site_code}")
    poid = _resolve_line_poid(
        po_no=po_no,
        po_line_no=po_line_no,
        shipment_number=line_dict.get("shipment_number"),
        fallback=line_dict.get("poid"),
    )
    payload = {
        "poid": poid,
        "po_intake": po_intake_name,
        "po_no": po_no,
        "po_line_no": po_line_no,
        "item_code": line_dict.get("item_code"),
        "item_description": line_dict.get("item_description"),
        "qty": flt(line_dict.get("qty", 0)),
        "rate": flt(line_dict.get("rate", 0)),
        "line_amount": flt(line_dict.get("line_amount", 0)),
        # Copy payment_terms from the intake line so the PO Dispatch validate
        # hook can parse MS1/MS2 percentages without joining back to the parent.
        "payment_terms": line_dict.get("payment_terms") or "",
        # PIC list view surfaces tax_rate alongside the line amount.
        "tax_rate": line_dict.get("tax_rate") or "",
        "customer": customer,
        "im": im,
        "target_month": target_month,
        "planning_mode": planning_mode,
        "dispatch_status": dispatch_status,
        "dispatch_mode": dispatch_mode,
        "center_area": center_area,
        "region_type": region_type_from_center_area(center_area),
        "site_code": site_code,
        "site_name": site_name,
    }
    if project_code and frappe.db.exists("Project Control Center", project_code):
        payload["project_code"] = project_code

    existing_name = frappe.db.get_value(
        "PO Dispatch", {"po_intake": po_intake_name, "po_line_no": po_line_no}, "name"
    )
    if existing_name:
        frappe.db.set_value("PO Dispatch", existing_name, payload, update_modified=False)
        system_id = frappe.db.get_value("PO Dispatch", existing_name, "system_id")
        if not system_id:
            frappe.db.set_value("PO Dispatch", existing_name, "system_id", existing_name, update_modified=False)
        # Ensure poid field is populated for docs created before this field existed
        if not frappe.db.get_value("PO Dispatch", existing_name, "poid"):
            frappe.db.set_value("PO Dispatch", existing_name, "poid", poid, update_modified=False)
        return existing_name

    doc = frappe.new_doc("PO Dispatch")
    for key, value in payload.items():
        if value is not None and value != "":
            setattr(doc, key, value)
    return _insert_po_dispatch_with_poid(doc, poid)


def _try_auto_dispatch(doc_name, po_no, child_rows):
    """
    After PO Intake save, auto-dispatch lines whose project has an IM assigned.

    Field team is chosen later by the IM at rollout planning (not on PO Dispatch).

    child_rows: list of (child_doc_name, line_dict) tuples
    Returns count of auto-dispatched lines.
    """
    # Batch-fetch IM + customer for all distinct project codes up front — replaces
    # 2 queries per child row with 1 IN query, critical when a PO has 1000+ lines.
    project_codes = list({
        (line_dict.get("project_code") or "").strip()
        for _, line_dict in child_rows
        if line_dict.get("project_code")
    })
    proj_map = {}
    if project_codes:
        for p in frappe.get_all(
            "Project Control Center",
            filters={"name": ["in", project_codes]},
            fields=["name", "implementation_manager", "customer"],
            limit_page_length=len(project_codes) + 1,
        ):
            proj_map[p.name] = p

    count = 0
    for child_doc_name, line_dict in child_rows:
        project_code = line_dict.get("project_code")
        if not project_code:
            continue
        pdata = proj_map.get(project_code)
        if not pdata or not pdata.get("implementation_manager"):
            continue

        _upsert_po_dispatch_for_line(
            doc_name,
            po_no,
            line_dict,
            customer=pdata.get("customer"),
            im=pdata.get("implementation_manager"),
            target_month=None,
            planning_mode="Plan",
            dispatch_status="Dispatched",
            dispatch_mode="Auto",
        )

        frappe.db.set_value(
            "PO Intake Line",
            child_doc_name,
            {"po_line_status": "Dispatched", "dispatch_mode": "Auto"},
            update_modified=False,
        )
        count += 1

    return count


@frappe.whitelist()
def backfill_po_dispatch_id_to_poid(limit=500):
    """
    One-time maintenance:
    Rename PO Dispatch docs to the PO Intake Line POID when possible.
    Returns summary counters.
    """
    limit = cint(limit or 500)
    rows = frappe.get_all(
        "PO Dispatch",
        fields=["name", "po_intake", "po_line_no", "po_no"],
        order_by="modified desc",
        limit_page_length=limit,
    )

    updated = 0
    skipped = 0
    failed = []

    for d in rows:
        line = frappe.db.get_value(
            "PO Intake Line",
            {"parent": d.po_intake, "po_line_no": d.po_line_no},
            ["poid", "shipment_number"],
            as_dict=True,
        )
        poid = _resolve_line_poid(
            po_no=d.po_no,
            po_line_no=d.po_line_no,
            shipment_number=(line or {}).get("shipment_number"),
            fallback=(line or {}).get("poid"),
        )
        if not poid or d.poid == poid:
            skipped += 1
            continue
        try:
            frappe.db.set_value("PO Dispatch", d.name, "poid", poid, update_modified=False)
            if not frappe.db.get_value("PO Dispatch", d.name, "system_id"):
                frappe.db.set_value("PO Dispatch", d.name, "system_id", d.name, update_modified=False)
            updated += 1
        except Exception:
            failed.append({"name": d.name, "poid": poid})

    frappe.db.commit()
    return {"checked": len(rows), "updated": updated, "skipped": skipped, "failed": failed}


# ---------------------------------------------------------------------------
# Task 12 — Pipeline Operations
# ---------------------------------------------------------------------------


@frappe.whitelist()
def export_po_dump(from_date=None, to_date=None, unique_inet_uid=1, statuses=None, limit=20000):
    """
    Export PO Intake lines whose parent PO was created in the date range (upload date).
    Returns uploaded PO lines in source column order for audit/export.

    ``statuses`` is a JSON list / comma string of buckets to include — any of
    ``OPEN`` (New / Dispatched / Completed), ``CLOSED``, ``CANCELLED``. Defaults
    to OPEN only so dumps stay small after archive imports.
    """
    fd = getdate(from_date) if from_date else add_days(getdate(), -30)
    td = getdate(to_date) if to_date else getdate()
    if td < fd:
        frappe.throw("to_date must be on or after from_date")

    if isinstance(statuses, str):
        try:
            parsed = frappe.parse_json(statuses)
            if isinstance(parsed, (list, tuple)):
                statuses = parsed
        except Exception:
            statuses = [s.strip() for s in statuses.split(",") if s.strip()]
    if not statuses:
        statuses = ["OPEN"]
    bucket_set = {str(s).strip().upper() for s in statuses}
    line_status_filter = []
    if "OPEN" in bucket_set:
        line_status_filter += ["New", "Dispatched", "Completed", ""]
    if "CLOSED" in bucket_set:
        line_status_filter.append("Closed")
    if "CANCELLED" in bucket_set or "CANCELED" in bucket_set:
        line_status_filter.append("Cancelled")

    parents = frappe.db.sql(
        """
        SELECT name, po_no, status, DATE(creation) AS upload_date
        FROM `tabPO Intake`
        WHERE DATE(creation) BETWEEN %s AND %s
        ORDER BY creation DESC
        """,
        (fd, td),
        as_dict=True,
    )
    parent_map = {p.name: p for p in parents}
    parent_names = list(parent_map.keys())
    rows_out = []

    # Total counts per bucket for the active date range — independent of which
    # buckets the caller is fetching, so the FE can always show full counts.
    bucket_totals = {"open": 0, "closed": 0, "cancelled": 0}
    if parent_names:
        ph_p = ", ".join(["%s"] * len(parent_names))
        agg = frappe.db.sql(
            f"""
            SELECT
              SUM(CASE WHEN IFNULL(po_line_status, 'New') IN ('New','Dispatched','Completed','') THEN 1 ELSE 0 END) AS open_cnt,
              SUM(CASE WHEN po_line_status = 'Closed' THEN 1 ELSE 0 END) AS closed_cnt,
              SUM(CASE WHEN po_line_status = 'Cancelled' THEN 1 ELSE 0 END) AS cancelled_cnt
            FROM `tabPO Intake Line`
            WHERE parent IN ({ph_p})
            """,
            tuple(parent_names),
            as_dict=True,
        )
        if agg:
            bucket_totals["open"] = cint(agg[0].open_cnt or 0)
            bucket_totals["closed"] = cint(agg[0].closed_cnt or 0)
            bucket_totals["cancelled"] = cint(agg[0].cancelled_cnt or 0)

    if not parent_names:
        return {
            "from_date": str(fd),
            "to_date": str(td),
            "rows": rows_out,
            "totals": bucket_totals,
        }

    fields = [
        "name",
        "parent",
        "source_id",
        "poid",
        "po_line_no",
        "shipment_number",
        "site_name",
        "site_code",
        "project_code",
        "project_name",
        "item_code",
        "item_description",
        "uom",
        "qty",
        "due_qty",
        "billed_quantity",
        "quantity_cancel",
        "start_date",
        "end_date",
        "sub_contract_no",
        "currency",
        "rate",
        "line_amount",
        "tax_rate",
        "payment_terms",
        "center_area",
        "publish_date",
        "po_line_status",
    ]

    # Allow up to 100k rows for archive dumps. Honor any positive limit (down
    # to 1) from the FE row-limit selector — previously we floored at 100,
    # which made "20" and "100" return the same result.
    lim = min(max(int(limit or 100000), 1), 100000)
    line_filters = {"parent": ["in", parent_names]}
    if line_status_filter:
        line_filters["po_line_status"] = ["in", list({s for s in line_status_filter})]
    lines = frappe.get_all(
        "PO Intake Line",
        filters=line_filters,
        fields=fields,
        order_by="parent desc, idx asc",
        limit_page_length=lim,
    )
    for ln in lines:
        par = parent_map.get(ln.parent) or {}
        # The PO Dump column reads from the per-line status (po_line_status)
        # so each POID shows its own state. Treat New / Dispatched / Completed
        # as Open for filter / display roll-ups.
        line_status = (ln.get("po_line_status") or "").strip()
        if not line_status:
            line_status = (par.get("status") or "OPEN").strip()
        upper = line_status.upper()
        if upper in ("NEW", "DISPATCHED", "COMPLETED", "OPEN"):
            display_status = "OPEN"
        elif upper in ("CLOSED",):
            display_status = "CLOSED"
        elif upper in ("CANCELLED", "CANCELED"):
            display_status = "CANCELLED"
        else:
            display_status = upper or "OPEN"
        rows_out.append(
            {
                "id": ln.get("source_id") or "",
                "po_status": display_status,
                "po_line_status": line_status,
                "po_no": par.get("po_no") or "",
                "po_line_no": ln.get("po_line_no"),
                "shipment_no": ln.get("shipment_number") or "",
                "site_name": ln.get("site_name") or "",
                "site_code": ln.get("site_code") or "",
                "item_code": ln.get("item_code") or "",
                "item_description": ln.get("item_description") or "",
                "unit": ln.get("uom") or "",
                "requested_qty": ln.get("qty"),
                "due_qty": ln.get("due_qty"),
                "billed_quantity": ln.get("billed_quantity"),
                "quantity_cancel": ln.get("quantity_cancel"),
                "start_date": ln.get("start_date"),
                "end_date": ln.get("end_date"),
                "sub_contract_no": ln.get("sub_contract_no") or "",
                "currency": ln.get("currency") or "",
                "unit_price": ln.get("rate"),
                "line_amount": ln.get("line_amount"),
                "tax_rate": ln.get("tax_rate") or "",
                "payment_terms": ln.get("payment_terms") or "",
                "poid": ln.get("poid") or "",
                "project_code": ln.get("project_code") or "",
                "project_name": ln.get("project_name") or "",
                "center_area": ln.get("center_area") or "",
                "publish_date": ln.get("publish_date"),
                "upload_date": str(par.get("upload_date") or ""),
            }
        )

    return {
        "from_date": str(fd),
        "to_date": str(td),
        "rows": rows_out,
        "totals": bucket_totals,
    }


@frappe.whitelist()
def upload_po_file(file_url, customer=None):
    """
    Parse an uploaded Huawei PO export (.xlsx/.csv) and return validated / error rows.

    Returns
    -------
    {"valid_rows": [...], "error_rows": [...], "total": N}
    """
    # ---- Resolve the physical file path --------------------------------
    if file_url.startswith("/private/files/"):
        file_path = frappe.get_site_path("private", "files", file_url[len("/private/files/"):])
    elif file_url.startswith("/files/"):
        file_path = frappe.get_site_path("public", "files", file_url[len("/files/"):])
    else:
        # Fall back to Frappe File document
        try:
            file_doc = frappe.get_doc("File", {"file_url": file_url})
            file_path = file_doc.get_full_path()
        except Exception:
            frappe.throw(f"Cannot resolve file path for: {file_url}")

    # ---- Column alias map ----------------------------------------------
    ALIAS = {
        "ID": "source_id",
        "PO Status": "po_status",
        "PO STATUS": "po_status",
        "Status": "po_status",
        "PO status": "po_status",
        "PO NO.": "po_no",
        "PO Line NO.": "po_line_no",
        "Shipment NO.": "shipment_no",
        "Site Name": "site_name",
        "Site Code": "site_code",
        "Item Code": "item_code",
        "Item Description": "item_description",
        "Unit": "unit",
        "Requested Qty": "qty",
        "Due Qty": "due_qty",
        "Billed Quantity": "billed_quantity",
        "Quantity Cancel": "quantity_cancel",
        "Start Date": "start_date",
        "End Date": "end_date",
        "Sub Contract NO.": "sub_contract_no",
        "Currency": "currency",
        "Unit Price": "rate",
        "Line Amount": "line_amount",
        "Tax Rate": "tax_rate",
        "Payment Terms": "payment_terms",
        "Project Code": "project_code",
        "Project Name": "project_name",
        "Center Area": "center_area",
        "Publish Date": "publish_date",
    }

    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            csv_rows = list(csv.reader(f))
        if not csv_rows:
            frappe.throw("The uploaded file appears to be empty.")
        raw_headers = csv_rows[0]
        rows_iter = iter(csv_rows[1:])
    else:
        try:
            import openpyxl
        except ImportError:
            frappe.throw("openpyxl is not installed. Run `bench pip install openpyxl`.")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)

    # First row = headers
    if not raw_headers:
        frappe.throw("The uploaded file appears to be empty.")

    headers = [str(h).strip() if h is not None else "" for h in raw_headers]

    valid_rows = []
    error_rows = []
    cim_ensured_keys = set()

    for row_idx, raw_row in enumerate(rows_iter, start=2):
        row_dict = {}
        for col_idx, cell_val in enumerate(raw_row):
            raw_header = headers[col_idx] if col_idx < len(headers) else ""
            std_key = ALIAS.get(raw_header)
            if std_key:
                row_dict[std_key] = cell_val

        row_dict["po_status"] = _normalize_po_intake_status(row_dict.get("po_status"))

        # ---- Validation ------------------------------------------------
        errors = []
        # Standard upload is for active workflow lines only. Closed / Cancelled
        # lines must be backfilled through the Archive tab so they don't
        # auto-dispatch into the IM queue.
        if row_dict["po_status"] in ("CLOSED", "CANCELLED"):
            errors.append(
                f"PO Status is {row_dict['po_status']} — use the Archive tab to import closed / cancelled lines"
            )
        if not row_dict.get("po_no"):
            errors.append("po_no is required")
        # Accept rows where item_code is blank / "NA" if we can fall back to
        # the description (helper rewrites the code at confirm time).
        resolved_code, used_desc = _resolve_item_code_with_fallback(row_dict)
        if not resolved_code:
            errors.append("item_code is required (or provide an item_description fallback)")
        elif used_desc:
            row_dict["item_code"] = resolved_code
            if not str(row_dict.get("item_description") or "").strip():
                row_dict["item_description"] = resolved_code
        if not row_dict.get("project_code"):
            errors.append("project_code is required")
        qty = flt(row_dict.get("qty", 0))
        rate = flt(row_dict.get("rate", 0))
        if qty <= 0:
            errors.append("qty must be > 0")
        if rate <= 0:
            errors.append("rate must be > 0")

        project_code = str(row_dict.get("project_code") or "").strip()
        if project_code and not frappe.db.exists("Project Control Center", project_code):
            cust_for_proj = (customer or "").strip() or str(row_dict.get("customer") or "").strip()
            if cust_for_proj:
                ok_p, err_p = ensure_project_control_center(
                    project_code,
                    cust_for_proj,
                    row_dict.get("project_name"),
                    row_dict.get("center_area"),
                )
                if not ok_p:
                    errors.append(err_p or "Could not create Project Control Center")
            else:
                errors.append(
                    "project_code not found in Project Control Center (select customer on upload to auto-create project)"
                )

        item_code = str(row_dict.get("item_code") or "").strip()
        if item_code:
            ok_i, err_i = ensure_item_master(item_code, row_dict.get("item_description"))
            if not ok_i:
                errors.append(err_i or "Could not create Item")
        if customer and item_code:
            ck = (customer.strip(), item_code)
            if ck not in cim_ensured_keys:
                ok_cim, cim_err = ensure_customer_item_master(customer, item_code)
                if ok_cim:
                    cim_ensured_keys.add(ck)
                else:
                    errors.append(cim_err or "Could not create Customer Item Master")
        elif item_code:
            cim_filters = {"item_code": item_code, "active_flag": 1}
            if not frappe.db.exists("Customer Item Master", cim_filters):
                errors.append(
                    "item_code not found in Customer Item Master (pick a customer in upload to auto-create per customer)"
                )

        if errors:
            row_dict["_row"] = row_idx
            row_dict["_errors"] = errors
            error_rows.append(row_dict)
        else:
            row_dict["qty"] = qty
            row_dict["rate"] = rate
            line_amount = flt(row_dict.get("line_amount", 0))
            row_dict["line_amount"] = line_amount if line_amount > 0 else (qty * rate)
            row_dict["item_exists"] = True
            row_dict["project_exists"] = True
            valid_rows.append(row_dict)

    return {
        "valid_rows": valid_rows,
        "error_rows": error_rows,
        "total": len(valid_rows) + len(error_rows),
    }


_ITEM_CODE_PLACEHOLDERS = {"", "NA", "N/A", "-", "N.A.", "NULL", "NONE"}


def _resolve_item_code_with_fallback(line):
    """Return (item_code, used_description_flag).

    Some Master Tracker rows leave Item Code blank or filled in with ``NA``.
    Standard PO uploads can't insert a PO Intake Line with a blank Item Code
    (it's a mandatory link to ``Item``), so when the value is missing or a
    placeholder, fall back to using the Item Description as the code (truncated
    to fit Frappe's Data length limit). The matching ``Item`` will be created
    by ``ensure_item_master`` further down the line.
    """
    raw = (line.get("item_code") or "")
    code = str(raw).strip()
    if code.upper() in _ITEM_CODE_PLACEHOLDERS:
        desc = str(line.get("item_description") or "").strip()
        if desc:
            return desc[:140].strip(), True
        # No description either — caller will need to skip this row.
        return "", True
    return code, False


def _poid_for_upload_line(po_no, line):
    """Stable POID for de-duplication (matches PO Intake Line / dispatch conventions)."""
    return _make_poid(po_no, line.get("po_line_no"), line.get("shipment_no"))


def _existing_poids_on_intake_doc(doc):
    seen = set()
    for row in doc.po_lines:
        pid = (row.poid or "").strip()
        if not pid:
            pid = _make_poid(doc.po_no, row.po_line_no, row.shipment_number)
        seen.add(pid)
    return seen


@frappe.whitelist()
def confirm_po_upload(rows):
    """
    Take validated rows (JSON string or list), group by po_no, and create or extend PO Intake.

    New PO numbers create a PO Intake document. If the PO already exists, new lines are appended
    when their POID is not already on the document (same PO No + line + shipment).

    Returns
    -------
    {
        "created": new PO Intake documents created,
        "lines_imported": child lines inserted,
        "lines_skipped_duplicate": rows skipped because POID already exists,
        "names": names of newly created PO Intake docs (append-only POs omitted),
    }
    """
    rows = _parse_rows(rows)

    # Group by po_no (normalize key for DB lookup)
    po_groups = {}
    for row in rows:
        po_no = row.get("po_no")
        if po_no is None or str(po_no).strip() == "":
            continue
        key = str(po_no).strip()
        po_groups.setdefault(key, []).append(row)

    created = 0
    names = []
    lines_imported = 0
    lines_skipped_duplicate = 0
    lines_skipped_terminal = 0   # = closed + cancelled (kept for backward-compat)
    lines_skipped_closed = 0
    lines_skipped_cancelled = 0
    terminal_dupe_samples = []  # first ~20 (poid, existing_status) pairs for the FE summary
    item_code_from_desc = 0  # rows where item_code was blank / "NA" — fell back to item_description
    auto_dispatched = 0
    po_summary = []  # per-PO breakdown for audit/UI: [{po_no, intake_name, lines_added, lines_skipped, is_new}]

    for po_no, lines in po_groups.items():
        first = lines[0]

        # Resolve customer (same rules for new + existing intake)
        row_customer = first.get("customer")
        doc_customer = row_customer
        if not doc_customer:
            project_code_hdr = first.get("project_code")
            if project_code_hdr:
                doc_customer = frappe.db.get_value(
                    "Project Control Center", project_code_hdr, "customer"
                )
        if not doc_customer:
            frappe.throw(f"Customer is required for PO {po_no}")
        resolved_cust = _resolve_customer_link_name(doc_customer)
        if not resolved_cust:
            frappe.throw(f"Customer does not exist: {doc_customer}")

        existing_name = frappe.db.get_value("PO Intake", {"po_no": po_no}, "name")
        existing_poids = set()
        existing_status_by_poid = {}
        if existing_name:
            # Fetch fields for POID dedup + the existing line's status so we can
            # tell the user which duplicates are sitting under terminal states.
            existing_rows = frappe.db.sql(
                "SELECT po_line_no, shipment_number, poid, po_line_status "
                "FROM `tabPO Intake Line` WHERE parent=%s",
                existing_name, as_dict=True,
            )
            for r in existing_rows:
                pid = (r.poid or "").strip() or _make_poid(po_no, r.po_line_no, r.shipment_number)
                existing_poids.add(pid)
                existing_status_by_poid[pid] = (r.po_line_status or "").strip()
            existing_customer = frappe.db.get_value("PO Intake", existing_name, "customer")
            if existing_customer != resolved_cust:
                frappe.throw(
                    f"PO {po_no} already exists for customer {existing_customer}; "
                    f"upload uses {resolved_cust}. Use the same customer or a different PO number."
                )

        # Bulk pre-fetch project codes to avoid a frappe.db.exists() call per line
        all_project_codes = {
            str(l.get("project_code") or "").strip() for l in lines if l.get("project_code")
        }
        if all_project_codes:
            _ph = ",".join(["%s"] * len(all_project_codes))
            existing_projects = {
                r[0]
                for r in frappe.db.sql(
                    f"SELECT name FROM `tabProject Control Center` WHERE name IN ({_ph})",
                    list(all_project_codes),
                )
            }
        else:
            existing_projects = set()

        cim_pairs_done = set()
        duid_done = set()  # site_codes whose DUID master was already ensured this batch
        new_entries = []  # (source_line_dict, append_row_dict)
        po_skipped = 0

        for line in lines:
            poid = _poid_for_upload_line(po_no, line)
            if poid in existing_poids:
                lines_skipped_duplicate += 1
                po_skipped += 1
                existing_status = existing_status_by_poid.get(poid, "")
                if existing_status == "Closed":
                    lines_skipped_terminal += 1
                    lines_skipped_closed += 1
                elif existing_status == "Cancelled":
                    lines_skipped_terminal += 1
                    lines_skipped_cancelled += 1
                if existing_status in ("Closed", "Cancelled") and len(terminal_dupe_samples) < 20:
                    terminal_dupe_samples.append({
                        "poid": poid,
                        "existing_status": existing_status,
                        "po_no": po_no,
                    })
                continue

            item_code, used_desc = _resolve_item_code_with_fallback(line)
            if used_desc and item_code:
                # Stamp the resolved code back on the source line so subsequent
                # lookups (CIM, dispatch upsert) all see the same value.
                line["item_code"] = item_code
                if not str(line.get("item_description") or "").strip():
                    line["item_description"] = item_code
                item_code_from_desc += 1
            if not item_code:
                # Truly empty (no description either) — skip rather than fail.
                lines_skipped_duplicate += 1
                po_skipped += 1
                continue
            qty = flt(line.get("qty", 0))
            rate = flt(line.get("rate", 0))
            if qty <= 0 or rate <= 0:
                frappe.throw(f"Invalid qty/rate in PO {po_no} for item {item_code}")
            project_code = str(line.get("project_code") or "").strip()
            if project_code and project_code not in existing_projects:
                ok_p, err_p = ensure_project_control_center(
                    project_code,
                    resolved_cust,
                    line.get("project_name") or first.get("project_name"),
                    line.get("center_area") or first.get("center_area"),
                )
                if not ok_p:
                    frappe.throw(err_p or f"Project code not found: {project_code}")
                existing_projects.add(project_code)

            cim_key = (resolved_cust, item_code)
            if cim_key not in cim_pairs_done:
                ok_cim, cim_err = ensure_customer_item_master(resolved_cust, item_code)
                if not ok_cim:
                    frappe.throw(
                        cim_err
                        or f"Customer Item Master could not be created for customer {resolved_cust} and item {item_code}"
                    )
                cim_pairs_done.add(cim_key)

            line_center_area = line.get("center_area") or first.get("center_area")
            line_region = region_type_from_center_area(line_center_area)
            site_code = str(line.get("site_code") or "").strip()
            if site_code not in duid_done:
                ok_duid, err_duid = ensure_duid_master(
                    site_code,
                    line.get("site_name"),
                    line_center_area,
                )
                if not ok_duid:
                    frappe.throw(err_duid or f"Could not create DUID Master for {site_code}")
                duid_done.add(site_code)

            append_row = {
                "source_id": line.get("source_id"),
                "po_line_no": cint(line.get("po_line_no") or 0),
                "shipment_number": line.get("shipment_no"),
                "poid": poid,
                "site_code": site_code,
                "site_name": line.get("site_name"),
                "item_code": item_code,
                "item_description": line.get("item_description"),
                "uom": line.get("unit"),
                "qty": qty,
                "due_qty": flt(line.get("due_qty", 0)),
                "billed_quantity": flt(line.get("billed_quantity", 0)),
                "quantity_cancel": flt(line.get("quantity_cancel", 0)),
                "start_date": line.get("start_date"),
                "end_date": line.get("end_date"),
                "sub_contract_no": line.get("sub_contract_no"),
                "currency": line.get("currency"),
                "rate": rate,
                "line_amount": flt(line.get("line_amount", 0)) or (qty * rate),
                "tax_rate": line.get("tax_rate"),
                "payment_terms": line.get("payment_terms"),
                "project_code": line.get("project_code"),
                "project_name": line.get("project_name"),
                "center_area": line_center_area,
                "region_type": line_region,
                "publish_date": line.get("publish_date"),
            }
            new_entries.append((line, append_row))
            existing_poids.add(poid)

        if not new_entries:
            # Every line was a duplicate — still record the PO for the audit summary
            if po_skipped > 0:
                po_summary.append({
                    "po_no": po_no,
                    "intake_name": existing_name,
                    "lines_added": 0,
                    "lines_skipped": po_skipped,
                    "is_new": False,
                })
            continue

        hdr_status = first.get("po_status") or first.get("status") or first.get("po_intake_status")

        if existing_name:
            doc = frappe.get_doc("PO Intake", existing_name)
            doc.status = _normalize_po_intake_status(hdr_status)
            for _src, append_row in new_entries:
                doc.append("po_lines", append_row)
            # Defence in depth: a few legacy rows leak into valid_rows with a
            # blank item_code (e.g. when the source cell is whitespace). Don't
            # blow up the whole upload — the helper above already filled in a
            # description-derived code where possible.
            doc.flags.ignore_mandatory = True
            doc.save(ignore_permissions=True)
            is_new_po = False
        else:
            doc = frappe.new_doc("PO Intake")
            doc.po_no = po_no
            doc.customer = resolved_cust
            doc.status = _normalize_po_intake_status(hdr_status)
            doc.publish_date = first.get("publish_date")
            doc.center_area = first.get("center_area")
            for _src, append_row in new_entries:
                doc.append("po_lines", append_row)
            doc.flags.ignore_mandatory = True
            doc.insert(ignore_permissions=True)
            created += 1
            names.append(doc.name)
            is_new_po = True

        frappe.db.commit()
        lines_imported += len(new_entries)
        po_summary.append({
            "po_no": po_no,
            "intake_name": doc.name,
            "lines_added": len(new_entries),
            "lines_skipped": po_skipped,
            "is_new": is_new_po,
        })

        # Resolve DB names for new child rows via SQL — avoids doc.reload() + O(n²) scan
        new_poid_map = {append_row["poid"]: append_row for _src, append_row in new_entries}
        child_sql_rows = frappe.db.sql(
            "SELECT name, po_line_no, shipment_number, poid FROM `tabPO Intake Line` WHERE parent=%s",
            doc.name, as_dict=True,
        )
        child_rows = []
        for r in child_sql_rows:
            got_poid = (r.poid or "").strip() or _make_poid(po_no, r.po_line_no, r.shipment_number)
            if got_poid in new_poid_map:
                row_dict = dict(new_poid_map[got_poid])
                row_dict["name"] = r.name
                row_dict["parent"] = doc.name
                child_rows.append((r.name, row_dict))

        # Ensure each new line has a PO Dispatch row + immutable system_id even before dispatch.
        for _child_name, line_dict in child_rows:
            _upsert_po_dispatch_for_line(
                doc.name,
                po_no,
                line_dict,
                customer=resolved_cust,
                dispatch_status="Pending",
                dispatch_mode="Manual",
            )
        auto_dispatched += _try_auto_dispatch(doc.name, po_no, child_rows)

    return {
        "created": created,
        "lines_imported": lines_imported,
        "lines_skipped_duplicate": lines_skipped_duplicate,
        "lines_skipped_terminal": lines_skipped_terminal,
        "lines_skipped_closed": lines_skipped_closed,
        "lines_skipped_cancelled": lines_skipped_cancelled,
        "terminal_dupe_samples": terminal_dupe_samples,
        "item_code_from_desc": item_code_from_desc,
        "auto_dispatched": auto_dispatched,
        "names": names,
        "po_summary": po_summary,
    }


@frappe.whitelist()
def record_po_upload_log(payload):
    """Persist a PO Upload Log for audit/history. Called after all chunks complete.

    payload = {
        "file_name": "...", "file_url": "...", "customer": "...",
        "total_rows": N, "lines_imported": N, "lines_skipped": N,
        "po_created": N, "po_updated": N, "auto_dispatched": N,
        "po_summary": [{po_no, intake_name, lines_added, lines_skipped, is_new/status}, ...],
        "status": "Completed"|"Partial"|"Failed",
        "notes": "...",
    }
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)
    payload = payload or {}

    doc = frappe.new_doc("PO Upload Log")
    doc.uploaded_by = frappe.session.user
    doc.uploaded_at = frappe.utils.now_datetime()
    doc.file_name = (payload.get("file_name") or "")[:140]
    doc.file_url = (payload.get("file_url") or "")[:240]
    doc.customer = payload.get("customer") or None
    doc.total_rows = cint(payload.get("total_rows") or 0)
    doc.lines_imported = cint(payload.get("lines_imported") or 0)
    doc.lines_skipped = cint(payload.get("lines_skipped") or 0)
    doc.lines_skipped_terminal = cint(payload.get("lines_skipped_terminal") or 0)
    doc.lines_skipped_closed = cint(payload.get("lines_skipped_closed") or 0)
    doc.lines_skipped_cancelled = cint(payload.get("lines_skipped_cancelled") or 0)
    samples = payload.get("terminal_dupe_samples")
    if samples:
        try:
            doc.terminal_dupe_samples = frappe.as_json(samples)[:8000]
        except Exception:
            doc.terminal_dupe_samples = None
    doc.po_created = cint(payload.get("po_created") or 0)
    doc.po_updated = cint(payload.get("po_updated") or 0)
    doc.auto_dispatched = cint(payload.get("auto_dispatched") or 0)
    doc.status = payload.get("status") or "Completed"
    doc.notes = (payload.get("notes") or "")[:2000]

    for item in (payload.get("po_summary") or []):
        if not isinstance(item, dict):
            continue
        status = item.get("status")
        if not status:
            if item.get("is_new"):
                status = "New"
            elif cint(item.get("lines_added") or 0) > 0:
                status = "Appended"
            else:
                status = "Duplicate"
        doc.append("po_details", {
            "po_no": (item.get("po_no") or "")[:140],
            "intake_name": item.get("intake_name") or None,
            "status": status,
            "lines_added": cint(item.get("lines_added") or 0),
            "lines_skipped": cint(item.get("lines_skipped") or 0),
        })

    doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": doc.name}


@frappe.whitelist()
def list_po_upload_logs(limit=50):
    """Return recent PO Upload Log entries with per-PO counts."""
    limit = cint(limit or 50)
    if limit <= 0 or limit > 500:
        limit = 50
    rows = frappe.get_all(
        "PO Upload Log",
        fields=[
            "name", "uploaded_by", "uploaded_at", "customer", "file_name",
            "total_rows", "lines_imported", "lines_skipped", "lines_skipped_terminal",
            "lines_skipped_closed", "lines_skipped_cancelled",
            "po_created", "po_updated", "auto_dispatched", "status",
        ],
        order_by="uploaded_at desc",
        limit_page_length=limit,
    )
    # Attach po count per log (quick summary; full details on demand)
    log_names = [r.name for r in rows]
    counts = {}
    if log_names:
        placeholders = ",".join(["%s"] * len(log_names))
        count_rows = frappe.db.sql(
            f"SELECT parent, COUNT(*) AS c FROM `tabPO Upload Log Detail` "
            f"WHERE parent IN ({placeholders}) GROUP BY parent",
            log_names,
            as_dict=True,
        )
        counts = {r.parent: r.c for r in count_rows}
    for r in rows:
        r["po_count"] = counts.get(r.name, 0)
    return rows


@frappe.whitelist()
def get_po_upload_log(name):
    """Return a PO Upload Log with its per-PO detail rows."""
    if not name:
        frappe.throw("name is required")
    if not frappe.db.exists("PO Upload Log", name):
        frappe.throw(f"PO Upload Log not found: {name}")
    doc = frappe.get_doc("PO Upload Log", name)
    samples = []
    if getattr(doc, "terminal_dupe_samples", None):
        try:
            parsed = frappe.parse_json(doc.terminal_dupe_samples)
            if isinstance(parsed, list):
                samples = parsed
        except Exception:
            samples = []
    return {
        "name": doc.name,
        "uploaded_by": doc.uploaded_by,
        "uploaded_at": doc.uploaded_at,
        "customer": doc.customer,
        "file_name": doc.file_name,
        "file_url": doc.file_url,
        "status": doc.status,
        "total_rows": doc.total_rows,
        "lines_imported": doc.lines_imported,
        "lines_skipped": doc.lines_skipped,
        "lines_skipped_terminal": getattr(doc, "lines_skipped_terminal", 0) or 0,
        "lines_skipped_closed": getattr(doc, "lines_skipped_closed", 0) or 0,
        "lines_skipped_cancelled": getattr(doc, "lines_skipped_cancelled", 0) or 0,
        "terminal_dupe_samples": samples,
        "po_created": doc.po_created,
        "po_updated": doc.po_updated,
        "auto_dispatched": doc.auto_dispatched,
        "notes": doc.notes,
        "po_details": [
            {
                "po_no": d.po_no,
                "intake_name": d.intake_name,
                "status": d.status,
                "lines_added": d.lines_added,
                "lines_skipped": d.lines_skipped,
            }
            for d in (doc.po_details or [])
        ],
    }


def _get_dispatch_for_intake_line(parent_name, po_line_no):
    """Load dispatch row for a PO Intake line; tolerate DB missing `region_type` before migrate."""
    filters = {"po_intake": parent_name, "po_line_no": po_line_no}
    fields_full = [
        "name", "system_id", "im", "dispatch_mode", "target_month", "region_type", "center_area",
    ]
    try:
        return frappe.db.get_value("PO Dispatch", filters, fields_full, as_dict=True) or {}
    except frappe.db.OperationalError as e:
        if not frappe.db.is_missing_column(e):
            raise
        row = frappe.db.get_value(
            "PO Dispatch",
            filters,
            ["name", "system_id", "im", "dispatch_mode", "target_month", "center_area"],
            as_dict=True,
        ) or {}
        row["region_type"] = region_type_from_center_area(row.get("center_area"))
        return row


@frappe.whitelist()
def list_po_intake_lines(status="New", limit=None, portal_filters=None):
    """
    Return PO Intake child lines that match given po_line_status (or all when status='all').
    Each row is enriched with parent PO Intake fields and, for dispatched lines, dispatch info.

    ``portal_filters`` (JSON dict) may include: search, project_code, site_code, dispatched_im (or im),
    from_date, to_date (dispatch ``target_month`` when a dispatch row exists),
    intake_tab (e.g. ``New`` — skips IM-on-dispatch filter so undispatched lines are not dropped).
    """
    filters = {}
    if status and status.lower() != "all":
        filters["po_line_status"] = status

    limit_page_length = _portal_row_limit(limit, 500)
    pf = _portal_filters_dict(portal_filters)

    def _portal_active():
        if not pf:
            return False
        for k in ("search", "q", "project_code", "site_code", "dispatched_im", "im", "from_date", "to_date"):
            if (pf.get(k) or "").strip() if isinstance(pf.get(k), str) else pf.get(k):
                return True
        return False

    line_fields_full = [
        "name", "parent", "po_line_no", "poid", "shipment_number",
        "item_code", "item_description", "qty", "rate", "line_amount",
        "project_code", "site_code", "site_name", "area", "center_area", "region_type",
        "po_line_status", "activity_code", "dispatch_mode",
    ]
    line_fields_base = [
        "name", "parent", "po_line_no", "poid", "shipment_number",
        "item_code", "item_description", "qty", "rate", "line_amount",
        "project_code", "site_code", "site_name", "area",
        "po_line_status", "activity_code", "dispatch_mode",
    ]

    lines = []
    if _portal_active():
        wheres = ["1=1"]
        params = []
        if filters.get("po_line_status"):
            wheres.append("pil.po_line_status = %s")
            params.append(filters["po_line_status"])
        for col, key in (("IFNULL(pil.project_code,'')", "project_code"),
                         ("IFNULL(pil.site_code,'')", "site_code")):
            c, p = _sql_in_or_eq(col, pf.get(key))
            if c:
                wheres.append(c)
                params.extend(p)
        im_vals = _ensure_list(pf.get("dispatched_im") or pf.get("im"))
        intake_tab = (pf.get("intake_tab") or "").strip().lower()
        if im_vals and intake_tab != "new":
            c, p = _sql_in_or_eq("IFNULL(pd.im,'')", im_vals)
            wheres.append(c)
            params.extend(p)
        if pf.get("from_date") and pf.get("to_date"):
            wheres.append(
                "(pd.name IS NULL OR (pd.target_month BETWEEN %s AND %s))"
            )
            params.extend([pf["from_date"], pf["to_date"]])
        elif pf.get("from_date"):
            wheres.append("(pd.name IS NULL OR pd.target_month >= %s)")
            params.append(pf["from_date"])
        elif pf.get("to_date"):
            wheres.append("(pd.name IS NULL OR pd.target_month <= %s)")
            params.append(pf["to_date"])
        concat_expr_intake = (
            "CONCAT_WS(' ', IFNULL(pil.name,''), IFNULL(pil.poid,''), IFNULL(pil.item_code,''), "
            "IFNULL(pil.project_code,''), IFNULL(pil.site_code,''), IFNULL(pi.po_no,''), "
            "IFNULL(pi.customer,''), IFNULL(pil.center_area,''), IFNULL(pil.region_type,''))"
        )
        clause, cparams = _sql_search_clause(concat_expr_intake, pf.get("search") or pf.get("q") or "")
        if clause:
            wheres.append(clause)
            params.extend(cparams)
        id_sql = (
            "SELECT pil.name AS line_id "
            "FROM `tabPO Intake Line` pil "
            "INNER JOIN `tabPO Intake` pi ON pi.name = pil.parent "
            "LEFT JOIN `tabPO Dispatch` pd ON pd.po_intake = pil.parent AND pd.po_line_no = pil.po_line_no "
            f"WHERE {' AND '.join(wheres)} "
            "ORDER BY pil.parent DESC, pil.idx ASC "
            f"{_sql_limit_suffix(limit_page_length)}"
        )
        id_rows = frappe.db.sql(id_sql, tuple(params), as_dict=True)
        line_ids = [r.line_id for r in (id_rows or []) if r.get("line_id")]
        if not line_ids:
            lines = []
        else:
            order_index = {n: i for i, n in enumerate(line_ids)}
            try:
                lines = frappe.get_all(
                    "PO Intake Line",
                    filters={"name": ["in", line_ids]},
                    fields=line_fields_full,
                    limit_page_length=len(line_ids) + 1,
                )
            except frappe.db.OperationalError as e:
                if not frappe.db.is_missing_column(e):
                    raise
                lines = frappe.get_all(
                    "PO Intake Line",
                    filters={"name": ["in", line_ids]},
                    fields=line_fields_base,
                    limit_page_length=len(line_ids) + 1,
                )
                for row in lines:
                    row.setdefault("center_area", None)
                    row.setdefault("region_type", None)
            lines.sort(key=lambda r: order_index.get(r.name, 10**9))
    else:
        ga_kwargs = dict(
            filters=filters,
            order_by="parent desc, idx asc",
        )
        if limit_page_length:
            ga_kwargs["limit_page_length"] = limit_page_length
        try:
            lines = frappe.get_all("PO Intake Line", fields=line_fields_full, **ga_kwargs)
        except frappe.db.OperationalError as e:
            if not frappe.db.is_missing_column(e):
                raise
            lines = frappe.get_all("PO Intake Line", fields=line_fields_base, **ga_kwargs)
            for row in lines:
                row.setdefault("center_area", None)
                row.setdefault("region_type", None)

    # Batch-enrich: two bulk fetches replace 1 query per parent + 1 query per line.
    # For 4k+ rows this collapses ~4000 round-trips into 2, cutting load from ~40s
    # to sub-second on realistic sizes.
    parent_names = list({line.get("parent") for line in lines if line.get("parent")})
    parent_map = {}
    if parent_names:
        for p in frappe.get_all(
            "PO Intake",
            filters={"name": ["in", parent_names]},
            fields=["name", "po_no", "customer", "center_area"],
            limit_page_length=len(parent_names) + 1,
        ):
            parent_map[p.name] = p

    dispatch_map = {}
    if parent_names:
        disp_fields_full = [
            "name", "po_intake", "po_line_no", "system_id", "im",
            "dispatch_mode", "target_month", "region_type", "center_area",
        ]
        disp_fields_base = [
            "name", "po_intake", "po_line_no", "system_id", "im",
            "dispatch_mode", "target_month", "center_area",
        ]
        try:
            all_disp = frappe.get_all(
                "PO Dispatch",
                filters={"po_intake": ["in", parent_names]},
                fields=disp_fields_full,
                limit_page_length=len(parent_names) * 100 + 1,
            )
        except frappe.db.OperationalError as e:
            if not frappe.db.is_missing_column(e):
                raise
            all_disp = frappe.get_all(
                "PO Dispatch",
                filters={"po_intake": ["in", parent_names]},
                fields=disp_fields_base,
                limit_page_length=len(parent_names) * 100 + 1,
            )
            for d in all_disp:
                d["region_type"] = region_type_from_center_area(d.get("center_area"))
        for d in all_disp:
            dispatch_map[(d.po_intake, cint(d.po_line_no))] = d

    for line in lines:
        parent_name = line.get("parent")
        pdata = parent_map.get(parent_name) or {}
        line["po_no"] = pdata.get("po_no", "")
        line["customer"] = pdata.get("customer", "")
        line["po_intake"] = parent_name
        line["center_area"] = line.get("center_area") or pdata.get("center_area")

        dispatch_data = dispatch_map.get((parent_name, cint(line.get("po_line_no")))) or {}
        if dispatch_data:
            line["dispatch_name"] = dispatch_data.get("name")
            line["system_id"] = dispatch_data.get("system_id")
        if line.get("po_line_status") == "Dispatched" and dispatch_data:
            line["dispatched_im"] = dispatch_data.get("im")
            line["dispatch_target_month"] = dispatch_data.get("target_month")
            if not line.get("dispatch_mode"):
                line["dispatch_mode"] = dispatch_data.get("dispatch_mode")
            if dispatch_data.get("region_type"):
                line["region_type"] = dispatch_data.get("region_type")
            elif not line.get("region_type"):
                line["region_type"] = region_type_from_center_area(
                    dispatch_data.get("center_area") or line.get("center_area")
                )

        if not line.get("region_type"):
            line["region_type"] = region_type_from_center_area(line.get("center_area"))

    im_dispatched_ids = list(
        {
            line.get("dispatched_im")
            for line in lines
            if line.get("dispatched_im") and line.get("po_line_status") == "Dispatched"
        }
    )
    im_fn_map = _batch_im_master_full_names(im_dispatched_ids)
    for line in lines:
        imn = line.get("dispatched_im")
        if imn:
            line["dispatched_im_full_name"] = im_fn_map.get(imn)

    return lines


def _require_inet_im_session():
    user = frappe.session.user
    if not user or user == "Guest":
        frappe.throw("Not permitted", frappe.PermissionError)
    roles = set(frappe.get_roles(user))
    if (
        "INET IM" not in roles
        and "Administrator" not in roles
        and "System Manager" not in roles
    ):
        frappe.throw(
            "Only Implementation Managers can use this action.",
            frappe.PermissionError,
        )
    _im_resolved, im_identifiers, _ = resolve_im_for_session()
    if not im_identifiers:
        frappe.throw(
            "Could not resolve IM from your user profile.",
            frappe.ValidationError,
        )
    return _im_resolved, im_identifiers


def _pcc_im_allows_project(project_code, im_identifiers):
    if not project_code or not frappe.db.exists("Project Control Center", project_code):
        return False
    im_on = frappe.db.get_value(
        "Project Control Center", project_code, "implementation_manager"
    )
    if not im_on:
        return False
    return im_on in set(im_identifiers)


@frappe.whitelist()
def create_im_dummy_po_dispatch(payload=None):
    """
    Dummy PO Dispatch: only project is required at create time.

    Placeholder DUID / item / qty are set; real details are applied in map_im_dummy_po_to_intake_line.

    payload: {"project_code": required, "target_month": optional "YYYY-MM" or
    "YYYY-MM-01". If provided, the dummy dispatch skips PO Intake and lands
    directly in My Dispatches under that month.}
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)
    payload = payload or {}

    _im_resolved, im_identifiers = _require_inet_im_session()

    project_code = (payload.get("project_code") or "").strip()
    if not project_code:
        frappe.throw("project_code is required")
    if not _pcc_im_allows_project(project_code, im_identifiers):
        frappe.throw("You are not the Implementation Manager for this project.")

    target_month = (payload.get("target_month") or "").strip()
    if target_month:
        # Accept "YYYY-MM" or a full date — always store as first-of-month.
        try:
            if len(target_month) == 7:
                target_month = f"{target_month}-01"
            target_month = str(getdate(target_month).replace(day=1))
        except Exception:
            frappe.throw("Invalid target_month (expected YYYY-MM or YYYY-MM-DD)")

    site_code = None
    for _attempt in range(40):
        candidate = f"DUMMY-{frappe.generate_hash(length=14)}"
        if not frappe.db.exists("DUID Master", candidate):
            site_code = candidate
            break
    if not site_code:
        frappe.throw("Could not allocate a placeholder DUID — retry.")

    site_name = "Pending — map to PO line"
    ok_duid, err_duid = ensure_duid_master(site_code, site_name=site_name, center_area=None)
    if not ok_duid:
        frappe.throw(err_duid or f"Could not create DUID Master for {site_code}")

    customer = frappe.db.get_value("Project Control Center", project_code, "customer")

    item_code = "PENDING"
    item_description = "Fill by mapping to PO Intake line"
    qty = flt(1)
    rate = flt(0)
    line_amount = flt(0)
    center_area = None

    # Placeholder PO until map: DUMMY-yymmdd-XXXXXX (6 hex) → short POID e.g. DUMMY-260416-A3B4C5-1
    dt = now_datetime()
    for _attempt in range(30):
        po_no = f"DUMMY-{dt.strftime('%y%m%d')}-{frappe.generate_hash(length=6).upper()}"
        if not frappe.db.exists("PO Dispatch", {"po_no": po_no}):
            break
    else:
        frappe.throw("Could not allocate a unique dummy PO number — retry.")

    poid = _make_poid(po_no, 1, None)

    doc = frappe.new_doc("PO Dispatch")
    doc.project_code = project_code
    doc.customer = customer
    doc.im = _im_resolved
    doc.po_no = po_no
    doc.po_line_no = 1
    doc.item_code = item_code
    if item_description:
        doc.item_description = item_description
    doc.qty = qty
    doc.rate = rate
    doc.line_amount = line_amount
    doc.site_code = site_code
    doc.site_name = site_name
    doc.center_area = center_area
    doc.region_type = region_type_from_center_area(center_area)
    doc.planning_mode = "Plan"
    doc.dispatch_status = "Dispatched"
    doc.dispatch_mode = "Manual"
    if target_month and frappe.db.has_column("PO Dispatch", "target_month"):
        doc.target_month = target_month
    if frappe.db.has_column("PO Dispatch", "is_dummy_po"):
        doc.is_dummy_po = 1

    final_name = _insert_po_dispatch_with_poid(doc, poid)
    stamp = {}
    if frappe.db.has_column("PO Dispatch", "original_dummy_poid"):
        # Store the business-visible POID (DUMMY-YYMMDD-XXXXXX-1), not the
        # doctype name (SYS-YYYY-NNNNN). The Dummy POID column everywhere
        # shows this value verbatim.
        stamp["original_dummy_poid"] = poid
    if frappe.db.has_column("PO Dispatch", "was_dummy_po"):
        stamp["was_dummy_po"] = 1
    if stamp:
        frappe.db.set_value("PO Dispatch", final_name, stamp, update_modified=False)
    frappe.db.commit()
    return {"name": final_name, "po_no": po_no, "poid": poid}


def _po_dispatch_portal_pf_active(pf):
    if not pf:
        return False
    if (pf.get("search") or pf.get("q") or "").strip():
        return True
    for k in (
        "project_code",
        "site_code",
        "team",
        "im",
        "from_date",
        "to_date",
        "dispatch_mode",
        "dummy_preset",
        "has_target_month",
    ):
        v = pf.get(k)
        if v is None or v == "":
            continue
        if k == "dummy_preset" and str(v).strip().lower() == "all":
            continue
        if k == "has_target_month" and str(v).strip().lower() in ("", "any"):
            continue
        return True
    return False


def _po_dispatch_portal_sql_where(filters, pf, fields):
    """Build WHERE clauses + bind values for portal-filtered PO Dispatch queries."""
    base_fl = (
        [[k, "=", v] for k, v in filters.items()]
        if isinstance(filters, dict)
        else [list(x) for x in (filters or [])]
    )
    for key in ("project_code", "site_code", "team"):
        vals = _ensure_list(pf.get(key))
        if vals:
            base_fl.append([key, "in" if len(vals) > 1 else "=", vals if len(vals) > 1 else vals[0]])
    im_vals = _ensure_list(pf.get("im"))
    if im_vals and "im" in fields:
        base_fl.append(["im", "in" if len(im_vals) > 1 else "=", im_vals if len(im_vals) > 1 else im_vals[0]])
    if pf.get("dispatch_mode"):
        base_fl.append(["dispatch_mode", "=", (pf.get("dispatch_mode") or "").strip()])
    if pf.get("from_date") and pf.get("to_date"):
        base_fl.append(["target_month", "between", [pf["from_date"], pf["to_date"]]])
    elif pf.get("from_date"):
        base_fl.append(["target_month", ">=", pf["from_date"]])
    elif pf.get("to_date"):
        base_fl.append(["target_month", "<=", pf["to_date"]])

    wheres = ["1=1"]
    params = []
    for cond in base_fl:
        if not isinstance(cond, (list, tuple)) or len(cond) != 3:
            continue
        k, op, v = cond[0], cond[1], cond[2]
        if k not in fields:
            continue
        op_l = str(op).lower()
        if op_l == "=":
            wheres.append(f"`{k}` = %s")
            params.append(v)
        elif op_l == "in" and isinstance(v, (list, tuple)) and v:
            ph = ", ".join(["%s"] * len(v))
            wheres.append(f"`{k}` IN ({ph})")
            params.extend(list(v))
        elif op_l == "between" and isinstance(v, (list, tuple)) and len(v) == 2:
            wheres.append(f"`{k}` BETWEEN %s AND %s")
            params.extend([v[0], v[1]])
        elif op_l in (">=", "<=", ">", "<"):
            wheres.append(f"`{k}` {op_l} %s")
            params.append(v)

    dummy_preset = (pf.get("dummy_preset") or "all").strip().lower()
    if dummy_preset == "dummy" and "is_dummy_po" in fields:
        wheres.append("IFNULL(`is_dummy_po`, 0) = 1")
    elif dummy_preset == "standard" and "is_dummy_po" in fields:
        wheres.append("(IFNULL(`is_dummy_po`, 0) = 0 OR `is_dummy_po` IS NULL)")
    elif dummy_preset == "mapped_dummy":
        wheres.append(
            "((IFNULL(was_dummy_po, 0) = 1 AND IFNULL(is_dummy_po, 0) = 0) OR "
            "(IFNULL(original_dummy_poid, '') != '' AND TRIM(original_dummy_poid) != IFNULL(TRIM(name), '')))"
        )

    # has_target_month: "yes" (target_month set), "no" (null/empty), "any" / "" (no filter)
    htm = (pf.get("has_target_month") or "").strip().lower()
    if htm == "yes" and "target_month" in fields:
        wheres.append("`target_month` IS NOT NULL AND `target_month` != ''")
    elif htm == "no" and "target_month" in fields:
        wheres.append("(`target_month` IS NULL OR `target_month` = '')")

    tokens = _sql_like_tokens(pf.get("search") or pf.get("q") or "")
    if tokens:
        like_cols = [
            c
            for c in (
                "name",
                "po_no",
                "item_code",
                "project_code",
                "site_code",
                "im",
                "center_area",
                "region_type",
                "original_dummy_poid",
            )
            if c in fields
        ]
        if like_cols:
            # Each token matches if ANY column contains it; any-token-matches
            # wins, so OR everything together.
            ors = []
            for _ in tokens:
                ors.extend(f"IFNULL(`{c}`, '') LIKE %s" for c in like_cols)
            wheres.append("(" + " OR ".join(ors) + ")")
            for tok in tokens:
                params.extend([tok] * len(like_cols))

    return wheres, params


@frappe.whitelist()
def list_po_dispatches(filters=None, order_by="modified desc", limit_page_length=100, portal_filters=None):
    """
    List PO Dispatch rows using only columns that exist in the database.

    ``frappe.client.get_list(..., fields=[\"*\"])`` expands ``*`` from DocType meta, so new
    fields in the JSON that are not yet migrated into MySQL cause OperationalError 1054.
    This endpoint avoids that by selecting physical table columns only.

    ``portal_filters`` (JSON dict) applies **before** ``limit_page_length`` so search / UI
    filters operate on the full dataset: search, project_code, site_code, team, im,
    from_date, to_date (on ``target_month``), dispatch_mode, dummy_preset
    (``all`` | ``dummy`` | ``standard`` | ``mapped_dummy``).
    """
    if isinstance(filters, str):
        filters = frappe.parse_json(filters) if filters else {}
    filters = filters or {}
    pf = _portal_filters_dict(portal_filters)
    limit_page_length = _portal_row_limit(limit_page_length, 100)
    fields = list(frappe.db.get_table_columns("PO Dispatch"))

    if not _po_dispatch_portal_pf_active(pf):
        gl_pd = dict(
            filters=filters,
            fields=fields,
            order_by=order_by or "modified desc",
        )
        if limit_page_length:
            gl_pd["limit_page_length"] = limit_page_length
        rows = frappe.get_list("PO Dispatch", **gl_pd)
    else:
        wheres, params = _po_dispatch_portal_sql_where(filters, pf, fields)
        ob = (order_by or "modified desc").strip()
        if not ob.replace(" ", "").replace(",", "").replace("_", "").isalnum():
            ob = "modified desc"
        cols_sql = ", ".join(f"`{c}`" for c in fields)
        sql = (
            f"SELECT {cols_sql} FROM `tabPO Dispatch` "
            f"WHERE {' AND '.join(wheres)} ORDER BY {ob}{_sql_limit_suffix(limit_page_length)}"
        )
        rows = frappe.db.sql(sql, tuple(params), as_dict=True)

    if not rows:
        return rows
    im_names = list({r.get("im") for r in rows if r.get("im")})
    im_names = [n for n in im_names if n]
    im_labels = {}
    if im_names:
        for imm in frappe.get_all(
            "IM Master",
            filters={"name": ["in", im_names]},
            fields=["name", "full_name"],
            limit_page_length=len(im_names) + 1,
        ):
            im_labels[imm.name] = imm.full_name or imm.name
    # Batch-fetch Customer Activity Type from Customer Item Master, keyed by
    # (customer, item_code). One query instead of per-row lookups.
    cim_map = _batch_customer_activity_types(rows)
    for r in rows:
        imn = r.get("im")
        r["im_full_name"] = im_labels.get(imn) if imn else None
        r["customer_activity_type"] = cim_map.get((r.get("customer") or "", r.get("item_code") or ""))
    return rows


def _next_visit_number_for_dispatch(po_dispatch_name):
    """Return the visit number for a NEW Rollout Plan on this POID.

    The first plan is visit #1, the next (re-visit) is #2, and so on. Uses
    MAX(visit_number) and COUNT(*) together so an already-filled column with
    gaps still advances monotonically.
    """
    if not po_dispatch_name:
        return 1
    try:
        row = frappe.db.sql(
            "SELECT COALESCE(MAX(visit_number), 0) AS max_v, COUNT(*) AS cnt "
            "FROM `tabRollout Plan` WHERE po_dispatch = %s",
            (po_dispatch_name,),
            as_dict=True,
        )
    except Exception:
        return 1
    if not row:
        return 1
    max_v = cint(row[0].get("max_v") or 0)
    cnt = cint(row[0].get("cnt") or 0)
    return max(max_v, cnt) + 1


def _batch_customer_activity_types(rows, customer_key="customer", item_key="item_code"):
    """Return {(customer, item_code): customer_activity_type} from Customer
    Item Master for every unique (customer, item) pair in `rows`. Picks the
    active-flag row first when multiple exist for a pair."""
    pairs = {(r.get(customer_key) or "", r.get(item_key) or "") for r in rows or []}
    pairs = {p for p in pairs if p[0] and p[1]}
    if not pairs:
        return {}
    customers = list({p[0] for p in pairs})
    items = list({p[1] for p in pairs})
    c_ph = ",".join(["%s"] * len(customers))
    i_ph = ",".join(["%s"] * len(items))
    try:
        cim_rows = frappe.db.sql(
            f"SELECT customer, item_code, customer_activity_type, IFNULL(active_flag, 0) AS active "
            f"FROM `tabCustomer Item Master` "
            f"WHERE customer IN ({c_ph}) AND item_code IN ({i_ph}) "
            f"ORDER BY IFNULL(active_flag, 0) DESC",
            tuple(customers) + tuple(items),
            as_dict=True,
        )
    except Exception:
        return {}
    out = {}
    for r in cim_rows:
        k = (r.customer, r.item_code)
        if k not in out and r.customer_activity_type:
            out[k] = r.customer_activity_type
    return out


@frappe.whitelist()
def get_po_dispatch_stats(filters=None, portal_filters=None):
    """
    Return aggregate counters for PO Dispatch rows matching the given filters.
    These counters are independent of any row/page limit so UI KPI cards stay
    stable no matter how many rows the table is currently rendering.

    Optional ``portal_filters`` uses the same rules as ``list_po_dispatches`` so
    KPI counts stay aligned when the UI applies search / IM / project / dummy
    presets (still no row limit).

    Response shape:
        { "total": int, "auto": int, "manual": int, "dispatched": int }
    """
    if isinstance(filters, str):
        filters = frappe.parse_json(filters) if filters else {}
    filters = filters or {}
    pf = _portal_filters_dict(portal_filters)
    fields = list(frappe.db.get_table_columns("PO Dispatch"))

    if _po_dispatch_portal_pf_active(pf):
        wheres, params = _po_dispatch_portal_sql_where(filters, pf, fields)
        base_sql = f"FROM `tabPO Dispatch` WHERE {' AND '.join(wheres)}"

        def _cnt(extra_sql="", extra_params=None):
            xp = list(extra_params or ())
            if extra_sql:
                q = f"SELECT COUNT(*) {base_sql} AND ({extra_sql})"
            else:
                q = f"SELECT COUNT(*) {base_sql}"
            row = frappe.db.sql(q, tuple(params) + tuple(xp))
            return int(row[0][0]) if row else 0

        total = _cnt()
        auto = _cnt("`dispatch_mode` = %s", ["Auto"])
        manual = _cnt("`dispatch_mode` = %s", ["Manual"])
        dispatched = _cnt("`dispatch_status` = %s", ["Dispatched"])
        return {
            "total": total,
            "auto": auto,
            "manual": manual,
            "dispatched": dispatched,
        }

    total = frappe.db.count("PO Dispatch", filters=filters) or 0

    def _count_with(extra):
        merged_filters = list(filters) if isinstance(filters, list) else dict(filters)
        if isinstance(merged_filters, list):
            for k, v in extra.items():
                merged_filters.append([k, "=", v])
        else:
            merged_filters.update(extra)
        try:
            return frappe.db.count("PO Dispatch", filters=merged_filters) or 0
        except Exception:
            return 0

    auto = _count_with({"dispatch_mode": "Auto"})
    manual = _count_with({"dispatch_mode": "Manual"})
    dispatched = _count_with({"dispatch_status": "Dispatched"})
    return {
        "total": int(total),
        "auto": int(auto),
        "manual": int(manual),
        "dispatched": int(dispatched),
    }


@frappe.whitelist()
def list_po_intake_lines_for_im_map(project_code=None):
    """
    PO Intake lines for a project — IM only, for mapping a dummy dispatch to a real line.
    """
    project_code = (project_code or "").strip()
    if not project_code:
        frappe.throw("project_code is required")

    _im_resolved, im_identifiers = _require_inet_im_session()
    if not _pcc_im_allows_project(project_code, im_identifiers):
        frappe.throw("Not permitted for this project.")

    line_fields = [
        "name",
        "parent",
        "po_line_no",
        "poid",
        "shipment_number",
        "item_code",
        "item_description",
        "qty",
        "rate",
        "line_amount",
        "project_code",
        "site_code",
        "site_name",
        "center_area",
        "region_type",
        "po_line_status",
    ]
    try:
        lines = frappe.get_all(
            "PO Intake Line",
            filters={"project_code": project_code, "po_line_status": ["!=", "Completed"]},
            fields=line_fields,
            order_by="modified desc",
            limit_page_length=400,
        )
    except frappe.db.OperationalError as e:
        if not frappe.db.is_missing_column(e):
            raise
        line_fields.remove("center_area")
        line_fields.remove("region_type")
        lines = frappe.get_all(
            "PO Intake Line",
            filters={"project_code": project_code, "po_line_status": ["!=", "Completed"]},
            fields=line_fields,
            order_by="modified desc",
            limit_page_length=400,
        )
        for row in lines:
            row["center_area"] = None
            row["region_type"] = None

    # Batch enrichment — one query for parents, one for dispatches.
    parent_names = list({row.get("parent") for row in lines if row.get("parent")})
    parent_map = {}
    if parent_names:
        for p in frappe.get_all(
            "PO Intake",
            filters={"name": ["in", parent_names]},
            fields=["name", "po_no", "customer"],
            limit_page_length=len(parent_names) + 1,
        ):
            parent_map[p.name] = p

    dispatch_map = {}
    if parent_names:
        for d in frappe.get_all(
            "PO Dispatch",
            filters={"po_intake": ["in", parent_names]},
            fields=["name", "po_intake", "po_line_no", "dispatch_status"],
            limit_page_length=len(parent_names) * 100 + 1,
        ):
            dispatch_map[(d.po_intake, cint(d.po_line_no))] = d

    for row in lines:
        pdata = parent_map.get(row.get("parent")) or {}
        row["po_no"] = pdata.get("po_no", "")
        row["customer"] = pdata.get("customer", "")
        row["po_intake"] = row.get("parent")
        drow = dispatch_map.get((row.get("parent"), cint(row.get("po_line_no"))))
        if drow:
            row["existing_dispatch"] = drow.get("name")
            row["existing_dispatch_status"] = drow.get("dispatch_status")
        else:
            row["existing_dispatch"] = None
            row["existing_dispatch_status"] = None

    return lines


@frappe.whitelist()
def assign_im_target_month(payload=None):
    """Bulk-assign `target_month` on a set of PO Dispatch rows so the IM
    promotes them from "PO Intake" (dispatched but not yet scheduled) into
    "My Dispatches" (ready for rollout planning).

    payload = {
        "dispatches": ["SYS-2026-0001", ...],
        "target_month": "2026-04"  (or "YYYY-MM-DD")
    }
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)
    payload = payload or {}

    dispatches = payload.get("dispatches") or []
    target_month = (payload.get("target_month") or "").strip()
    if not dispatches or not isinstance(dispatches, list):
        frappe.throw("dispatches (list of PO Dispatch names) is required")
    if not target_month:
        frappe.throw("target_month is required")

    # YYYY-MM shorthand → first day of month
    if len(target_month) == 7 and target_month[4] == "-":
        target_month = target_month + "-01"

    _im_resolved, im_identifiers = _require_inet_im_session()
    # Only allow the IM to update their own dispatches.
    allowed = frappe.db.sql(
        "SELECT name FROM `tabPO Dispatch` WHERE name IN ({ph}) AND IFNULL(im, '') IN ({im_ph})".format(
            ph=",".join(["%s"] * len(dispatches)),
            im_ph=",".join(["%s"] * len(im_identifiers)),
        ),
        tuple(dispatches) + tuple(im_identifiers),
    )
    allowed_names = [r[0] for r in (allowed or [])]
    if not allowed_names:
        frappe.throw("No matching dispatches belong to the current IM.")

    updated = 0
    for name in allowed_names:
        frappe.db.set_value("PO Dispatch", name, "target_month", target_month, update_modified=True)
        updated += 1
    frappe.db.commit()
    return {"updated": updated, "target_month": target_month, "names": allowed_names}


@frappe.whitelist()
def map_im_dummy_po_to_intake_line(payload=None):
    """
    Link a dummy PO Dispatch to a real PO Intake line and rename to business POID.

    payload: {
        "dummy_po_dispatch": "PO Dispatch name (current POID)",
        "po_intake_line": "PO Intake Line child row name",
    }

    If the intake line already has another dispatch from intake/PM dispatch (often Pending or
    Dispatched) with **no** rollout plans and **no** executions on that dispatch, it is removed
    so the dummy can take the real POID. If that dispatch has rollouts/executions or is
    Completed, mapping is blocked.
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)
    payload = payload or {}

    dummy_name = (payload.get("dummy_po_dispatch") or "").strip()
    line_name = (payload.get("po_intake_line") or "").strip()
    if not dummy_name or not line_name:
        frappe.throw("dummy_po_dispatch and po_intake_line are required")

    _im_resolved, im_identifiers = _require_inet_im_session()

    if not frappe.db.exists("PO Dispatch", dummy_name):
        frappe.throw("PO Dispatch not found.")

    d_im = frappe.db.get_value("PO Dispatch", dummy_name, "im")
    if d_im not in set(im_identifiers):
        frappe.throw("Not permitted for this dispatch.")

    if frappe.db.has_column("PO Dispatch", "is_dummy_po"):
        if not cint(frappe.db.get_value("PO Dispatch", dummy_name, "is_dummy_po")):
            frappe.throw("This dispatch is not marked as dummy PO.")

    if not frappe.db.exists("PO Intake Line", line_name):
        frappe.throw("PO Intake Line not found.")

    line_row = frappe.db.get_value("PO Intake Line", line_name, "*", as_dict=True)
    parent_intake = line_row.get("parent")
    project_line = (line_row.get("project_code") or "").strip()
    if not parent_intake or not frappe.db.exists("PO Intake", parent_intake):
        frappe.throw("Invalid PO Intake Line parent.")

    if not _pcc_im_allows_project(project_line, im_identifiers):
        frappe.throw("Not permitted for this PO line’s project.")

    d_proj = (frappe.db.get_value("PO Dispatch", dummy_name, "project_code") or "").strip()
    if project_line != d_proj:
        frappe.throw(
            "Project on the PO line must match the dummy dispatch project."
        )

    parent_po_no = frappe.db.get_value("PO Intake", parent_intake, "po_no") or ""
    customer = frappe.db.get_value("PO Intake", parent_intake, "customer")

    po_line_no = cint(line_row.get("po_line_no") or 0)
    shipment_number = line_row.get("shipment_number")
    poid_target = (line_row.get("poid") or "").strip() or _resolve_line_poid(
        parent_po_no,
        po_line_no,
        shipment_number,
    )

    stub = _get_dispatch_for_intake_line(parent_intake, po_line_no)
    stub_name = stub.get("name")
    if stub_name and stub_name != dummy_name:
        rp_count = frappe.db.count("Rollout Plan", {"po_dispatch": stub_name})
        ex_count = frappe.db.sql(
            """
            SELECT COUNT(*) FROM `tabDaily Execution` de
            INNER JOIN `tabRollout Plan` rp ON rp.name = de.rollout_plan
            WHERE rp.po_dispatch = %s
            """,
            (stub_name,),
        )[0][0]
        if rp_count or ex_count:
            frappe.throw(
                f"PO line already has dispatch {stub_name} with rollout or execution activity. "
                "Finish or cancel that work before mapping this dummy PO."
            )
        stub_status = (frappe.db.get_value("PO Dispatch", stub_name, "dispatch_status") or "").strip()
        if stub_status == "Completed":
            frappe.throw(
                f"This PO line is linked to a completed dispatch ({stub_name}). Mapping is not allowed."
            )
        frappe.delete_doc("PO Dispatch", stub_name, force=True, ignore_permissions=True)

    center_area = line_row.get("center_area") or line_row.get("area")
    site_code = (line_row.get("site_code") or "").strip()
    site_name = line_row.get("site_name")
    if site_code:
        ok_duid, err_duid = ensure_duid_master(site_code, site_name, center_area)
        if not ok_duid:
            frappe.throw(err_duid or f"Could not ensure DUID Master for {site_code}")

    update_vals = {
        "po_intake": parent_intake,
        "po_no": parent_po_no,
        "po_line_no": po_line_no,
        "item_code": line_row.get("item_code"),
        "item_description": line_row.get("item_description"),
        "qty": flt(line_row.get("qty", 0)),
        "rate": flt(line_row.get("rate", 0)),
        "line_amount": flt(line_row.get("line_amount", 0)),
        "project_code": project_line,
        "customer": customer or frappe.db.get_value(
            "Project Control Center", project_line, "customer"
        ),
        "center_area": center_area,
        "region_type": line_row.get("region_type")
        or region_type_from_center_area(center_area),
        "site_code": site_code or None,
        "site_name": site_name,
    }
    if frappe.db.has_column("PO Dispatch", "is_dummy_po"):
        update_vals["is_dummy_po"] = 0
    if frappe.db.has_column("PO Dispatch", "was_dummy_po"):
        update_vals["was_dummy_po"] = 1
    if frappe.db.has_column("PO Dispatch", "original_dummy_poid"):
        ex_orig = (
            frappe.db.get_value("PO Dispatch", dummy_name, "original_dummy_poid") or ""
        ).strip()
        if not ex_orig:
            # Store the dummy's own POID (not its SYS- doc name) as the origin
            dummy_poid = (frappe.db.get_value("PO Dispatch", dummy_name, "poid") or dummy_name).strip()
            update_vals["original_dummy_poid"] = dummy_poid

    # Set real POID on the poid field — no rename needed
    if poid_target:
        existing_with_poid = frappe.db.get_value("PO Dispatch", {"poid": poid_target}, "name")
        if existing_with_poid and existing_with_poid != dummy_name:
            frappe.throw(
                f"POID {poid_target} already assigned to another dispatch — resolve duplicates first."
            )
        update_vals["poid"] = poid_target

    frappe.db.set_value("PO Dispatch", dummy_name, update_vals, update_modified=True)

    frappe.db.set_value(
        "PO Intake Line",
        line_name,
        {"po_line_status": "Dispatched", "dispatch_mode": "Manual"},
        update_modified=True,
    )

    frappe.db.commit()
    gv_fields = ["po_no", "po_line_no", "poid"]
    if frappe.db.has_column("PO Dispatch", "original_dummy_poid"):
        gv_fields.append("original_dummy_poid")
    if frappe.db.has_column("PO Dispatch", "was_dummy_po"):
        gv_fields.append("was_dummy_po")
    meta = frappe.db.get_value("PO Dispatch", dummy_name, gv_fields, as_dict=True) or {}
    return {
        "name": dummy_name,
        "poid": meta.get("poid") or poid_target,
        "po_intake": parent_intake,
        "po_intake_line": line_name,
        "original_dummy_poid": (meta.get("original_dummy_poid") or "").strip(),
        "was_dummy_po": cint(meta.get("was_dummy_po")),
        "po_no": meta.get("po_no") or parent_po_no,
        "po_line_no": cint(meta.get("po_line_no") or po_line_no),
    }


@frappe.whitelist()
def dispatch_po_lines(payload):
    """
    Dispatch PO lines to an Implementation Manager for a target month.

    payload = {
        "lines": [{"po_intake": "PIP-...", "item_code": "...", ...}],
        "im": "IM Master document name",  # required (or legacy "team" to derive im)
        "target_month": "2026-04",
        "planning_mode": "Plan",
    }

    Returns
    -------
    {"created": N, "poids": [...]}
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)

    lines = payload.get("lines") or []
    im = (payload.get("im") or "").strip() or None
    team_id = payload.get("team")  # deprecated: derive IM from team if im not sent
    raw_month = payload.get("target_month") or ""
    # Handle "2026-04" from HTML month input -> "2026-04-01"
    if raw_month and len(raw_month) == 7:
        target_month = raw_month + "-01"
    else:
        target_month = raw_month
    planning_mode = payload.get("planning_mode", "Plan")

    if not im and team_id:
        team_doc = frappe.db.get_value("INET Team", team_id, ["im"], as_dict=True)
        if team_doc:
            im = team_doc.im

    if not im:
        frappe.throw(
            frappe._("Implementation Manager (im) is required to dispatch PO lines.")
        )

    created = 0
    poids = []

    for line in lines:
        po_intake_name = line.get("po_intake") or line.get("parent")
        po_line_no = cint(line.get("po_line_no") or 0)
        item_code = line.get("item_code")
        item_description = line.get("item_description")
        qty = flt(line.get("qty", 0))
        rate = flt(line.get("rate", 0))
        line_amount = flt(line.get("line_amount", 0))
        project_code = line.get("project_code")
        center_area = line.get("center_area") or line.get("area")
        site_code = line.get("site_code")
        site_name = line.get("site_name")
        po_no = line.get("po_no")
        line_child_name = line.get("name")  # child table row name
        poid = _resolve_line_poid(
            po_no=po_no,
            po_line_no=po_line_no,
            shipment_number=line.get("shipment_number"),
            fallback=line.get("poid"),
        )

        # Resolve customer from PO Intake parent
        customer = line.get("customer")
        if not customer and project_code:
            customer = frappe.db.get_value(
                "Project Control Center", project_code, "customer"
            )

        final_dispatch_name = _upsert_po_dispatch_for_line(
            po_intake_name,
            po_no,
            {
                "po_line_no": po_line_no,
                "item_code": item_code,
                "item_description": item_description,
                "qty": qty,
                "rate": rate,
                "line_amount": line_amount,
                "project_code": project_code,
                "center_area": center_area,
                "site_code": site_code,
                "site_name": site_name,
                "shipment_number": line.get("shipment_number"),
                "poid": poid,
            },
            customer=customer,
            im=im,
            target_month=target_month,
            planning_mode=planning_mode,
            dispatch_status="Dispatched",
            dispatch_mode="Manual",
        )

        # Mark the PO Intake Line as "Dispatched"
        if line_child_name:
            frappe.db.set_value(
                "PO Intake Line", line_child_name,
                {"po_line_status": "Dispatched", "dispatch_mode": "Manual"},
                update_modified=False,
            )

        frappe.db.commit()
        created += 1
        poids.append(final_dispatch_name)

    return {"created": created, "poids": poids}


@frappe.whitelist()
def convert_dispatch_mode(payload):
    """
    Convert Auto-dispatched lines to Manual (or vice-versa).

    payload = {
        "scope": "lines" | "project",
        "line_names": [...],        # PO Intake Line child row names (for scope=lines)
        "project_code": "...",      # required when scope=project
        "new_im": "im_name",        # optional — re-assign IM (IM Master name)
        "target_mode": "Manual",    # "Manual" or "Auto" (default "Manual")
    }

    Returns {"converted": N}
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)

    scope = payload.get("scope", "lines")
    project_code = payload.get("project_code")
    line_names = payload.get("line_names") or []
    new_im = payload.get("new_im")
    target_mode = payload.get("target_mode", "Manual")

    if scope == "project" and project_code:
        line_names = frappe.get_all(
            "PO Intake Line",
            filters={"project_code": project_code, "po_line_status": "Dispatched"},
            pluck="name",
        )

    count = 0
    for line_name in line_names:
        frappe.db.set_value(
            "PO Intake Line", line_name, "dispatch_mode", target_mode,
            update_modified=False,
        )

        line_data = frappe.db.get_value(
            "PO Intake Line", line_name, ["parent", "po_line_no"], as_dict=True
        )
        if line_data:
            dispatch_names = frappe.get_all(
                "PO Dispatch",
                filters={"po_intake": line_data.parent, "po_line_no": line_data.po_line_no},
                pluck="name",
            )
            for dispatch_name in dispatch_names:
                update_vals = {"dispatch_mode": target_mode}
                if new_im:
                    update_vals["im"] = new_im
                frappe.db.set_value(
                    "PO Dispatch", dispatch_name, update_vals, update_modified=False
                )
            count += len(dispatch_names)

        frappe.db.commit()

    return {"converted": count}


@frappe.whitelist()
def create_rollout_plans(payload):
    """
    Create Rollout Plans for a list of PO Dispatch system IDs.

    payload = {
        "dispatches": ["SYS-2026-00001", ...],
        "plan_date": "2026-04-04",
        "plan_end_date": "2026-04-05",  # optional; defaults to plan_date
        "team": "TEAM-001",  # required INET Team name (IM chooses at planning; stored on Rollout Plan only)
        "access_time": "",  # optional
        "access_period": "Day" | "Night" | "",  # optional
        "visit_type": "Work Done",
    }

    Returns
    -------
    {"created": N, "names": [...]}
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)

    dispatches = payload.get("dispatches") or []
    plan_date = payload.get("plan_date") or nowdate()
    plan_end_date = payload.get("plan_end_date") or plan_date
    team_override = (payload.get("team") or "").strip()
    access_time = (payload.get("access_time") or "").strip()
    access_period = (payload.get("access_period") or "").strip()
    if access_period and access_period not in ("Day", "Night"):
        access_period = ""
    visit_type = payload.get("visit_type") or "Execution"

    pd = getdate(plan_date)
    ped = getdate(plan_end_date)
    if ped < pd:
        frappe.throw(frappe._("Planned end date cannot be before plan start date"))

    if not team_override:
        frappe.throw(
            frappe._(
                "Select a field team for this rollout. Team is stored on the Rollout Plan (not on PO Dispatch)."
            )
        )

    if not frappe.db.exists("INET Team", team_override):
        frappe.throw(frappe._("Invalid team selected"))

    # Look up multiplier. "Execution" is the new label for what used to be
    # "Work Done"; fall back so renames don't require touching the master.
    multiplier_val = frappe.db.get_value("Visit Multiplier Master", visit_type, "multiplier")
    if multiplier_val is None and visit_type == "Execution":
        multiplier_val = frappe.db.get_value("Visit Multiplier Master", "Work Done", "multiplier")
    visit_multiplier = flt(multiplier_val or 1.0)

    created = 0
    names = []

    # Optional remark fields the caller can stamp on each PO Dispatch as part
    # of plan creation. Same write rules as `update_po_remark`:
    #   general    → PM only
    #   manager    → PM + IM
    #   team_lead  → PM + IM + Field
    role = _user_role_class()
    remark_updates = {}
    if (payload.get("general_remark") is not None) and role == "pm":
        remark_updates["general_remark"] = str(payload.get("general_remark") or "")[:8000]
    if (payload.get("manager_remark") is not None) and role in ("pm", "im"):
        remark_updates["manager_remark"] = str(payload.get("manager_remark") or "")[:8000]
    if (payload.get("team_lead_remark") is not None) and role in ("pm", "im", "field"):
        remark_updates["team_lead_remark"] = str(payload.get("team_lead_remark") or "")[:8000]

    for dispatch_name in dispatches:
        dispatch = frappe.db.get_value(
            "PO Dispatch",
            dispatch_name,
            ["name", "line_amount", "im", "region_type", "center_area"],
            as_dict=True,
        )
        if not dispatch:
            continue

        target_team = team_override

        # Reuse-instead-of-recreate: when this dispatch's latest plan was
        # created from an issue (Planning-with-Issue placeholder, or a
        # previously-planned issue plan still pre-execution), don't allocate
        # a NEW visit_number — repurpose that plan as the actionable target.
        # This keeps visit_number aligned with the physical attempt count
        # (1, 2, 2, 3, …) and lets "Create Plans from I&R" be clicked
        # repeatedly without each click incrementing the visit.
        latest_plan = frappe.db.sql(
            """
            SELECT name, visit_number, plan_status,
                   IFNULL(issue_category, '') AS issue_category,
                   IFNULL(visit_type, '') AS visit_type
            FROM `tabRollout Plan`
            WHERE po_dispatch = %s
            ORDER BY IFNULL(visit_number, 0) DESC, modified DESC
            LIMIT 1
            """,
            (dispatch_name,),
            as_dict=True,
        )
        latest = latest_plan[0] if latest_plan else None
        # Reuse the latest plan when it's an open issue placeholder of any kind:
        # currently in Planning with Issue, or already Planned but originated
        # from a re-plan (visit_type = Re-Visit, or has an issue_category set).
        reusable = bool(
            latest and (
                latest.plan_status == "Planning with Issue"
                or (
                    latest.plan_status == "Planned"
                    and (latest.issue_category or latest.visit_type == "Re-Visit")
                )
            )
        )
        if reusable:
            # Move it into "Planned" so it shows up on the Planning page
            # alongside other actionable plans, while the I&R row stays
            # visible (issue_category is preserved).
            updates = {
                "team": target_team,
                "plan_date": plan_date,
                "plan_end_date": plan_end_date,
                "access_time": access_time,
                "access_period": access_period or None,
                "visit_type": visit_type,
                "plan_status": "Planned",
            }
            if payload.get("issue_category"):
                updates["issue_category"] = str(payload["issue_category"])[:140]
            if payload.get("issue_remarks") and frappe.db.has_column("Rollout Plan", "issue_remarks"):
                updates["issue_remarks"] = str(payload["issue_remarks"])[:2000]
            frappe.db.set_value(
                "Rollout Plan", latest.name, updates, update_modified=True
            )
            disp_updates = {"dispatch_status": "Planned"}
            disp_updates.update(remark_updates)
            frappe.db.set_value(
                "PO Dispatch",
                dispatch_name,
                disp_updates,
                update_modified=False,
            )
            frappe.db.commit()
            created += 1
            names.append(latest.name)
            continue

        doc = frappe.new_doc("Rollout Plan")
        doc.po_dispatch = dispatch_name
        doc.team = target_team
        if dispatch.im and hasattr(doc, "im"):
            doc.im = dispatch.im
        doc.plan_date = plan_date
        doc.plan_end_date = plan_end_date
        doc.access_time = access_time
        doc.access_period = access_period or None
        doc.visit_type = visit_type
        # Visit # advances per POID: 1st plan = 1, 2nd = 2 (Re-Visit), etc.
        doc.visit_number = _next_visit_number_for_dispatch(dispatch_name)
        doc.visit_multiplier = visit_multiplier
        doc.target_amount = flt(dispatch.line_amount) * visit_multiplier
        doc.plan_status = "Planned"
        # Issue & Risk fields — only attached when a re-plan carries them.
        if payload.get("issue_category"):
            doc.issue_category = str(payload["issue_category"])[:140]
        if payload.get("issue_remarks") and hasattr(doc, "issue_remarks"):
            doc.issue_remarks = str(payload["issue_remarks"])[:2000]
        if hasattr(doc, "region_type"):
            doc.region_type = dispatch.get("region_type") or region_type_from_center_area(
                dispatch.get("center_area")
            )

        doc.insert(ignore_permissions=True)
        frappe.db.commit()

        disp_updates = {"dispatch_status": "Planned"}
        disp_updates.update(remark_updates)
        frappe.db.set_value(
            "PO Dispatch",
            dispatch_name,
            disp_updates,
            update_modified=False,
        )

        created += 1
        names.append(doc.name)

    return {"created": created, "names": names}


_EXEC_STATUSES_ROLLOUT_IN_PROGRESS_LIKE = frozenset(
    (
        "In Progress",
        "POD Pending",
        "PO Required",
        "Span Loss",
        "Spare Parts",
        "Extra Visit",
        "Late Arrival",
        "Quality Issue",
        "Travel",
    )
)


def _sync_rollout_plan_from_daily_execution(rollout_plan, exec_doc):
    """
    Keep Rollout Plan plan_status (and optional amounts) in sync with Daily Execution.
    Options: Planned, Planning with Issue, In Execution, Completed, Cancelled.

    Two completion signals are honoured, in priority order:
    1. ``execution_status`` — the IM's authoritative confirmation.
    2. ``tl_status``        — the field team lead's "I'm done" flag, which
       reaches the system before the IM has a chance to confirm.

    Either one flipping to *Completed* moves the Rollout Plan to *Completed*
    so the line surfaces in the field team's QC / CIAG queue immediately;
    the IM's later confirmation only updates ``execution_status`` (read-only
    on the field side). This removes the chicken-and-egg where the field
    team marked work done but couldn't see it on the QC page until the IM
    happened to flip a separate flag.
    """
    if not rollout_plan or not exec_doc:
        return
    st = exec_doc.execution_status
    tls = getattr(exec_doc, "tl_status", None) or ""
    # ``effective`` is the status that drives plan_status: prefer the IM's
    # confirmed value, fall back to the team lead's signal.
    effective = st or tls
    updates = {}

    if (st in _EXEC_STATUSES_ROLLOUT_IN_PROGRESS_LIKE) or (
        not st and tls in _EXEC_STATUSES_ROLLOUT_IN_PROGRESS_LIKE
    ):
        cur = frappe.db.get_value("Rollout Plan", rollout_plan, "plan_status")
        if cur == "Planned":
            updates["plan_status"] = "In Execution"

    elif effective == "Completed":
        qc = str(getattr(exec_doc, "qc_status", None) or "")
        ach_amt = flt(getattr(exec_doc, "achieved_amount", None) or 0)
        tgt = flt(frappe.db.get_value("Rollout Plan", rollout_plan, "target_amount") or 0)
        if qc == "Fail":
            # QC failure returns work to planning; do not close the original plan as Completed.
            updates["plan_status"] = "Planning with Issue"
        else:
            # Completed execution should reflect as Completed in planning monitor.
            # QC/CIAG is a follow-up review step and should not revert plan to In Execution.
            updates["plan_status"] = "Completed"
        if ach_amt > 0:
            updates["achieved_amount"] = ach_amt
        if tgt > 0 and ach_amt >= 0:
            updates["completion_pct"] = min(100.0, (ach_amt / tgt) * 100.0)

    elif effective == "Cancelled":
        updates["plan_status"] = "Cancelled"

    elif effective == "Postponed":
        updates["plan_status"] = "Planned"

    # Hold: keep current plan status (still on site / in progress)

    if updates:
        frappe.db.set_value("Rollout Plan", rollout_plan, updates)


_ALLOWED_DAILY_EXECUTION_STATUSES = frozenset(
    (
        "In Progress",
        "Completed",
        "Hold",
        "Cancelled",
        "Postponed",
        "POD Pending",
        "PO Required",
        "Span Loss",
        "Spare Parts",
        "Extra Visit",
        "Late Arrival",
        "Quality Issue",
        "Travel",
    )
)


def _normalize_execution_status(value):
    s = str(value or "").strip()
    if not s:
        return s
    sl = s.lower().replace("_", " ")
    if sl in ("in progress", "inprogress"):
        return "In Progress"
    if sl in ("complete", "completed"):
        return "Completed"
    if sl in ("cancel", "cancelled", "canceled"):
        return "Cancelled"
    if sl in ("postpone", "postponed"):
        return "Postponed"
    if sl in ("hold", "on hold"):
        return "Hold"
    if sl in ("pod pending", "podpending"):
        return "POD Pending"
    if sl in ("po required", "porequired"):
        return "PO Required"
    if sl in ("span loss", "spanloss"):
        return "Span Loss"
    if sl in ("spare parts", "spareparts"):
        return "Spare Parts"
    if sl in ("extra visit", "extravisit"):
        return "Extra Visit"
    if sl in ("late arrival", "latearrival"):
        return "Late Arrival"
    if sl in ("quality issue", "qualityissue"):
        return "Quality Issue"
    if sl == "travel":
        return "Travel"
    # Title-style match against allowed labels (e.g. user-typed casing)
    for opt in _ALLOWED_DAILY_EXECUTION_STATUSES:
        if sl == opt.lower():
            return opt
    return s


def _get_rollout_billing_rate(rollout_plan):
    if not rollout_plan:
        return 0.0

    rp = frappe.db.get_value("Rollout Plan", rollout_plan, ["po_dispatch"], as_dict=True)
    if not rp or not rp.po_dispatch:
        return 0.0

    dispatch = frappe.db.get_value(
        "PO Dispatch",
        rp.po_dispatch,
        ["item_code", "center_area", "region_type", "customer"],
        as_dict=True,
    )
    if not dispatch or not dispatch.item_code:
        return 0.0

    is_hard = is_hard_region(dispatch.region_type, dispatch.center_area)
    cim_filters = {"item_code": dispatch.item_code, "active_flag": 1}
    if dispatch.customer:
        cim_filters["customer"] = dispatch.customer
    cim = frappe.db.get_value(
        "Customer Item Master",
        cim_filters,
        ["standard_rate_sar", "hard_rate_sar"],
        as_dict=True,
    )
    if not cim:
        cim = frappe.db.get_value(
            "Customer Item Master",
            {"item_code": dispatch.item_code, "active_flag": 1},
            ["standard_rate_sar", "hard_rate_sar"],
            as_dict=True,
        )
    if not cim:
        return 0.0
    return flt(cim.hard_rate_sar if is_hard else cim.standard_rate_sar)


def _user_execution_update_mode(user, doc):
    """
    IM/Admin: full update. INET Field Team members (team lead + roster): qc_status / ciag_status only.
    """
    if not user or user == "Guest":
        return None
    roles = set(frappe.get_roles(user))
    if (
        "Administrator" in roles
        or "System Manager" in roles
        or "INET Admin" in roles
        or "INET IM" in roles
    ):
        return "full"
    if "INET Field Team" not in roles:
        return None
    team_id = getattr(doc, "team", None)
    if not team_id or not frappe.db.exists("INET Team", team_id):
        return None
    td = frappe.get_doc("INET Team", team_id)
    members = set()
    if td.field_user:
        members.add(td.field_user)
    for r in td.team_members or []:
        emp = getattr(r, "employee", None)
        if not emp:
            continue
        uid = frappe.db.get_value("Employee", emp, "user_id")
        if uid:
            members.add(uid)
    if user in members:
        return "qc_ciag"
    return None


@frappe.whitelist()
def get_field_execution_for_rollout(rollout_plan):
    """Latest Daily Execution for this rollout.

    Field team: scoped to the logged-in user's INET team (QC/CIAG flow).
    IM / desk admin: any team on that rollout so IM can record execution without a field_user.
    """
    rollout_plan = (rollout_plan or "").strip()
    if not rollout_plan:
        return None
    user = frappe.session.user
    roles = set(frappe.get_roles(user))
    im_scope = (
        "Administrator" in roles
        or "System Manager" in roles
        or "INET Admin" in roles
        or "INET IM" in roles
    )
    if im_scope:
        rows = frappe.get_all(
            "Daily Execution",
            filters={"rollout_plan": rollout_plan},
            fields=["name", "qc_status", "ciag_status", "execution_status", "photos"],
            order_by="modified desc",
            limit_page_length=1,
        )
        return rows[0] if rows else None
    team_id = _session_inet_field_team_id()
    if not team_id:
        return None
    rows = frappe.get_all(
        "Daily Execution",
        filters={"rollout_plan": rollout_plan, "team": team_id},
        fields=["name", "qc_status", "ciag_status", "execution_status", "photos"],
        order_by="modified desc",
        limit_page_length=1,
    )
    return rows[0] if rows else None


@frappe.whitelist()
def get_rollout_plan_details(rollout_plan):
    """Return a single Rollout Plan enriched with PO Dispatch context.

    Used by the Field Execute form to render the Plan Details panel —
    item description, activity type, DUID, POID, qty, target amount, and
    the IM's confirmed ``execution_status`` (read-only badge).

    The plain ``frappe.client.get_list`` call only returns Rollout Plan
    columns, so the FE was missing item / qty / customer / poid until now.
    """
    rollout_plan = (rollout_plan or "").strip()
    if not rollout_plan:
        return None

    rp_cols = (
        "name, po_dispatch, im, region_type, team, plan_date, plan_end_date, "
        "visit_type, visit_number, visit_multiplier, target_amount, "
        "achieved_amount, completion_pct, plan_status, modified"
    )
    rp_rows = frappe.db.sql(
        f"SELECT {rp_cols} FROM `tabRollout Plan` WHERE name = %s LIMIT 1",
        (rollout_plan,),
        as_dict=True,
    )
    if not rp_rows:
        return None
    plan = rp_rows[0]

    pd_name = plan.get("po_dispatch")
    if pd_name:
        pd = frappe.db.get_value(
            "PO Dispatch",
            pd_name,
            [
                "poid", "po_no", "po_line_no",
                "item_code", "item_description", "qty", "rate", "line_amount",
                "project_code", "customer", "site_code", "site_name",
                "center_area", "region_type",
            ],
            as_dict=True,
        ) or {}
        plan["poid"] = pd.get("poid")
        plan["po_no"] = pd.get("po_no")
        plan["po_line_no"] = pd.get("po_line_no")
        plan["item_code"] = pd.get("item_code")
        plan["item_description"] = pd.get("item_description")
        plan["qty"] = pd.get("qty")
        plan["rate"] = pd.get("rate")
        plan["line_amount"] = pd.get("line_amount")
        plan["project_code"] = pd.get("project_code")
        plan["customer"] = pd.get("customer")
        plan["site_code"] = pd.get("site_code")
        plan["site_name"] = pd.get("site_name")
        plan["center_area"] = pd.get("center_area")
        if not plan.get("region_type"):
            plan["region_type"] = pd.get("region_type") or region_type_from_center_area(
                pd.get("center_area")
            )
        # Activity type is keyed by (customer, item_code) in CIM master.
        cim_map = _batch_customer_activity_types([{"customer": pd.get("customer"), "item_code": pd.get("item_code")}])
        plan["customer_activity_type"] = cim_map.get(
            (pd.get("customer") or "", pd.get("item_code") or "")
        )

    # Pull the latest Daily Execution for this plan so the form can show
    # the IM's confirmed ``execution_status`` and the field team's own
    # ``tl_status`` without a second round-trip.
    de_rows = frappe.db.sql(
        "SELECT name, execution_status, tl_status, qc_status, ciag_status, "
        "       achieved_qty, achieved_amount "
        "FROM `tabDaily Execution` WHERE rollout_plan = %s "
        "ORDER BY modified DESC LIMIT 1",
        (rollout_plan,),
        as_dict=True,
    )
    if de_rows:
        de = de_rows[0]
        plan["execution_status"] = de.get("execution_status")
        plan["tl_status"] = de.get("tl_status")
        plan["qc_status"] = de.get("qc_status")
        plan["ciag_status"] = de.get("ciag_status")
        plan["execution_name"] = de.get("name")
    else:
        plan["execution_status"] = None
        plan["tl_status"] = None
        plan["qc_status"] = None
        plan["ciag_status"] = None
        plan["execution_name"] = None

    return plan


@frappe.whitelist()
def update_execution(payload):
    """
    Create or update a Daily Execution record.

    payload keys:
        name            — existing Execution name (to update)
        rollout_plan    — required when creating
        execution_date
        execution_status
        achieved_qty
        achieved_amount
        gps_location
        qc_status
        ciag_status
        revisit_flag
        remarks

    Returns
    -------
    {"name": ..., "status": ...}
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)

    exec_name = payload.get("name")

    if exec_name:
        # Update existing
        doc = frappe.get_doc("Daily Execution", exec_name)
        mode = _user_execution_update_mode(frappe.session.user, doc)
        if mode is None:
            frappe.throw("Not permitted", frappe.PermissionError)
        if mode == "qc_ciag":
            slim = {}
            if "qc_status" in payload:
                slim["qc_status"] = payload.get("qc_status")
            if "ciag_status" in payload:
                slim["ciag_status"] = payload.get("ciag_status")
            payload = slim
    else:
        # Upsert by rollout_plan: only ONE Daily Execution per plan.
        rollout_plan = payload.get("rollout_plan")
        if not rollout_plan:
            frappe.throw("rollout_plan is required when creating a new Daily Execution.")

        existing_exec = frappe.db.get_value(
            "Daily Execution",
            {"rollout_plan": rollout_plan},
            "name",
            order_by="modified desc",
        )
        if existing_exec:
            doc = frappe.get_doc("Daily Execution", existing_exec)
        else:
            rp = frappe.db.get_value(
                "Rollout Plan", rollout_plan, ["po_dispatch", "team", "region_type"], as_dict=True
            )

            doc = frappe.new_doc("Daily Execution")
            doc.rollout_plan = rollout_plan
            doc.system_id = rp.get("po_dispatch") if rp else None
            doc.team = rp.get("team") if rp else None
            pd_name = rp.get("po_dispatch") if rp else None
            im_v = None
            if pd_name:
                im_v = frappe.db.get_value("PO Dispatch", pd_name, "im")
            if im_v and hasattr(doc, "im"):
                doc.im = im_v

            if hasattr(doc, "region_type"):
                doc.region_type = (rp.get("region_type") if rp else None) or None
                if not doc.region_type and pd_name:
                    pd_row = frappe.db.get_value(
                        "PO Dispatch", pd_name, ["region_type", "center_area"], as_dict=True
                    )
                    if pd_row:
                        doc.region_type = pd_row.get("region_type") or region_type_from_center_area(
                            pd_row.get("center_area")
                        )
                if not doc.region_type:
                    doc.region_type = "Standard"

    # Apply updatable fields
    for field in [
        "execution_date",
        "execution_status",
        "tl_status",
        "achieved_qty",
        "achieved_amount",
        "gps_location",
        "photos",
        "qc_status",
        "issue_category",
        "ciag_status",
        "revisit_flag",
        "remarks",
    ]:
        if field in payload and hasattr(doc, field):
            setattr(doc, field, payload[field])

    if hasattr(doc, "execution_status"):
        doc.execution_status = _normalize_execution_status(doc.execution_status)
        if doc.execution_status and doc.execution_status not in _ALLOWED_DAILY_EXECUTION_STATUSES:
            frappe.throw(
                frappe._("Invalid execution_status. Allowed values: {0}").format(
                    ", ".join(sorted(_ALLOWED_DAILY_EXECUTION_STATUSES))
                )
            )
    if hasattr(doc, "tl_status") and "tl_status" in payload:
        doc.tl_status = _normalize_execution_status(doc.tl_status)
        if doc.tl_status and doc.tl_status not in _ALLOWED_DAILY_EXECUTION_STATUSES:
            frappe.throw(
                frappe._("Invalid tl_status. Allowed values: {0}").format(
                    ", ".join(sorted(_ALLOWED_DAILY_EXECUTION_STATUSES))
                )
            )

    # QC / CIAG is only allowed once the work is done — either side counts.
    # Field team's tl_status reaches us before the IM's execution_status, so
    # gating only on execution_status would block the field user from
    # recording QC / CIAG until the IM confirms. tl_status === Completed is
    # the field-side "done" signal.
    if "qc_status" in payload or "ciag_status" in payload:
        _exec_done = (doc.execution_status == "Completed")
        _tl_done = (getattr(doc, "tl_status", None) == "Completed")
        if not (_exec_done or _tl_done):
            frappe.throw(
                "QC and CIAG can only be updated once Execution is marked Completed "
                "(by the field team via TL Status, or confirmed by the IM)."
            )

    # Keep achieved amount server-driven from achieved qty * billing rate.
    if hasattr(doc, "achieved_qty") and hasattr(doc, "achieved_amount"):
        if "achieved_qty" in payload or "achieved_amount" not in payload:
            rate = _get_rollout_billing_rate(doc.rollout_plan)
            doc.achieved_amount = flt(doc.achieved_qty or 0) * flt(rate)

    if payload.get("activity_code") and hasattr(doc, "activity_code"):
        doc.activity_code = payload["activity_code"]
        if hasattr(doc, "activity_cost_sar"):
            acm_cost = frappe.db.get_value(
                "Activity Cost Master", payload["activity_code"], "base_cost_sar"
            )
            doc.activity_cost_sar = flt(acm_cost or 0)

    if doc.is_new():
        doc.insert(ignore_permissions=True)
    else:
        doc.save(ignore_permissions=True)

    _sync_rollout_plan_from_daily_execution(doc.rollout_plan, doc)
    qc = str(getattr(doc, "qc_status", None) or "")

    # Stage 6: QC fail should return work to planning and create a new revisit plan.
    if (doc.execution_status == "Completed") and (qc == "Fail") and doc.rollout_plan:
        reopen_rollout_for_revisit(
            doc.rollout_plan,
            issue_category=(payload.get("issue_category") or "QC Rejection"),
            planning_route="with_issue",
        )

    frappe.db.commit()
    return {"name": doc.name, "status": doc.execution_status}


@frappe.whitelist()
def generate_work_done(execution_name):
    """
    Create a Work Done record from a completed Daily Execution.

    Returns
    -------
    {"name": ...}
    """
    roles = set(frappe.get_roles(frappe.session.user))
    if not (
        "Administrator" in roles
        or "System Manager" in roles
        or "INET Admin" in roles
        or "INET IM" in roles
    ):
        frappe.throw(
            "Only an Implementation Manager or administrator can create Work Done.",
            frappe.PermissionError,
        )

    exec_doc = frappe.get_doc("Daily Execution", execution_name)

    if exec_doc.execution_status != "Completed":
        frappe.throw(
            f"Execution {execution_name} is not Completed (status: {exec_doc.execution_status})."
        )

    qc = str(getattr(exec_doc, "qc_status", None) or "")
    if qc != "Pass":
        frappe.throw("QC must be Pass before creating Work Done.")

    # Check if Work Done already exists
    existing = frappe.db.get_value("Work Done", {"execution": execution_name}, "name")
    if existing:
        frappe.throw(f"Work Done already exists for execution {execution_name}: {existing}")

    # Trace back chain: execution → rollout_plan → po_dispatch
    rp_name = exec_doc.rollout_plan
    rp = frappe.db.get_value(
        "Rollout Plan", rp_name, ["po_dispatch", "visit_multiplier", "team"], as_dict=True
    )
    if not rp:
        frappe.throw(f"Rollout Plan {rp_name} not found.")

    dispatch_name = rp.po_dispatch
    dispatch = frappe.db.get_value(
        "PO Dispatch",
        dispatch_name,
        ["item_code", "center_area", "region_type", "project_code", "customer"],
        as_dict=True,
    )
    if not dispatch:
        frappe.throw(f"PO Dispatch {dispatch_name} not found.")

    item_code = dispatch.item_code
    center_area = dispatch.center_area or ""
    team_id = rp.team

    # Determine billing_rate from Customer Item Master
    is_hard = is_hard_region(dispatch.region_type, center_area)
    billing_rate = 0.0
    cim_filters = {"item_code": item_code, "active_flag": 1}
    if dispatch.customer:
        cim_filters["customer"] = dispatch.customer
    cim = frappe.db.get_value(
        "Customer Item Master",
        cim_filters,
        ["standard_rate_sar", "hard_rate_sar"],
        as_dict=True,
    )
    if not cim:
        cim = frappe.db.get_value(
            "Customer Item Master",
            {"item_code": item_code, "active_flag": 1},
            ["standard_rate_sar", "hard_rate_sar"],
            as_dict=True,
        )
    if cim:
        billing_rate = flt(cim.hard_rate_sar if is_hard else cim.standard_rate_sar)

    executed_qty = flt(exec_doc.achieved_qty or 0)
    revenue = billing_rate * executed_qty

    # Team cost from INET Team
    team_cost = 0.0
    team_type = None
    subcontractor = None
    if team_id:
        team_info = frappe.db.get_value(
            "INET Team",
            team_id,
            ["daily_cost", "team_type", "subcontractor", "daily_cost_applies"],
            as_dict=True,
        )
        if team_info:
            team_type = team_info.team_type
            subcontractor = team_info.subcontractor
            if team_info.daily_cost_applies:
                team_cost = flt(team_info.daily_cost)

    # Subcontract cost (only for SUB teams)
    subcontract_cost = 0.0
    inet_margin_pct = 0.0
    if team_type == "SUB" and subcontractor:
        scc = frappe.db.get_value(
            "Subcontract Cost Master",
            {"subcontractor": subcontractor, "active_flag": 1},
            "expected_cost_sar",
            as_dict=False,
        )
        subcontract_cost = flt(scc or 0)

        # INET margin % from Subcontractor Master
        margin_pct = frappe.db.get_value(
            "Subcontractor Master", subcontractor, "inet_margin_pct"
        )
        inet_margin_pct = flt(margin_pct or 0)

    # Activity cost from Activity Cost Master (linked via execution)
    activity_cost = 0.0
    if getattr(exec_doc, "activity_code", None):
        acm_cost = frappe.db.get_value(
            "Activity Cost Master", exec_doc.activity_code, "base_cost_sar"
        )
        activity_cost = flt(acm_cost or 0)

    visit_multiplier = flt(rp.visit_multiplier or 1.0)
    subcontract_cost = subcontract_cost * visit_multiplier
    total_cost = team_cost + subcontract_cost + activity_cost
    margin = revenue - total_cost

    wd = frappe.new_doc("Work Done")
    wd.execution = execution_name
    wd.system_id = exec_doc.system_id
    wd.region_type = dispatch.get("region_type") or region_type_from_center_area(center_area)
    wd.item_code = item_code
    wd.executed_qty = executed_qty
    wd.billing_rate_sar = billing_rate
    wd.revenue_sar = revenue
    wd.team_cost_sar = team_cost
    wd.subcontract_cost_sar = subcontract_cost
    wd.activity_cost_sar = activity_cost
    wd.total_cost_sar = total_cost
    wd.margin_sar = margin
    wd.inet_margin_pct = inet_margin_pct
    wd.billing_status = "Pending"

    wd.insert(ignore_permissions=True)
    frappe.db.commit()

    return {"name": wd.name}


def _ensure_work_done_for_execution(execution_name):
    if not execution_name:
        return None
    existing = frappe.db.get_value("Work Done", {"execution": execution_name}, "name")
    if existing:
        return existing
    exec_status = frappe.db.get_value("Daily Execution", execution_name, "execution_status")
    if exec_status != "Completed":
        return None
    try:
        result = generate_work_done(execution_name)
        return result.get("name")
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Admin list APIs — Execution monitor / Work Done
# ---------------------------------------------------------------------------


def _batch_inet_team_names(team_ids):
    ids = list({t for t in (team_ids or []) if t})
    if not ids:
        return {}
    out = {}
    for row in frappe.get_all(
        "INET Team",
        filters={"name": ["in", ids]},
        fields=["name", "team_name"],
        limit_page_length=len(ids) + 1,
    ):
        out[row.name] = row.team_name or row.name
    return out


def _batch_im_master_full_names(im_ids):
    ids = list({n for n in (im_ids or []) if n})
    if not ids:
        return {}
    out = {}
    for row in frappe.get_all(
        "IM Master",
        filters={"name": ["in", ids]},
        fields=["name", "full_name"],
        limit_page_length=len(ids) + 1,
    ):
        out[row.name] = row.full_name or row.name
    return out


@frappe.whitelist()
def list_execution_monitor_rows(filters=None, limit=500):
    """
    Rich rows for PM Execution Monitor (Rollout + latest execution + dispatch context).

    filters:
      {
        "status": "Planned|In Execution|Completed|Cancelled|Planning with Issue",
        "visit_type": "...",
        "team": "...",
        "from_date": "YYYY-MM-DD",
        "to_date": "YYYY-MM-DD",
        "search": "free text (matches plan, dispatch, team, IM, …)",
        "project_code": "exact project",
        "site_code": "exact DUID / site code",
        "execution_status": "Daily Execution.execution_status (e.g. Completed for Field QC)",
        "execution_team": "Daily Execution.team when filtering by execution_status",
      }
    """
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
    filters = filters or {}

    rp_fields = [
        "name",
        "po_dispatch",
        "team",
        "plan_date",
        "visit_type",
        "visit_number",
        "target_amount",
        "achieved_amount",
        "completion_pct",
        "plan_status",
        "modified",
    ]
    if frappe.db.has_column("Rollout Plan", "region_type"):
        rp_fields.append("region_type")
    if frappe.db.has_column("Rollout Plan", "im"):
        rp_fields.append("im")
    lim = _portal_row_limit(limit, 500)

    wheres = ["1=1"]
    params = []
    for col, key in (("rp.plan_status", "status"),
                     ("rp.visit_type", "visit_type"),
                     ("rp.team", "team"),
                     ("IFNULL(pd.project_code,'')", "project_code"),
                     ("IFNULL(pd.site_code,'')", "site_code")):
        c, p = _sql_in_or_eq(col, filters.get(key))
        if c:
            wheres.append(c)
            params.extend(p)
    if filters.get("from_date") and filters.get("to_date"):
        wheres.append("rp.plan_date BETWEEN %s AND %s")
        params.extend([filters["from_date"], filters["to_date"]])
    elif filters.get("from_date"):
        wheres.append("rp.plan_date >= %s")
        params.append(filters["from_date"])
    elif filters.get("to_date"):
        wheres.append("rp.plan_date <= %s")
        params.append(filters["to_date"])

    exec_status_vals = _ensure_list(filters.get("execution_status"))
    if exec_status_vals:
        if len(exec_status_vals) == 1:
            ex_match = "de0.execution_status = %s"
            ex_params = [exec_status_vals[0]]
        else:
            ex_match = "de0.execution_status IN ({})".format(", ".join(["%s"] * len(exec_status_vals)))
            ex_params = list(exec_status_vals)
        ex_sql = (
            "EXISTS (SELECT 1 FROM `tabDaily Execution` de0 "
            f"WHERE de0.rollout_plan = rp.name AND {ex_match}"
        )
        params.extend(ex_params)
        if filters.get("execution_team"):
            ex_sql += " AND de0.team = %s"
            params.append(filters["execution_team"])
        ex_sql += " LIMIT 1)"
        wheres.append(ex_sql)

    # Once the line is fully closed (exec Completed + QC Pass + CIAG Approved)
    # AND the IM has created a Work Done record for that execution, hide it
    # from the monitor — operators don't need to act on it anymore.
    ciag_col = "de_wd.ciag_status" if frappe.db.has_column("Daily Execution", "ciag_status") else "''"
    wheres.append(
        "NOT EXISTS ("
        " SELECT 1 FROM `tabDaily Execution` de_wd"
        " INNER JOIN `tabWork Done` wd0 ON wd0.execution = de_wd.name"
        " WHERE de_wd.rollout_plan = rp.name"
        " AND de_wd.execution_status = 'Completed'"
        " AND IFNULL(de_wd.qc_status, '') = 'Pass'"
        f" AND IFNULL({ciag_col}, '') = 'Approved'"
        ")"
    )

    # When a plan was re-planned (a higher-visit_number plan exists for the
    # same dispatch), the older plan represents a past attempt — drop it
    # from the live monitor so operators only see the current visit.
    wheres.append(
        "NOT EXISTS ("
        " SELECT 1 FROM `tabRollout Plan` rp_later"
        " WHERE rp_later.po_dispatch = rp.po_dispatch"
        " AND IFNULL(rp_later.visit_number, 0) > IFNULL(rp.visit_number, 0)"
        ")"
    )

    like_pat = _sql_like_pattern(filters.get("search") or filters.get("q") or "")
    if like_pat:
        concat_parts = [
            "IFNULL(rp.name,'')",
            "IFNULL(rp.team,'')",
            "CAST(rp.plan_date AS CHAR)",
            "IFNULL(rp.visit_type,'')",
            "IFNULL(rp.plan_status,'')",
            "IFNULL(rp.po_dispatch,'')",
            "IFNULL(pd.po_no,'')",
            "IFNULL(pd.item_code,'')",
            "IFNULL(pd.item_description,'')",
            "IFNULL(pd.project_code,'')",
            "IFNULL(pd.site_code,'')",
            "IFNULL(pd.site_name,'')",
            "IFNULL(pd.center_area,'')",
            "IFNULL(pd.region_type,'')",
            "IFNULL(pd.original_dummy_poid,'')",
            "IFNULL(it.team_name,'')",
            "IFNULL(pd.im,'')",
        ]
        if frappe.db.has_column("Rollout Plan", "region_type"):
            concat_parts.append("IFNULL(rp.region_type,'')")
        rp_im_join = ""
        if frappe.db.has_column("Rollout Plan", "im"):
            rp_im_join = "LEFT JOIN `tabIM Master` rim_rp ON rim_rp.name = rp.im"
            concat_parts.append("IFNULL(rim_rp.full_name,'')")
        pd_im_join = ""
        if frappe.db.has_column("PO Dispatch", "im"):
            pd_im_join = "LEFT JOIN `tabIM Master` rim_pd ON rim_pd.name = pd.im"
            concat_parts.append("IFNULL(rim_pd.full_name,'')")
        concat_expr = "CONCAT_WS(' ', " + ", ".join(concat_parts) + ")"
        clause, cparams = _sql_search_clause(concat_expr, filters.get("search") or filters.get("q") or "")
        if clause:
            wheres.append(clause)
            params.extend(cparams)
    else:
        rp_im_join = (
            "LEFT JOIN `tabIM Master` rim_rp ON rim_rp.name = rp.im"
            if frappe.db.has_column("Rollout Plan", "im")
            else ""
        )
        pd_im_join = (
            "LEFT JOIN `tabIM Master` rim_pd ON rim_pd.name = pd.im"
            if frappe.db.has_column("PO Dispatch", "im")
            else ""
        )

    sql_from = (
        "FROM `tabRollout Plan` rp "
        "LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch "
        "LEFT JOIN `tabINET Team` it ON it.name = rp.team "
        f"{rp_im_join} {pd_im_join}"
    )
    id_sql = (
        "SELECT rp.name AS plan_id "
        f"{sql_from} "
        f"WHERE {' AND '.join(wheres)} "
        "ORDER BY rp.modified DESC "
        f"{_sql_limit_suffix(lim)}"
    )
    plan_rows = frappe.db.sql(id_sql, tuple(params), as_dict=True)
    plan_names = [r.plan_id for r in (plan_rows or []) if r.get("plan_id")]
    if not plan_names:
        return []

    order_index = {n: i for i, n in enumerate(plan_names)}
    plans = frappe.get_all(
        "Rollout Plan",
        filters={"name": ["in", plan_names]},
        fields=rp_fields,
        limit_page_length=len(plan_names) + 1,
    )
    plans.sort(key=lambda x: order_index.get(x.name, 10**9))
    dispatch_names = [p.po_dispatch for p in plans if p.po_dispatch]

    if lim:
        exec_cap = min(max(len(plan_names) * 25, lim), 5000)
    else:
        exec_cap = min(max(len(plan_names) * 25, 500), 50000)
    execution_rows = frappe.get_all(
        "Daily Execution",
        filters={"rollout_plan": ["in", plan_names]},
        fields=[
            "name",
            "rollout_plan",
            "execution_date",
            "execution_status",
            "tl_status",
            "issue_category",
            "achieved_qty",
            "achieved_amount",
            "qc_status",
            "ciag_status",
            "photos",
            "gps_location",
            "modified",
        ],
        order_by="modified desc",
        limit_page_length=exec_cap,
    )
    latest_exec_by_plan = {}
    for ex in execution_rows:
        if ex.rollout_plan not in latest_exec_by_plan:
            latest_exec_by_plan[ex.rollout_plan] = ex

    dispatch_map = {}
    if dispatch_names:
        d_fields = [
            "name",
            "po_no",
            "item_code",
            "item_description",
            "project_code",
            "site_name",
            "site_code",
            "customer",
        ]
        if frappe.db.has_column("PO Dispatch", "poid"):
            d_fields.append("poid")
        if frappe.db.has_column("PO Dispatch", "im"):
            d_fields.append("im")
        if frappe.db.has_column("PO Dispatch", "center_area"):
            d_fields.append("center_area")
        if frappe.db.has_column("PO Dispatch", "region_type"):
            d_fields.append("region_type")
        if frappe.db.has_column("PO Dispatch", "original_dummy_poid"):
            d_fields.append("original_dummy_poid")
        for rk in ("general_remark", "manager_remark", "team_lead_remark"):
            if frappe.db.has_column("PO Dispatch", rk):
                d_fields.append(rk)
        drows = frappe.get_all(
            "PO Dispatch",
            filters={"name": ["in", dispatch_names]},
            fields=d_fields,
            limit_page_length=1000,
        )
        dispatch_map = {d.name: d for d in drows}

    team_ids = [p.team for p in plans if p.team]
    im_keys = []
    for p in plans:
        if p.get("im"):
            im_keys.append(p.im)
    for d in dispatch_map.values():
        if d.get("im"):
            im_keys.append(d.im)
    team_name_map = _batch_inet_team_names(team_ids)
    im_name_map = _batch_im_master_full_names(im_keys)

    out = []
    for p in plans:
        ex = latest_exec_by_plan.get(p.name)
        d = dispatch_map.get(p.po_dispatch) if p.po_dispatch else None
        pd_im = d.get("im") if d else None
        rp_im = p.get("im") if frappe.db.has_column("Rollout Plan", "im") else None
        im_key = rp_im or pd_im
        out.append(
            {
                "name": p.name,
                "system_id": p.po_dispatch or p.name,
                "po_dispatch": p.po_dispatch,
                # Prefer the `poid` field when populated; fall back to the
                # dispatch name (which was the POID on pre-refactor docs).
                "poid": ((d.get("poid") if d else None) or (d.name if d else None) or p.po_dispatch),
                "po_no": d.po_no if d else None,
                "item_code": d.item_code if d else None,
                "item_description": d.item_description if d else None,
                "project_code": d.project_code if d else None,
                "site_name": d.site_name if d else None,
                "site_code": d.site_code if d else None,
                "center_area": (d.get("center_area") if d else None),
                "region_type": (p.get("region_type") or (d.get("region_type") if d else None)),
                "team": p.team,
                "team_name": team_name_map.get(p.team) if p.team else None,
                "im": im_key,
                "im_full_name": im_name_map.get(im_key) if im_key else None,
                "plan_date": p.plan_date,
                "visit_type": p.visit_type,
                "visit_number": p.get("visit_number") if p else None,
                "target_amount": flt(p.target_amount or 0),
                "achieved_amount": flt(p.achieved_amount or 0),
                "completion_pct": flt(p.completion_pct or 0),
                "plan_status": p.plan_status,
                "execution_name": ex.name if ex else None,
                "execution_date": ex.execution_date if ex else None,
                "execution_status": ex.execution_status if ex else None,
                "tl_status": ex.get("tl_status") if ex else None,
                "issue_category": ex.get("issue_category") if ex else None,
                "execution_achieved_qty": flt(ex.achieved_qty or 0) if ex else 0,
                "execution_achieved_amount": flt(ex.achieved_amount or 0) if ex else 0,
                "qc_status": ex.qc_status if ex else None,
                "ciag_status": ex.ciag_status if ex else None,
                "photos": ex.photos if ex else None,
                "gps_location": ex.gps_location if ex else None,
                "modified": p.modified,
                "customer": d.get("customer") if d else None,
                "general_remark": d.get("general_remark") if d else None,
                "manager_remark": d.get("manager_remark") if d else None,
                "team_lead_remark": d.get("team_lead_remark") if d else None,
                "original_dummy_poid": (
                    (d.get("original_dummy_poid") or "").strip()
                    if d and frappe.db.has_column("PO Dispatch", "original_dummy_poid")
                    else None
                ),
            }
        )
    if out:
        cim_map = _batch_customer_activity_types(out)
        for r in out:
            r["customer_activity_type"] = cim_map.get((r.get("customer") or "", r.get("item_code") or ""))
    return out


@frappe.whitelist()
def list_work_done_rows(filters=None, limit=500):
    """
    Rich Work Done rows for PM page.
    filters: {
      billing_status, from_date, to_date, team, project_code, site_code, im,
      search (free text across work done, execution, dispatch, team, IM),
    }
    """
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
    filters = filters or {}

    wd_fields = [
        "name",
        "execution",
        "system_id",
        "item_code",
        "executed_qty",
        "billing_rate_sar",
        "revenue_sar",
        "team_cost_sar",
        "subcontract_cost_sar",
        "total_cost_sar",
        "margin_sar",
        "billing_status",
        "modified",
    ]
    if frappe.db.has_column("Work Done", "region_type"):
        wd_fields.append("region_type")
    if frappe.db.has_column("Work Done", "submission_status"):
        wd_fields.append("submission_status")
    lim = _portal_row_limit(limit, 500)

    wheres = ["1=1"]
    params = []
    # Billing status filter uses the same PIC roll-up that the response shows
    # — picking "Closed" returns rows where the PIC has marked
    # "Commercial Invoice Closed" / "PO Line Canceled". Falls back to the
    # legacy wd.billing_status for rows that haven't been touched by PIC yet,
    # OR for sites where the pic_status column hasn't migrated.
    if frappe.db.has_column("PO Dispatch", "pic_status"):
        billing_expr = (
            "CASE "
            "WHEN pd.pic_status IN ('Commercial Invoice Closed','PO Line Canceled') THEN 'Closed' "
            "WHEN pd.pic_status IN ('Commercial Invoice Submitted','Ready for Invoice','Under I-BUY','Under ISDP') THEN 'Invoiced' "
            "WHEN IFNULL(pd.pic_status,'') != '' THEN 'Pending' "
            "ELSE IFNULL(wd.billing_status, 'Pending') END"
        )
    else:
        billing_expr = "IFNULL(wd.billing_status, 'Pending')"
    for col, key in ((billing_expr, "billing_status"),
                     ("IFNULL(rp.team, de.team)", "team"),
                     ("IFNULL(pd.project_code,'')", "project_code"),
                     ("IFNULL(pd.site_code,'')", "site_code")):
        c, p = _sql_in_or_eq(col, filters.get(key))
        if c:
            wheres.append(c)
            params.extend(p)
    im_vals = _ensure_list(filters.get("im"))
    if im_vals:
        rp_im_col = frappe.db.has_column("Rollout Plan", "im")
        ph = ", ".join(["%s"] * len(im_vals))
        if rp_im_col:
            wheres.append(f"(IFNULL(pd.im,'') IN ({ph}) OR IFNULL(rp.im,'') IN ({ph}))")
            params.extend(im_vals + im_vals)
        else:
            wheres.append(f"IFNULL(pd.im,'') IN ({ph})")
            params.extend(im_vals)
    if filters.get("from_date"):
        wheres.append("de.execution_date >= %s")
        params.append(filters["from_date"])
    if filters.get("to_date"):
        wheres.append("de.execution_date <= %s")
        params.append(filters["to_date"])

    like_pat = _sql_like_pattern(filters.get("search") or filters.get("q") or "")
    rp_im_join_wd = (
        "LEFT JOIN `tabIM Master` rim_rp ON rim_rp.name = rp.im"
        if frappe.db.has_column("Rollout Plan", "im")
        else ""
    )
    pd_im_join_wd = (
        "LEFT JOIN `tabIM Master` rim_pd ON rim_pd.name = pd.im"
        if frappe.db.has_column("PO Dispatch", "im")
        else ""
    )
    if like_pat:
        concat_parts = [
            "IFNULL(wd.name,'')",
            "IFNULL(wd.execution,'')",
            "IFNULL(wd.system_id,'')",
            "IFNULL(wd.item_code,'')",
            "CAST(wd.executed_qty AS CHAR)",
            "IFNULL(pd.po_no,'')",
            "IFNULL(pd.project_code,'')",
            "IFNULL(pd.site_code,'')",
            "IFNULL(pd.site_name,'')",
            "IFNULL(pd.item_description,'')",
            "IFNULL(pd.center_area,'')",
            "IFNULL(pd.region_type,'')",
            "IFNULL(pd.original_dummy_poid,'')",
            "IFNULL(it.team_name,'')",
            "IFNULL(pd.im,'')",
        ]
        if frappe.db.has_column("Work Done", "region_type"):
            concat_parts.append("IFNULL(wd.region_type,'')")
        if frappe.db.has_column("Rollout Plan", "im"):
            concat_parts.append("IFNULL(rim_rp.full_name,'')")
        if frappe.db.has_column("PO Dispatch", "im"):
            concat_parts.append("IFNULL(rim_pd.full_name,'')")
        concat_expr = "CONCAT_WS(' ', " + ", ".join(concat_parts) + ")"
        clause, cparams = _sql_search_clause(concat_expr, filters.get("search") or filters.get("q") or "")
        if clause:
            wheres.append(clause)
            params.extend(cparams)

    id_sql = (
        "SELECT wd.name AS wd_name "
        "FROM `tabWork Done` wd "
        "INNER JOIN `tabDaily Execution` de ON de.name = wd.execution "
        "LEFT JOIN `tabRollout Plan` rp ON rp.name = de.rollout_plan "
        "LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch "
        "LEFT JOIN `tabINET Team` it ON it.name = IFNULL(rp.team, de.team) "
        f"{rp_im_join_wd} {pd_im_join_wd} "
        f"WHERE {' AND '.join(wheres)} "
        "ORDER BY wd.modified DESC "
        f"{_sql_limit_suffix(lim)}"
    )
    wd_id_rows = frappe.db.sql(id_sql, tuple(params), as_dict=True)
    wd_ids = [r.wd_name for r in (wd_id_rows or []) if r.get("wd_name")]
    if not wd_ids:
        return []

    order_index = {n: i for i, n in enumerate(wd_ids)}
    rows = frappe.get_all(
        "Work Done",
        filters={"name": ["in", wd_ids]},
        fields=wd_fields,
        limit_page_length=len(wd_ids) + 1,
    )
    rows.sort(key=lambda r: order_index.get(r.name, 10**9))

    rel_cap = min(max(len(rows) * 4, 200), 8000)
    exec_names = [r.execution for r in rows if r.execution]
    ex_map = {}
    if exec_names:
        de_fields = ["name", "rollout_plan", "execution_date", "execution_status", "team", "qc_status", "remarks"]
        if frappe.db.has_column("Daily Execution", "ciag_status"):
            de_fields.append("ciag_status")
        ex_rows = frappe.get_all(
            "Daily Execution",
            filters={"name": ["in", exec_names]},
            fields=de_fields,
            limit_page_length=rel_cap,
        )
        ex_map = {e.name: e for e in ex_rows}

    plan_names = [e.rollout_plan for e in ex_map.values() if e.rollout_plan]
    rp_map = {}
    if plan_names:
        rp_fields_wd = ["name", "po_dispatch", "plan_date", "visit_type", "visit_number", "team", "creation"]
        if frappe.db.has_column("Rollout Plan", "region_type"):
            rp_fields_wd.append("region_type")
        if frappe.db.has_column("Rollout Plan", "im"):
            rp_fields_wd.append("im")
        rp_rows = frappe.get_all(
            "Rollout Plan",
            filters={"name": ["in", plan_names]},
            fields=rp_fields_wd,
            limit_page_length=rel_cap,
        )
        rp_map = {r.name: r for r in rp_rows}

    dispatch_names = [r.po_dispatch for r in rp_map.values() if r.po_dispatch]
    pd_map = {}
    if dispatch_names:
        pd_fields_wd = ["name", "po_no", "project_code", "site_name", "site_code", "item_code", "item_description", "customer", "line_amount", "po_line_no", "dispatch_status"]
        for opt in ("general_remark", "manager_remark", "team_lead_remark"):
            if frappe.db.has_column("PO Dispatch", opt):
                pd_fields_wd.append(opt)
        if frappe.db.has_column("PO Dispatch", "poid"):
            pd_fields_wd.append("poid")
        if frappe.db.has_column("PO Dispatch", "im"):
            pd_fields_wd.append("im")
        if frappe.db.has_column("PO Dispatch", "center_area"):
            pd_fields_wd.append("center_area")
        if frappe.db.has_column("PO Dispatch", "region_type"):
            pd_fields_wd.append("region_type")
        if frappe.db.has_column("PO Dispatch", "original_dummy_poid"):
            pd_fields_wd.append("original_dummy_poid")
        if frappe.db.has_column("PO Dispatch", "pic_status"):
            pd_fields_wd.append("pic_status")
        if frappe.db.has_column("PO Dispatch", "subcon_submission_status"):
            pd_fields_wd.append("subcon_submission_status")
        pd_rows = frappe.get_all(
            "PO Dispatch",
            filters={"name": ["in", dispatch_names]},
            fields=pd_fields_wd,
            limit_page_length=rel_cap,
        )
        pd_map = {p.name: p for p in pd_rows}

    team_prefetch = set()
    for ex in ex_map.values():
        if ex.team:
            team_prefetch.add(ex.team)
    for rp in rp_map.values():
        if rp.team:
            team_prefetch.add(rp.team)
    im_prefetch = set()
    for pd in pd_map.values():
        if pd.get("im"):
            im_prefetch.add(pd.im)
    if frappe.db.has_column("Rollout Plan", "im"):
        for rp in rp_map.values():
            if rp.get("im"):
                im_prefetch.add(rp.im)
    team_name_map_wd = _batch_inet_team_names(list(team_prefetch))
    im_name_map_wd = _batch_im_master_full_names(list(im_prefetch))

    project_codes = {pd.project_code for pd in pd_map.values() if pd.get("project_code")}
    project_name_map = {}
    if project_codes:
        try:
            proj_rows = frappe.get_all(
                "Project Control Center",
                filters={"name": ["in", list(project_codes)]},
                fields=["name", "project_name"],
                limit_page_length=len(project_codes) + 1,
            )
            project_name_map = {p.name: p.project_name for p in proj_rows}
        except Exception:
            project_name_map = {}

    out = []
    for r in rows:
        ex = ex_map.get(r.execution)
        rp = rp_map.get(ex.rollout_plan) if ex and ex.rollout_plan else None
        pd = pd_map.get(rp.po_dispatch) if rp and rp.po_dispatch else None
        team = (rp.team if rp else None) or (ex.team if ex else None)
        project_code = pd.project_code if pd else None
        ex_date = ex.execution_date if ex else None
        rp_im_row = rp.get("im") if rp and frappe.db.has_column("Rollout Plan", "im") else None
        pd_im_row = pd.get("im") if pd else None
        im_row = rp_im_row or pd_im_row
        # Billing status now follows the PIC's invoice status. The PM/IM
        # bucket (Pending / Invoiced / Closed) is rolled up from the 11-value
        # PIC scale so existing widgets/filters keep working. Falls back to
        # the legacy stored billing_status on sites that haven't migrated
        # the new pic_status column yet.
        pic_status_val = pd.get("pic_status") if pd and "pic_status" in pd else None
        billing_override = _billing_status_from_pic(pic_status_val, r.get("billing_status"))
        out.append(
            {
                **r,
                "billing_status": billing_override,
                "pic_status": pic_status_val,
                "po_dispatch": rp.po_dispatch if rp else None,
                # Business POID from dispatch (fall back to dispatch name on legacy docs).
                "poid": ((pd.get("poid") if pd else None) or (pd.name if pd else None) or (rp.po_dispatch if rp else None)),
                "po_no": pd.po_no if pd else None,
                "project_code": project_code,
                "site_name": pd.site_name if pd else None,
                "site_code": pd.site_code if pd else None,
                "center_area": pd.get("center_area") if pd else None,
                "region_type": r.get("region_type")
                or (rp.get("region_type") if rp else None)
                or (pd.get("region_type") if pd else None),
                "team": team,
                "team_name": team_name_map_wd.get(team) if team else None,
                "im": im_row,
                "im_full_name": im_name_map_wd.get(im_row) if im_row else None,
                "item_code": pd.item_code if pd else None,
                "item_description": pd.item_description if pd else None,
                "customer": pd.get("customer") if pd else None,
                "execution_date": ex_date,
                "execution_status": ex.execution_status if ex else None,
                "qc_status": ex.get("qc_status") if ex else None,
                "ciag_status": ex.get("ciag_status") if ex else None,
                "execution_remarks": ex.get("remarks") if ex else None,
                "plan_date": rp.plan_date if rp else None,
                "visit_type": rp.visit_type if rp else None,
                "visit_number": rp.get("visit_number") if rp else None,
                "planning_timestamp": rp.get("creation") if rp else None,
                "dispatch_seq": pd.get("po_line_no") if pd else None,
                "dispatch_status": pd.get("dispatch_status") if pd else None,
                "line_amount": pd.get("line_amount") if pd else None,
                "project_name": project_name_map.get(project_code) if project_code else None,
                "general_remark": pd.get("general_remark") if pd else None,
                "manager_remark": pd.get("manager_remark") if pd else None,
                "team_lead_remark": pd.get("team_lead_remark") if pd else None,
                "original_dummy_poid": (
                    (pd.get("original_dummy_poid") or "").strip()
                    if pd and frappe.db.has_column("PO Dispatch", "original_dummy_poid")
                    else None
                ),
            }
        )
    if out:
        cim_map = _batch_customer_activity_types(out)
        for r in out:
            r["customer_activity_type"] = cim_map.get((r.get("customer") or "", r.get("item_code") or ""))

    # ── Sub-Contract completions: synthesize Work Done rows for PO Dispatches
    # whose subcon_status='Work Done'. Subcon flow lives outside the rollout
    # chain (no Daily Execution / Rollout Plan / Work Done record), so we union
    # them in here as flagged synthetic rows so they show up in the same list.
    subcon_rows = _synthesize_subcon_workdone_rows(filters)
    if subcon_rows:
        out.extend(subcon_rows)
        out.sort(
            key=lambda r: (r.get("execution_date") or "", r.get("modified") or ""),
            reverse=True,
        )
    return out


def _synthesize_subcon_workdone_rows(filters):
    """Build Work-Done-shaped rows from PO Dispatches with subcon_status='Work Done'.

    Honors the same filters as ``list_work_done_rows``:
        billing_status, from_date / to_date, team, project_code, site_code, im, search.
    Subcon dispatches don't carry a billing_status; if the caller filtered to a
    specific real billing_status (Confirmed/Submitted/etc.), they're excluded.
    The ``team`` filter is matched against ``pd.subcon_team`` (not the rollout team).
    """
    f = filters or {}

    # Billing status filter — subcon rows are emitted as `billing_status='Pending'`.
    # Only include them when the filter allows Pending (or is unset).
    bs_vals = _ensure_list(f.get("billing_status"))
    if bs_vals:
        accept = {"", "pending", "all"}
        if not any(str(v).strip().lower() in accept for v in bs_vals):
            return []

    where = ["IFNULL(pd.subcon_status,'') = 'Work Done'"]
    params = []
    for col, key in (
        ("pd.project_code", "project_code"),
        ("pd.site_code", "site_code"),
        ("pd.subcon_team", "team"),
    ):
        c, p = _sql_in_or_eq(col, f.get(key))
        if c:
            where.append(c)
            params.extend(p)
    im_vals = _ensure_list(f.get("im"))
    if im_vals:
        ph = ", ".join(["%s"] * len(im_vals))
        where.append(f"IFNULL(pd.im,'') IN ({ph})")
        params.extend(im_vals)
    if f.get("from_date"):
        where.append("pd.subcon_completed_on >= %s")
        params.append(f["from_date"])
    if f.get("to_date"):
        where.append("pd.subcon_completed_on <= %s")
        params.append(f["to_date"])

    search = f.get("search") or f.get("q") or ""
    if search:
        clause, like_params = _sql_search_clause(
            "CONCAT_WS(' ', "
            "IFNULL(pd.poid,''), IFNULL(pd.po_no,''), IFNULL(pd.item_code,''), "
            "IFNULL(pd.item_description,''), IFNULL(pd.site_name,''), "
            "IFNULL(pd.site_code,''), IFNULL(pd.project_code,''), "
            "IFNULL(pd.center_area,''), IFNULL(pd.region_type,''), "
            "IFNULL(t.team_id,''), IFNULL(t.team_name,''), IFNULL(pd.im,''))",
            search,
        )
        if clause:
            where.append(clause)
            params.extend(like_params)

    sub_sub_col = (
        "pd.subcon_submission_status AS subcon_submission_status"
        if frappe.db.has_column("PO Dispatch", "subcon_submission_status")
        else "'' AS subcon_submission_status"
    )
    sql = (
        "SELECT pd.name AS po_dispatch, pd.poid AS poid, pd.po_no, pd.po_line_no, "
        "pd.project_code, pd.site_code, pd.site_name, pd.center_area, pd.region_type, "
        "pd.item_code, pd.item_description, pd.customer, pd.line_amount, "
        "pd.dispatch_status, pd.im, "
        "pd.subcon_team, pd.subcon_status, pd.subcon_completed_on, pd.subcon_remark, "
        f"{sub_sub_col}, "
        f"{_remark_select()}, "
        "t.team_id AS subcon_team_id, t.team_name AS subcon_team_name, "
        "pd.modified "
        "FROM `tabPO Dispatch` pd "
        "LEFT JOIN `tabINET Team` t ON t.name = pd.subcon_team "
        f"WHERE {' AND '.join(where)} "
        "ORDER BY pd.subcon_completed_on DESC, pd.modified DESC"
    )
    rows = frappe.db.sql(sql, tuple(params), as_dict=True) or []
    if not rows:
        return []

    project_codes = {r.project_code for r in rows if r.get("project_code")}
    project_name_map = {}
    if project_codes:
        try:
            proj_rows = frappe.get_all(
                "Project Control Center",
                filters={"name": ["in", list(project_codes)]},
                fields=["name", "project_name"],
                limit_page_length=len(project_codes) + 1,
            )
            project_name_map = {p.name: p.project_name for p in proj_rows}
        except Exception:
            project_name_map = {}

    im_prefetch = list({r.im for r in rows if r.get("im")})
    im_name_map = _batch_im_master_full_names(im_prefetch) if im_prefetch else {}

    out = []
    cim_map = _batch_customer_activity_types(rows)
    for r in rows:
        out.append({
            # Synthetic name so React/UI keys stay unique. There's no Work Done
            # doctype row backing this entry — anything that tries to mutate by
            # `name` should check `is_subcon` first.
            "name": f"SUBCON-{r.get('po_dispatch')}",
            "execution": None,
            "system_id": r.get("po_dispatch"),
            "poid": r.get("poid") or r.get("po_dispatch"),
            "po_dispatch": r.get("po_dispatch"),
            "po_no": r.get("po_no"),
            "project_code": r.get("project_code"),
            "project_name": project_name_map.get(r.get("project_code")),
            "site_code": r.get("site_code"),
            "site_name": r.get("site_name"),
            "center_area": r.get("center_area"),
            "region_type": r.get("region_type"),
            "team": r.get("subcon_team"),
            "team_name": r.get("subcon_team_name") or r.get("subcon_team_id") or r.get("subcon_team"),
            "im": r.get("im"),
            "im_full_name": im_name_map.get(r.get("im")),
            "item_code": r.get("item_code"),
            "item_description": r.get("item_description"),
            "customer": r.get("customer"),
            "customer_activity_type": cim_map.get((r.get("customer") or "", r.get("item_code") or "")),
            "executed_qty": None,
            "billing_rate_sar": None,
            "revenue_sar": r.get("line_amount"),
            "team_cost_sar": None,
            "subcontract_cost_sar": None,
            "total_cost_sar": None,
            "margin_sar": None,
            "billing_status": "Pending",
            "submission_status": r.get("subcon_submission_status") or "",
            "execution_date": r.get("subcon_completed_on"),
            "execution_status": "Completed",
            "qc_status": None,
            "ciag_status": None,
            "execution_remarks": r.get("subcon_remark"),
            "plan_date": None,
            "visit_type": "Sub-Contract",
            "visit_number": None,
            "planning_timestamp": None,
            "dispatch_seq": r.get("po_line_no"),
            "dispatch_status": r.get("dispatch_status"),
            "line_amount": r.get("line_amount"),
            "general_remark": r.get("general_remark"),
            "manager_remark": r.get("manager_remark"),
            "team_lead_remark": r.get("team_lead_remark"),
            "original_dummy_poid": None,
            "modified": r.get("modified"),
            "is_subcon": 1,
        })
    return out


@frappe.whitelist()
def update_work_done_submission(name, submission_status):
    """IM sets Work Done submission status: 'Ready for Confirmation' or
    'Confirmation Done'."""
    name = (name or "").strip()
    status = (submission_status or "").strip()
    if not name:
        frappe.throw("name is required")
    if status not in ("Ready for Confirmation", "Confirmation Done", ""):
        frappe.throw("Invalid submission_status")
    if not frappe.db.exists("Work Done", name):
        frappe.throw(f"Work Done not found: {name}")
    frappe.db.set_value("Work Done", name, "submission_status", status, update_modified=True)
    frappe.db.commit()
    return {"name": name, "submission_status": status}


@frappe.whitelist()
def update_subcon_submission(po_dispatch, submission_status):
    """Set submission_status on a sub-contracted PO Dispatch (no Work Done row exists).

    Stored on PO Dispatch.subcon_submission_status so synthetic Work Done rows can
    surface it in the regular list. Allowed values: '', 'Ready for Confirmation',
    'Confirmation Done'. Permission: PM, or IM owning the dispatch.
    """
    role = _user_role_class()
    if role not in ("pm", "im"):
        frappe.throw("Not permitted", frappe.PermissionError)
    status = (submission_status or "").strip()
    if status not in ("Ready for Confirmation", "Confirmation Done", ""):
        frappe.throw("Invalid submission_status")
    name = _resolve_dispatch_for_remarks(po_dispatch)
    pd = frappe.db.get_value(
        "PO Dispatch", name,
        ["name", "im", "subcon_status", "poid"],
        as_dict=True,
    ) or {}
    if not pd.get("name"):
        frappe.throw("PO Dispatch not found")
    if (pd.get("subcon_status") or "") != "Work Done":
        frappe.throw("Submission status can only be set on completed sub-contract POIDs.")
    if role == "im":
        _, im_identifiers, _ = resolve_im_for_session()
        if not _can_subcon_dispatch(role, im_identifiers, pd):
            frappe.throw("Not permitted", frappe.PermissionError)
    frappe.db.set_value(
        "PO Dispatch", name,
        "subcon_submission_status", status,
        update_modified=True,
    )
    frappe.db.commit()
    return {"po_dispatch": name, "poid": pd.get("poid") or name, "submission_status": status}


@frappe.whitelist()
def list_issue_risk_rows(im=None, limit=1000, search=None, portal_filters=None):
    """
    Issue & Risk rows from rollout plans that are in issue state or carry an issue category.
    - Admin roles can view all rows (or filter by im argument).
    - IM role is restricted to its own dispatches.
    ``search`` matches rollout / dispatch / project / team / IM fields across the full dataset
    before the row limit is applied.

    ``portal_filters`` (JSON dict) may set exact ``project_code``, ``site_code``, ``team``
    (rollout plan team id) — applied in SQL before the row limit.
    """
    user = frappe.session.user
    roles = set(frappe.get_roles(user))

    is_admin = (
        "Administrator" in roles
        or "System Manager" in roles
        or "INET Admin" in roles
    )
    is_im = "INET IM" in roles
    if not (is_admin or is_im):
        frappe.throw("Not permitted", frappe.PermissionError)

    im_filter_value = (im or "").strip()
    if is_im and not is_admin:
        im_filter_value, _, _ = resolve_im_for_session(im_filter_value)
        if not im_filter_value:
            return []

    rp_im_join_ir = ""
    im_full_sql_ir = "im_pd.full_name AS im_full_name"
    if frappe.db.has_column("Rollout Plan", "im"):
        rp_im_join_ir = "\n        LEFT JOIN `tabIM Master` im_rp ON im_rp.name = rp.im"
        im_full_sql_ir = "COALESCE(im_rp.full_name, im_pd.full_name) AS im_full_name"
    ciag_sel_ir = "de.ciag_status" if frappe.db.has_column("Daily Execution", "ciag_status") else "NULL AS ciag_status"
    lim_ir = _portal_row_limit(limit, 1000)

    # I&R = "open issue per dispatch". The row stays visible across the full
    # lifecycle of the issue: from re-plan → date/team filled in → execution
    # → and only disappears when (a) a Work Done is created for the
    # dispatch, or (b) another re-plan supersedes it with a higher visit.
    #
    # Conditions:
    #   1. Show only the latest plan per dispatch (NOT EXISTS higher visit).
    #   2. The dispatch has an unresolved issue — any of these persistent
    #      markers signals that this plan was created from a re-plan flow:
    #        - visit_type = 'Re-Visit' (always set by reopen_rollout_for_revisit)
    #        - issue_category is set
    #        - plan_status is currently or was historically 'Planning with Issue'
    #   3. No Work Done exists for the dispatch yet.
    wheres = [
        "NOT EXISTS ("
        " SELECT 1 FROM `tabRollout Plan` rp_later"
        " WHERE rp_later.po_dispatch = rp.po_dispatch"
        " AND IFNULL(rp_later.visit_number, 0) > IFNULL(rp.visit_number, 0)"
        ")",
        "("
        " rp.visit_type = 'Re-Visit'"
        " OR IFNULL(rp.issue_category,'') != ''"
        " OR rp.plan_status = 'Planning with Issue'"
        " OR EXISTS ("
        "   SELECT 1 FROM `tabRollout Plan` rp_pwi"
        "   WHERE rp_pwi.po_dispatch = rp.po_dispatch"
        "   AND rp_pwi.plan_status = 'Planning with Issue'"
        " )"
        ")",
        "NOT EXISTS ("
        " SELECT 1 FROM `tabWork Done` wd_ir"
        " INNER JOIN `tabDaily Execution` de_ir ON de_ir.name = wd_ir.execution"
        " INNER JOIN `tabRollout Plan` rp_wd ON rp_wd.name = de_ir.rollout_plan"
        " WHERE rp_wd.po_dispatch = rp.po_dispatch"
        ")",
    ]
    params = []
    if im_filter_value:
        if frappe.db.has_column("Rollout Plan", "im"):
            wheres.append("(IFNULL(pd.im,'') = %s OR IFNULL(rp.im,'') = %s)")
            params.extend([im_filter_value, im_filter_value])
        else:
            wheres.append("IFNULL(pd.im,'') = %s")
            params.append(im_filter_value)

    pf_ir = _portal_filters_dict(portal_filters)
    for col, key in (("IFNULL(pd.project_code,'')", "project_code"),
                     ("IFNULL(pd.site_code,'')", "site_code"),
                     ("IFNULL(rp.team,'')", "team")):
        c, p = _sql_in_or_eq(col, pf_ir.get(key))
        if c:
            wheres.append(c)
            params.extend(p)

    like_pat = _sql_like_pattern(search or "")
    if like_pat:
        concat_parts = [
            "IFNULL(rp.name,'')",
            "IFNULL(rp.po_dispatch,'')",
            "IFNULL(rp.issue_category,'')",
            "IFNULL(rp.team,'')",
            "IFNULL(it.team_name,'')",
            "IFNULL(pd.po_no,'')",
            "IFNULL(pd.project_code,'')",
            "IFNULL(pd.site_code,'')",
            "IFNULL(pd.site_name,'')",
            "IFNULL(pd.im,'')",
            "IFNULL(im_pd.full_name,'')",
        ]
        if frappe.db.has_column("Rollout Plan", "im"):
            concat_parts.append("IFNULL(im_rp.full_name,'')")
        concat_expr = "CONCAT_WS(' ', " + ", ".join(concat_parts) + ")"
        clause, cparams = _sql_search_clause(concat_expr, search or "")
        if clause:
            wheres.append(clause)
            params.extend(cparams)

    id_sql = (
        "SELECT rp.name AS plan_id "
        "FROM `tabRollout Plan` rp "
        "LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch "
        "LEFT JOIN `tabINET Team` it ON it.name = rp.team "
        f"{rp_im_join_ir}"
        "\n        LEFT JOIN `tabIM Master` im_pd ON im_pd.name = pd.im"
        f"\n        WHERE {' AND '.join(wheres)} "
        "ORDER BY rp.modified DESC "
        f"{_sql_limit_suffix(lim_ir)}"
    )
    plan_rows = frappe.db.sql(id_sql, tuple(params), as_dict=True)
    plan_ids = [r.plan_id for r in (plan_rows or []) if r.get("plan_id")]
    if not plan_ids:
        return []

    ph = ", ".join(["%s"] * len(plan_ids))
    rows = frappe.db.sql(
        f"""
        SELECT
            rp.name AS rollout_plan,
            rp.po_dispatch,
            COALESCE(NULLIF(pd.poid, ''), pd.name) AS poid,
            rp.plan_status,
            rp.issue_category,
            rp.issue_remarks,
            rp.visit_number,
            rp.plan_date,
            rp.visit_type,
            rp.team,
            it.team_name AS team_name,
            {im_full_sql_ir},
            rp.modified,
            pd.im,
            pd.po_no,
            pd.project_code,
            pd.site_code,
            pd.site_name,
            pd.item_code,
            pd.item_description,
            pd.line_amount,
            pd.center_area,
            pd.region_type,
            {_remark_select()},
            de.name AS execution_name,
            de.execution_date,
            de.execution_status,
            de.tl_status,
            de.remarks AS execution_remarks,
            de.qc_status,
            {ciag_sel_ir}
        FROM `tabRollout Plan` rp
        LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        LEFT JOIN `tabINET Team` it ON it.name = rp.team
        {rp_im_join_ir}
        LEFT JOIN `tabIM Master` im_pd ON im_pd.name = pd.im
        LEFT JOIN `tabDaily Execution` de ON de.name = (
            SELECT de2.name
            FROM `tabDaily Execution` de2
            INNER JOIN `tabRollout Plan` rp2 ON rp2.name = de2.rollout_plan
            WHERE rp2.po_dispatch = rp.po_dispatch
            AND IFNULL(rp2.visit_number, 0) <= IFNULL(rp.visit_number, 0)
            ORDER BY IFNULL(rp2.visit_number,0) DESC, de2.modified DESC
            LIMIT 1
        )
        WHERE rp.name IN ({ph})
        ORDER BY rp.modified DESC, de.modified DESC
        """,
        tuple(plan_ids),
        as_dict=True,
    )

    out = []
    seen_plans = set()
    for r in rows or []:
        rp_name = r.get("rollout_plan")
        if not rp_name or rp_name in seen_plans:
            continue
        seen_plans.add(rp_name)
        out.append(
            {
                "rollout_plan": rp_name,
                "po_dispatch": r.get("po_dispatch"),
                "poid": r.get("poid"),
                "plan_status": r.get("plan_status"),
                "issue_category": r.get("issue_category"),
                "issue_remarks": r.get("issue_remarks"),
                "visit_number": r.get("visit_number"),
                "plan_date": r.get("plan_date"),
                "visit_type": r.get("visit_type"),
                "team": r.get("team"),
                "team_name": r.get("team_name"),
                "im": r.get("im"),
                "im_full_name": r.get("im_full_name"),
                "po_no": r.get("po_no"),
                "project_code": r.get("project_code"),
                "site_code": r.get("site_code"),
                "site_name": r.get("site_name"),
                "item_code": r.get("item_code"),
                "item_description": r.get("item_description"),
                "line_amount": r.get("line_amount"),
                "center_area": r.get("center_area"),
                "region_type": r.get("region_type"),
                "general_remark": r.get("general_remark"),
                "manager_remark": r.get("manager_remark"),
                "team_lead_remark": r.get("team_lead_remark"),
                "execution_name": r.get("execution_name"),
                "execution_date": r.get("execution_date"),
                "execution_status": r.get("execution_status"),
                "tl_status": r.get("tl_status"),
                "execution_remarks": r.get("execution_remarks"),
                "qc_status": r.get("qc_status"),
                "ciag_status": r.get("ciag_status"),
                "modified": r.get("modified"),
            }
        )
    return out


# ---------------------------------------------------------------------------
# Task 13 — Dashboard Aggregation
# ---------------------------------------------------------------------------


def _month_bounds():
    """Return (first_day_str, last_day_str, today_str) for current month."""
    today = getdate(nowdate())
    first = get_first_day(today)
    last = get_last_day(today)
    return str(first), str(last), str(today)


def _days_in_month(today=None):
    today = getdate(today or nowdate())
    return calendar.monthrange(today.year, today.month)[1]


@frappe.whitelist()
def get_command_dashboard(from_date=None, to_date=None, etag=None):
    """
    Return ALL data for the admin Command Dashboard in a single API call.

    Optional ``from_date`` / ``to_date`` (ISO YYYY-MM-DD) override the default
    month window for date-scoped metrics (closed activities, revenue, cost,
    team activity, etc.). Strategic targets still use the running-month
    baseline so the pro-rated "target today" calc stays meaningful.

    If the caller passes the ``etag`` it last received, and nothing in the
    underlying tables has changed since, returns a small ``{"unchanged":
    True, "etag": ...}`` payload so polling clients save a full re-render.
    """
    current_etag = _dashboard_etag("cmd", from_date, to_date)
    if etag and etag == current_etag:
        return {"unchanged": True, "etag": current_etag, "last_updated": _iso_now()}
    today_str = nowdate()
    today = getdate(today_str)
    first_day, last_day, _ = _month_bounds()
    if from_date:
        try: first_day = getdate(from_date)
        except Exception: pass
    if to_date:
        try: last_day = getdate(to_date)
        except Exception: pass
    days_in_month = _days_in_month(today)
    day_of_month = today.day

    # ---- Operational KPIs --------------------------------------------------
    # Open lines = PO Intake Lines whose per-line status is NOT terminal.
    # The line-wise status is the source of truth — the parent PO Intake's
    # status field is just a roll-up and isn't authoritative for KPIs.
    open_po_result = frappe.db.sql(
        """
        SELECT COALESCE(SUM(line_amount), 0) AS total_value, COUNT(*) AS line_count
        FROM `tabPO Intake Line`
        WHERE IFNULL(po_line_status, 'New') NOT IN ('Closed', 'Cancelled')
        """,
        as_dict=True,
    )
    _open_po_row = open_po_result[0] if open_po_result else None
    total_open_po_line_value = flt(_open_po_row.total_value if _open_po_row else 0)
    total_open_po_lines = cint(_open_po_row.line_count if _open_po_row else 0)
    # Legacy key: same as total open line amount (SAR)
    total_open_po = total_open_po_line_value

    # Teams with at least one execution today
    active_team_rows = frappe.db.sql(
        """
        SELECT DISTINCT team FROM `tabDaily Execution`
        WHERE execution_date = %s
        AND execution_status NOT IN ('Cancelled')
        """,
        (today_str,),
        as_dict=True,
    )
    active_teams_count = len(active_team_rows)
    active_team_ids = {r.team for r in active_team_rows}

    total_teams = frappe.db.count("INET Team", {"status": "Active"})
    idle_teams_count = max(0, total_teams - active_teams_count)

    planned_activities = frappe.db.count("Rollout Plan", {"plan_status": "Planned"})

    closed_activities = frappe.db.sql(
        """
        SELECT COUNT(*) AS cnt FROM `tabDaily Execution`
        WHERE execution_status = 'Completed'
        AND execution_date BETWEEN %s AND %s
        """,
        (first_day, last_day),
        as_dict=True,
    )[0].cnt or 0

    revisits = frappe.db.sql(
        """
        SELECT COUNT(*) AS cnt FROM `tabRollout Plan`
        WHERE visit_type = 'Re-Visit'
        AND plan_date BETWEEN %s AND %s
        """,
        (first_day, last_day),
        as_dict=True,
    )[0].cnt or 0

    # ---- INET KPIs ---------------------------------------------------------
    inet_teams = frappe.get_all(
        "INET Team",
        filters={"team_type": "INET", "status": "Active"},
        fields=["name", "team_id", "daily_cost", "im"],
    )
    active_inet_teams = len(inet_teams)

    inet_monthly_cost = sum(flt(t.daily_cost) * 26 for t in inet_teams)

    # Monthly target from Project Control Center
    pcc_targets = frappe.db.sql(
        "SELECT COALESCE(SUM(monthly_target), 0) AS total FROM `tabProject Control Center` WHERE active_flag = 'Yes'",
        as_dict=True,
    )
    inet_monthly_target = flt(pcc_targets[0].total if pcc_targets else 0)

    inet_target_today = (
        (inet_monthly_target * day_of_month) / days_in_month if days_in_month else 0.0
    )

    inet_achieved_rows = frappe.db.sql(
        """
        SELECT COALESCE(SUM(wd.revenue_sar), 0) AS total
        FROM `tabWork Done` wd
        JOIN `tabDaily Execution` exe ON exe.name = wd.execution
        JOIN `tabINET Team` it ON it.name = exe.team
        WHERE it.team_type = 'INET'
        AND exe.execution_date BETWEEN %s AND %s
        """,
        (first_day, last_day),
        as_dict=True,
    )
    inet_achieved = flt(inet_achieved_rows[0].total if inet_achieved_rows else 0)
    inet_gap_today = inet_target_today - inet_achieved

    # ---- Subcontractor KPIs ------------------------------------------------
    sub_teams = frappe.get_all(
        "INET Team",
        filters={"team_type": "SUB", "status": "Active"},
        fields=["name", "team_id"],
    )
    active_sub_teams = len(sub_teams)
    # Real MTD target: sum of rollout plan target_amount for active SUB teams this month
    sub_target = 0.0
    if sub_teams:
        sub_names = [t.name for t in sub_teams]
        sub_ph = ", ".join(["%s"] * len(sub_names))
        sub_tgt_rows = frappe.db.sql(
            f"""
            SELECT COALESCE(SUM(rp.target_amount), 0) AS total
            FROM `tabRollout Plan` rp
            WHERE rp.team IN ({sub_ph})
            AND rp.plan_date BETWEEN %s AND %s
            """,
            tuple(sub_names) + (first_day, last_day),
            as_dict=True,
        )
        sub_target = flt(sub_tgt_rows[0].total if sub_tgt_rows else 0)

    sub_revenue_rows = frappe.db.sql(
        """
        SELECT COALESCE(SUM(wd.revenue_sar), 0) AS rev,
               COALESCE(SUM(wd.total_cost_sar), 0) AS cost,
               COALESCE(AVG(wd.inet_margin_pct), 0) AS avg_margin
        FROM `tabWork Done` wd
        JOIN `tabDaily Execution` exe ON exe.name = wd.execution
        JOIN `tabINET Team` it ON it.name = exe.team
        WHERE it.team_type = 'SUB'
        AND exe.execution_date BETWEEN %s AND %s
        """,
        (first_day, last_day),
        as_dict=True,
    )
    sub_revenue = flt(sub_revenue_rows[0].rev if sub_revenue_rows else 0)
    sub_expense = flt(sub_revenue_rows[0].cost if sub_revenue_rows else 0)
    avg_margin_pct = flt(sub_revenue_rows[0].avg_margin if sub_revenue_rows else 0)
    inet_margin_sub = sub_revenue * (avg_margin_pct / 100.0)
    sub_gap = sub_target - sub_revenue

    # ---- Company-level KPIs ------------------------------------------------
    company_target = inet_monthly_target + sub_target
    total_achieved = inet_achieved + sub_revenue
    total_cost = inet_monthly_cost + sub_expense
    company_gap = company_target - total_achieved
    profit_loss = total_achieved - total_cost
    coverage_pct = (total_achieved / company_target * 100.0) if company_target else 0.0

    # ---- Top 5 teams by revenue this month ---------------------------------
    top_teams = frappe.db.sql(
        """
        SELECT exe.team AS team, COALESCE(SUM(wd.revenue_sar), 0) AS revenue
        FROM `tabWork Done` wd
        JOIN `tabDaily Execution` exe ON exe.name = wd.execution
        WHERE exe.execution_date BETWEEN %s AND %s
        AND exe.team IS NOT NULL AND exe.team != ''
        GROUP BY exe.team
        ORDER BY revenue DESC
        LIMIT 5
        """,
        (first_day, last_day),
        as_dict=True,
    )
    team_rollout_targets = frappe.db.sql(
        """
        SELECT team AS team, COALESCE(SUM(target_amount), 0) AS target
        FROM `tabRollout Plan`
        WHERE plan_date BETWEEN %s AND %s
        AND team IS NOT NULL AND team != ''
        GROUP BY team
        """,
        (first_day, last_day),
        as_dict=True,
    )
    target_by_team = {r.team: flt(r.target) for r in (team_rollout_targets or [])}
    for _row in top_teams:
        tid = _row.get("team")
        _row["team_name"] = frappe.db.get_value("INET Team", tid, "team_name") or tid or "—"
        _row["achieved"] = flt(_row.get("revenue", 0))
        _row["target"] = target_by_team.get(tid, 0.0)

    # ---- IM performance ----------------------------------------------------
    im_perf = frappe.db.sql(
        """
        SELECT pd.im AS im,
               COUNT(DISTINCT exe.team) AS team_count,
               COALESCE(SUM(wd.revenue_sar), 0) AS revenue,
               COALESCE(SUM(wd.total_cost_sar), 0) AS cost
        FROM `tabWork Done` wd
        JOIN `tabDaily Execution` exe ON exe.name = wd.execution
        JOIN `tabRollout Plan` rp ON rp.name = exe.rollout_plan
        JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        WHERE exe.execution_date BETWEEN %s AND %s
        AND pd.im IS NOT NULL AND pd.im != ''
        GROUP BY pd.im
        """,
        (first_day, last_day),
        as_dict=True,
    )
    for row in im_perf:
        row["profit"] = flt(row.revenue) - flt(row.cost)
        row["teams"] = cint(row.get("team_count") or 0)
        row["team_cost"] = flt(row.get("cost") or 0)

    # ---- Team status summary -----------------------------------------------
    in_progress_count = frappe.db.sql(
        """
        SELECT COUNT(DISTINCT team) AS cnt FROM `tabDaily Execution`
        WHERE execution_date = %s AND execution_status = 'In Progress'
        """,
        (today_str,),
        as_dict=True,
    )[0].cnt or 0

    planned_today = frappe.db.sql(
        """
        SELECT COUNT(DISTINCT team) AS cnt FROM `tabRollout Plan`
        WHERE plan_date = %s AND plan_status = 'Planned'
        """,
        (today_str,),
        as_dict=True,
    )[0].cnt or 0

    team_status = {
        "active": active_teams_count,
        "idle": idle_teams_count,
        "planned": planned_today,
        "in_progress": in_progress_count,
    }

    # ---- Watchlist (action items) — shape matches CommandDashboard.jsx ------
    watchlist = []
    issue_rows = frappe.db.sql(
        """
        SELECT COALESCE(rp.issue_category, 'Uncategorized') AS issue_category, COUNT(*) AS cnt
        FROM `tabDaily Execution` de
        JOIN `tabRollout Plan` rp ON rp.name = de.rollout_plan
        WHERE de.qc_status = 'Fail'
        AND de.execution_date BETWEEN %s AND %s
        GROUP BY COALESCE(rp.issue_category, 'Uncategorized')
        ORDER BY cnt DESC
        LIMIT 5
        """,
        (first_day, last_day),
        as_dict=True,
    )
    for row in issue_rows:
        cnt = cint(row.cnt)
        watchlist.append(
            {
                "indicator": f"QC open issues — {row.issue_category}",
                "current": cnt,
                "target": None,
                "status": "Recover",
            }
        )

    if idle_teams_count > 0:
        watchlist.append(
            {
                "indicator": "Idle teams today",
                "current": idle_teams_count,
                "target": None,
                "status": "Behind",
            }
        )
    if inet_gap_today > 0:
        watchlist.append(
            {
                "indicator": "INET revenue gap (month-to-date vs prorated target)",
                "current": inet_gap_today,
                "target": None,
                "status": "Recover",
            }
        )
    if sub_gap > 0:
        watchlist.append(
            {
                "indicator": "Subcontractor revenue gap",
                "current": sub_gap,
                "target": None,
                "status": "Recover",
            }
        )
    if not watchlist:
        watchlist.append(
            {
                "indicator": "Operational health",
                "current": 0,
                "target": None,
                "status": "Optimized",
            }
        )

    return {
        "operational": {
            "total_open_po_lines": total_open_po_lines,
            "total_open_po_line_value": total_open_po_line_value,
            "total_open_po": total_open_po,
            "active_teams": active_teams_count,
            "idle_teams": idle_teams_count,
            "planned_activities": planned_activities,
            "closed_activities": closed_activities,
            "revisits": revisits,
        },
        "inet": {
            "active_inet_teams": active_inet_teams,
            "inet_monthly_cost": inet_monthly_cost,
            "inet_monthly_target": inet_monthly_target,
            "inet_target_today": inet_target_today,
            "inet_achieved": inet_achieved,
            "inet_gap_today": inet_gap_today,
        },
        "subcon": {
            "active_sub_teams": active_sub_teams,
            "sub_target": sub_target,
            "sub_revenue": sub_revenue,
            "sub_expense": sub_expense,
            "inet_margin_sub": inet_margin_sub,
            "sub_gap": sub_gap,
        },
        "company": {
            "company_target": company_target,
            "total_achieved": total_achieved,
            "company_gap": company_gap,
            "total_cost": total_cost,
            "profit_loss": profit_loss,
            "coverage_pct": coverage_pct,
        },
        "top_teams": top_teams,
        "im_performance": im_perf,
        "team_status": team_status,
        "watchlist": watchlist,
        "last_updated": _iso_now(),
        "etag": current_etag,
    }


def resolve_im_for_session(im=None):
    """
    Resolve IM Master document name + all identifier strings usable in PO Dispatch.im
    and similar Link fields. Shared by IM dashboard, pipeline lists, and overview APIs.
    """
    session_user = frappe.session.user
    user_full_name = frappe.db.get_value("User", session_user, "full_name") or ""
    im_rec = None

    def _find_im_rec(filters):
        if not frappe.db.exists("DocType", "IM Master"):
            return None
        try:
            rows = frappe.get_all(
                "IM Master", filters=filters,
                fields=["name", "full_name", "email"], limit=1,
                ignore_permissions=True,
            )
            return rows[0] if rows else None
        except Exception:
            return None

    if im and frappe.db.exists("IM Master", im):
        im_rec = _find_im_rec({"name": im})
    if not im_rec and im:
        im_rec = _find_im_rec({"full_name": im})
    if not im_rec and im:
        im_rec = _find_im_rec({"email": im})
    if not im_rec:
        im_rec = _find_im_rec({"user": session_user})
    if not im_rec and user_full_name:
        im_rec = _find_im_rec({"full_name": user_full_name})

    if im_rec:
        im_identifiers = list({
            v for v in [im_rec.name, im_rec.full_name, im_rec.email, im, user_full_name]
            if v
        })
        im_resolved = im_rec.name
    else:
        im_resolved = im or user_full_name or None
        im_identifiers = list({v for v in [im, user_full_name] if v})

    if not im_identifiers:
        return None, [], {}
    return im_resolved, im_identifiers, {"session_user": session_user, "user_full_name": user_full_name}


def im_action_counts(im_identifiers):
    """Counts for IM dashboard action strip (PM→IM→execution workflow)."""
    if not im_identifiers:
        return {"pending_plan_dispatches": 0, "qc_fail_needs_action": 0, "planned_ready_execution": 0}
    ph = ", ".join(["%s"] * len(im_identifiers))
    params = tuple(im_identifiers)
    pending = frappe.db.sql(
        f"""
        SELECT COUNT(*) AS c FROM `tabPO Dispatch`
        WHERE im IN ({ph}) AND dispatch_status = 'Dispatched'
        """,
        params,
        as_dict=True,
    )[0].c
    qc_open = frappe.db.sql(
        f"""
        SELECT COUNT(*) AS c FROM `tabDaily Execution` de
        INNER JOIN `tabRollout Plan` rp ON rp.name = de.rollout_plan
        INNER JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        WHERE pd.im IN ({ph}) AND de.qc_status = 'Fail'
        AND de.execution_status NOT IN ('Cancelled')
        """,
        params,
        as_dict=True,
    )[0].c
    planned_exec = frappe.db.sql(
        f"""
        SELECT COUNT(*) AS c FROM `tabRollout Plan` rp
        INNER JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        WHERE pd.im IN ({ph}) AND rp.plan_status = 'Planned'
        """,
        params,
        as_dict=True,
    )[0].c
    return {
        "pending_plan_dispatches": cint(pending),
        "qc_fail_needs_action": cint(qc_open),
        "planned_ready_execution": cint(planned_exec),
    }


@frappe.whitelist()
def list_im_rollout_plans(im=None, plan_status=None, limit=500, portal_filters=None):
    """Rollout plans for this IM (join PO Dispatch — works before im backfill on Rollout Plan)."""
    im_resolved, im_identifiers, _ = resolve_im_for_session(im)
    if not im_identifiers:
        return []
    pf = _portal_filters_dict(portal_filters)
    ph = ", ".join(["%s"] * len(im_identifiers))
    params = list(im_identifiers)
    status_clause = ""
    status_vals = _ensure_list(plan_status)
    if status_vals:
        if len(status_vals) == 1:
            status_clause = " AND rp.plan_status = %s"
            params.append(status_vals[0])
        else:
            status_clause = " AND rp.plan_status IN ({})".format(", ".join(["%s"] * len(status_vals)))
            params.extend(status_vals)
    portal_clause = ""
    for col, key in (("rp.visit_type", "visit_type"),
                     ("IFNULL(pd.project_code,'')", "project_code"),
                     ("IFNULL(pd.site_code,'')", "site_code"),
                     ("IFNULL(rp.team,'')", "team")):
        c, p = _sql_in_or_eq(col, pf.get(key))
        if c:
            portal_clause += f" AND {c}"
            params.extend(p)
    if pf.get("from_date") and pf.get("to_date"):
        portal_clause += " AND rp.plan_date BETWEEN %s AND %s"
        params.extend([pf["from_date"], pf["to_date"]])
    elif pf.get("from_date"):
        portal_clause += " AND rp.plan_date >= %s"
        params.append(pf["from_date"])
    elif pf.get("to_date"):
        portal_clause += " AND rp.plan_date <= %s"
        params.append(pf["to_date"])
    like_pat = _sql_like_pattern(pf.get("search") or pf.get("q") or "")
    if like_pat:
        concat_parts = [
            "IFNULL(rp.name,'')",
            "IFNULL(rp.po_dispatch,'')",
            "CAST(rp.plan_date AS CHAR)",
            "IFNULL(rp.visit_type,'')",
            "IFNULL(rp.team,'')",
            "IFNULL(it.team_name,'')",
            "IFNULL(pd.po_no,'')",
            "IFNULL(pd.project_code,'')",
            "IFNULL(pd.site_code,'')",
            "IFNULL(pd.item_code,'')",
            "IFNULL(pd.im,'')",
            "IFNULL(im_pd.full_name,'')",
        ]
        if frappe.db.has_column("Rollout Plan", "im"):
            concat_parts.append("IFNULL(im_rp.full_name,'')")
        concat_expr = "CONCAT_WS(' ', " + ", ".join(concat_parts) + ")"
        clause, cparams = _sql_search_clause(concat_expr, pf.get("search") or pf.get("q") or "")
        if clause:
            portal_clause += f" AND {clause}"
            params.extend(cparams)
    im_plan_extras = []
    if frappe.db.has_column("Rollout Plan", "region_type"):
        im_plan_extras.append("rp.region_type AS region_type")
    else:
        im_plan_extras.append("NULL AS region_type")
    if frappe.db.has_column("PO Dispatch", "center_area"):
        im_plan_extras.append("pd.center_area AS center_area")
    else:
        im_plan_extras.append("NULL AS center_area")
    im_plan_extra_sql = ", " + ", ".join(im_plan_extras)
    rp_im_join = ""
    im_full_sql = "im_pd.full_name AS im_full_name"
    if frappe.db.has_column("Rollout Plan", "im"):
        rp_im_join = "LEFT JOIN `tabIM Master` im_rp ON im_rp.name = rp.im"
        im_full_sql = "COALESCE(im_rp.full_name, im_pd.full_name) AS im_full_name"
    lim_rp = _portal_row_limit(limit, 500)
    rows = frappe.db.sql(
        f"""
        SELECT rp.name, rp.po_dispatch AS system_id, rp.po_dispatch,
               COALESCE(NULLIF(pd.poid, ''), pd.name) AS poid,
               rp.team, rp.plan_date, rp.visit_type,
               rp.visit_number, rp.visit_multiplier, rp.target_amount, rp.achieved_amount,
               rp.completion_pct, rp.plan_status,
               pd.qty AS qty,
               pd.im AS dispatch_im, pd.site_code, pd.po_no, pd.project_code, pd.item_code,
               pd.customer AS customer, pd.item_description,
               {_remark_select()},
               it.team_name AS team_name,
               {im_full_sql}
               {im_plan_extra_sql}
        FROM `tabRollout Plan` rp
        INNER JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        LEFT JOIN `tabINET Team` it ON it.name = rp.team
        {rp_im_join}
        LEFT JOIN `tabIM Master` im_pd ON im_pd.name = pd.im
        WHERE pd.im IN ({ph}){status_clause}{portal_clause}
        ORDER BY rp.plan_date DESC, rp.modified DESC
        {_sql_limit_suffix(lim_rp)}
        """,
        tuple(params),
        as_dict=True,
    )
    if rows:
        cim_map = _batch_customer_activity_types(rows)
        for r in rows:
            r["customer_activity_type"] = cim_map.get((r.get("customer") or "", r.get("item_code") or ""))
    return rows or []


@frappe.whitelist()
def list_im_daily_executions(im=None, execution_status=None, limit=500, portal_filters=None):
    """Daily executions for this IM's dispatches."""
    im_resolved, im_identifiers, _ = resolve_im_for_session(im)
    if not im_identifiers:
        return []
    pf = _portal_filters_dict(portal_filters)
    ph = ", ".join(["%s"] * len(im_identifiers))
    params = list(im_identifiers)
    status_clause = ""
    status_vals = _ensure_list(execution_status)
    if status_vals:
        norm_vals = [_normalize_execution_status(s) for s in status_vals]
        if len(norm_vals) == 1:
            status_clause = " AND de.execution_status = %s"
            params.append(norm_vals[0])
        else:
            status_clause = " AND de.execution_status IN ({})".format(", ".join(["%s"] * len(norm_vals)))
            params.extend(norm_vals)
    portal_clause = ""
    for col, key in (("IFNULL(de.qc_status,'')", "qc_status"),
                     ("IFNULL(pd.project_code,'')", "project_code"),
                     ("IFNULL(pd.site_code,'')", "site_code"),
                     ("IFNULL(de.team,'')", "team")):
        c, p = _sql_in_or_eq(col, pf.get(key))
        if c:
            portal_clause += f" AND {c}"
            params.extend(p)
    if frappe.db.has_column("Daily Execution", "ciag_status"):
        c, p = _sql_in_or_eq("IFNULL(de.ciag_status,'')", pf.get("ciag_status"))
        if c:
            portal_clause += f" AND {c}"
            params.extend(p)
    if pf.get("from_date") and pf.get("to_date"):
        portal_clause += " AND de.execution_date BETWEEN %s AND %s"
        params.extend([pf["from_date"], pf["to_date"]])
    elif pf.get("from_date"):
        portal_clause += " AND de.execution_date >= %s"
        params.append(pf["from_date"])
    elif pf.get("to_date"):
        portal_clause += " AND de.execution_date <= %s"
        params.append(pf["to_date"])

    # Once an execution is fully closed (Completed + Pass + Approved) AND has
    # been converted to Work Done, drop it — the IM has nothing left to do.
    ciag_col_hide = "de.ciag_status" if frappe.db.has_column("Daily Execution", "ciag_status") else "''"
    portal_clause += (
        " AND NOT ("
        " de.execution_status = 'Completed'"
        " AND IFNULL(de.qc_status,'') = 'Pass'"
        f" AND IFNULL({ciag_col_hide},'') = 'Approved'"
        " AND EXISTS (SELECT 1 FROM `tabWork Done` wd0 WHERE wd0.execution = de.name)"
        ")"
    )

    like_pat = _sql_like_pattern(pf.get("search") or pf.get("q") or "")
    if like_pat:
        concat_parts = [
            "IFNULL(de.name,'')",
            "IFNULL(de.rollout_plan,'')",
            "CAST(de.execution_date AS CHAR)",
            "IFNULL(de.team,'')",
            "IFNULL(it.team_name,'')",
            "IFNULL(de.execution_status,'')",
            "IFNULL(de.qc_status,'')",
            "IFNULL(pd.po_no,'')",
            "IFNULL(pd.project_code,'')",
            "IFNULL(pd.site_code,'')",
            "IFNULL(pd.site_name,'')",
            "IFNULL(pd.item_code,'')",
            "IFNULL(pd.item_description,'')",
            "IFNULL(pd.im,'')",
            "IFNULL(im_pd.full_name,'')",
        ]
        if frappe.db.has_column("Daily Execution", "ciag_status"):
            concat_parts.append("IFNULL(de.ciag_status,'')")
        if frappe.db.has_column("Rollout Plan", "im"):
            concat_parts.append("IFNULL(im_rp.full_name,'')")
        concat_expr = "CONCAT_WS(' ', " + ", ".join(concat_parts) + ")"
        clause, cparams = _sql_search_clause(concat_expr, pf.get("search") or pf.get("q") or "")
        if clause:
            portal_clause += f" AND {clause}"
            params.extend(cparams)
    ciag_sel = "de.ciag_status" if frappe.db.has_column("Daily Execution", "ciag_status") else "NULL"
    im_ex_extras = []
    if frappe.db.has_column("Daily Execution", "region_type"):
        im_ex_extras.append("de.region_type AS region_type")
    else:
        im_ex_extras.append("NULL AS region_type")
    if frappe.db.has_column("PO Dispatch", "center_area"):
        im_ex_extras.append("pd.center_area AS center_area")
    else:
        im_ex_extras.append("NULL AS center_area")
    if frappe.db.has_column("PO Dispatch", "original_dummy_poid"):
        im_ex_extras.append("pd.original_dummy_poid AS original_dummy_poid")
    else:
        im_ex_extras.append("NULL AS original_dummy_poid")
    im_ex_extra_sql = ", " + ", ".join(im_ex_extras)
    rp_im_join_ex = ""
    im_full_sql_ex = "im_pd.full_name AS im_full_name"
    if frappe.db.has_column("Rollout Plan", "im"):
        rp_im_join_ex = "LEFT JOIN `tabIM Master` im_rp ON im_rp.name = rp.im"
        im_full_sql_ex = "COALESCE(im_rp.full_name, im_pd.full_name) AS im_full_name"
    lim_de = _portal_row_limit(limit, 500)
    rows = frappe.db.sql(
        f"""
        SELECT de.name, rp.po_dispatch AS system_id, de.rollout_plan,
               COALESCE(NULLIF(pd.poid, ''), pd.name) AS poid,
               rp.visit_number, rp.visit_type,
               de.team, de.execution_date,
               de.execution_status, de.tl_status, de.issue_category,
               de.achieved_qty, de.achieved_amount, de.gps_location,
               de.qc_status, {ciag_sel} AS ciag_status, de.revisit_flag, de.photos,
               pd.im AS dispatch_im, pd.site_code, pd.site_name, pd.po_no, pd.project_code, pd.item_code, pd.item_description,
               pd.customer AS customer,
               {_remark_select()},
               (SELECT wd.name FROM `tabWork Done` wd WHERE wd.execution = de.name LIMIT 1) AS work_done,
               it.team_name AS team_name,
               {im_full_sql_ex}
               {im_ex_extra_sql}
        FROM `tabDaily Execution` de
        INNER JOIN `tabRollout Plan` rp ON rp.name = de.rollout_plan
        INNER JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        LEFT JOIN `tabINET Team` it ON it.name = de.team
        {rp_im_join_ex}
        LEFT JOIN `tabIM Master` im_pd ON im_pd.name = pd.im
        WHERE pd.im IN ({ph}){status_clause}{portal_clause}
        AND NOT EXISTS (
            SELECT 1 FROM `tabRollout Plan` rp_later
            WHERE rp_later.po_dispatch = rp.po_dispatch
            AND IFNULL(rp_later.visit_number, 0) > IFNULL(rp.visit_number, 0)
        )
        ORDER BY de.execution_date DESC, de.modified DESC
        {_sql_limit_suffix(lim_de)}
        """,
        tuple(params),
        as_dict=True,
    )
    if rows:
        cim_map = _batch_customer_activity_types(rows)
        for r in rows:
            r["customer_activity_type"] = cim_map.get((r.get("customer") or "", r.get("item_code") or ""))
    return rows or []


@frappe.whitelist()
def get_duid_overview(duid=None, po_no=None):
    """
    DUID-wise (site_code) rollout view: PO line, plans, executions.
    Optional po_no narrows dispatch rows. Expenses / acceptance: placeholders for Phase 2.

    PM / desk admin only — not for INET IM / field roles.
    """
    user = frappe.session.user
    if not user or user == "Guest":
        frappe.throw("Not permitted", frappe.PermissionError)
    roles = set(frappe.get_roles(user))
    if not (
        "Administrator" in roles
        or "System Manager" in roles
        or "INET Admin" in roles
    ):
        frappe.throw("Not permitted", frappe.PermissionError)

    duid = (duid or "").strip()
    po_no = (po_no or "").strip()
    if not duid and not po_no:
        frappe.throw("Provide duid (site / DUID) and/or po_no")

    dfilters = {}
    if duid:
        dfilters["site_code"] = duid
    if po_no:
        dfilters["po_no"] = po_no

    dispatches = frappe.get_all(
        "PO Dispatch",
        filters=dfilters,
        fields=["*"],
        order_by="modified desc",
        limit_page_length=50,
    )
    dispatch_names = [d.name for d in dispatches]
    plans = []
    executions = []
    if dispatch_names:
        plans = frappe.get_all(
            "Rollout Plan",
            filters={"po_dispatch": ["in", dispatch_names]},
            fields=["*"],
            order_by="plan_date desc",
            limit_page_length=200,
        )
        plan_names = [p.name for p in plans]
        if plan_names:
            executions = frappe.get_all(
                "Daily Execution",
                filters={"rollout_plan": ["in", plan_names]},
                fields=["*"],
                order_by="execution_date desc",
                limit_page_length=200,
            )

    return {
        "duid": duid or None,
        "po_no": po_no or None,
        "dispatches": dispatches,
        "rollout_plans": plans,
        "executions": executions,
        "additional_activities": [],
        "expenses": [],
        "acceptance": [],
        "notes": "Additional activities, expenses, and acceptance lines can be linked in a later phase.",
    }


@frappe.whitelist()
def reopen_rollout_for_revisit(rollout_plan, issue_category=None, planning_route=None, issue_remarks=None):
    """
    Re-visit workflow: move job back to planning with issue.
    Non-completed executions for this plan are cancelled.
    """
    if not rollout_plan or not frappe.db.exists("Rollout Plan", rollout_plan):
        frappe.throw("Invalid Rollout Plan")

    # Route selection removed: all re-plans are tracked as issue/risk.
    new_status = "Planning with Issue"

    source = frappe.get_doc("Rollout Plan", rollout_plan)
    existing_revisit = frappe.db.get_value(
        "Rollout Plan",
        {
            "po_dispatch": source.po_dispatch,
            "visit_type": "Re-Visit",
            "plan_status": ["in", ["Planned", "Planning with Issue", "In Execution"]],
        },
        "name",
    )

    if existing_revisit and existing_revisit != rollout_plan:
        # Re-stamp the existing re-visit with the new issue context so it
        # reappears on Issues & Risks. Also force plan_status back to
        # "Planning with Issue" in case it was moved elsewhere.
        revisit_name = existing_revisit
        restamp = {
            "plan_status": "Planning with Issue",
            "issue_category": (issue_category or "")[:140],
        }
        if frappe.db.has_column("Rollout Plan", "issue_remarks") and issue_remarks:
            restamp["issue_remarks"] = str(issue_remarks)[:2000]
        frappe.db.set_value(
            "Rollout Plan", existing_revisit, restamp, update_modified=True
        )
    else:
        revisit = frappe.new_doc("Rollout Plan")
        revisit.po_dispatch = source.po_dispatch
        revisit.im = source.im
        revisit.team = source.team
        revisit.plan_date = nowdate()
        if hasattr(revisit, "plan_end_date"):
            revisit.plan_end_date = nowdate()
        if hasattr(revisit, "access_time"):
            revisit.access_time = getattr(source, "access_time", None)
        if hasattr(revisit, "access_period"):
            revisit.access_period = getattr(source, "access_period", None)
        revisit.visit_type = "Re-Visit"
        revisit.visit_number = _next_visit_number_for_dispatch(source.po_dispatch)
        pd_reg = frappe.db.get_value(
            "PO Dispatch", source.po_dispatch, ["region_type", "center_area"], as_dict=True
        ) or {}
        if hasattr(revisit, "region_type"):
            revisit.region_type = pd_reg.get("region_type") or region_type_from_center_area(
                pd_reg.get("center_area")
            )
        rev_mult = flt(
            frappe.db.get_value("Visit Multiplier Master", "Re-Visit", "multiplier") or 0.5
        )
        revisit.visit_multiplier = rev_mult
        tgt_src = flt(source.target_amount or 0)
        revisit.target_amount = tgt_src * rev_mult if tgt_src > 0 else 0
        revisit.plan_status = new_status
        revisit.issue_category = (issue_category or "")[:140]
        if hasattr(revisit, "issue_remarks") and issue_remarks:
            revisit.issue_remarks = str(issue_remarks)[:2000]
        if hasattr(revisit, "source_rollout_plan"):
            revisit.source_rollout_plan = rollout_plan
        revisit.insert(ignore_permissions=True)
        revisit_name = revisit.name

    # Keep source plan as historical record; do not overwrite to Re-Visit.
    source_updates = {"issue_category": (issue_category or "")[:140]}
    if frappe.db.has_column("Rollout Plan", "issue_remarks") and issue_remarks:
        source_updates["issue_remarks"] = str(issue_remarks)[:2000]
    frappe.db.set_value(
        "Rollout Plan",
        rollout_plan,
        source_updates,
        update_modified=True,
    )

    for row in frappe.get_all(
        "Daily Execution",
        filters={"rollout_plan": rollout_plan, "execution_status": ["not in", ["Completed", "Cancelled"]]},
        pluck="name",
    ):
        frappe.db.set_value("Daily Execution", row, "execution_status", "Cancelled", update_modified=True)

    frappe.db.commit()
    return {
        "ok": True,
        "source_rollout_plan": rollout_plan,
        "revisit_rollout_plan": revisit_name,
        "plan_status": new_status,
    }


@frappe.whitelist()
def backfill_rollout_and_execution_im():
    """One-shot: set Rollout Plan.im and Daily Execution.im from PO Dispatch (Administrator)."""
    frappe.only_for("System Manager")
    updated_rp = 0
    updated_de = 0
    for rp in frappe.get_all("Rollout Plan", filters={"po_dispatch": ["!=", ""]}, fields=["name", "po_dispatch"]):
        im_v = frappe.db.get_value("PO Dispatch", rp.po_dispatch, "im")
        if im_v:
            frappe.db.set_value("Rollout Plan", rp.name, "im", im_v, update_modified=False)
            updated_rp += 1
    for de in frappe.get_all("Daily Execution", filters={"rollout_plan": ["!=", ""]}, fields=["name", "rollout_plan"]):
        pd = frappe.db.get_value("Rollout Plan", de.rollout_plan, "po_dispatch")
        if not pd:
            continue
        im_v = frappe.db.get_value("PO Dispatch", pd, "im")
        if im_v:
            frappe.db.set_value("Daily Execution", de.name, "im", im_v, update_modified=False)
            updated_de += 1
    frappe.db.commit()
    return {"rollout_plans_updated": updated_rp, "executions_updated": updated_de}


@frappe.whitelist()
def get_im_dashboard(im=None, from_date=None, to_date=None, etag=None):
    """
    Filtered dashboard for a single IM.
    Resolves the IM Master record name in multiple ways so mismatched
    im/full_name values don't silently return empty data.

    Optional ``from_date`` / ``to_date`` (ISO YYYY-MM-DD) override the
    default month window for date-scoped metrics.

    If the caller passes an ``etag`` matching the current data version,
    short-circuits with ``{"unchanged": True, "etag": ...}``.
    """
    current_etag = _dashboard_etag("im", im or "", from_date, to_date)
    if etag and etag == current_etag:
        return {"unchanged": True, "etag": current_etag, "last_updated": _iso_now()}
    im_resolved, im_identifiers, _ = resolve_im_for_session(im)
    if not im_identifiers:
        return {
            "im": None,
            "teams": [],
            "projects": [],
            "kpi": {},
            "action_items": {"pending_plan_dispatches": 0, "qc_fail_needs_action": 0, "planned_ready_execution": 0},
            "debug": {"error": "Could not resolve IM from session or parameter"},
            "etag": current_etag,
            "last_updated": _iso_now(),
        }

    im = im_resolved
    action_items = im_action_counts(im_identifiers)

    today_str = nowdate()
    today = getdate(today_str)
    first_day, last_day, _ = _month_bounds()
    if from_date:
        try: first_day = getdate(from_date)
        except Exception: pass
    if to_date:
        try: last_day = getdate(to_date)
        except Exception: pass
    days_in_month = _days_in_month(today)
    day_of_month = today.day

    # Teams belonging to this IM — match any known identifier value
    teams = frappe.get_all(
        "INET Team",
        filters={"im": ["in", im_identifiers], "status": "Active"},
        fields=["name", "team_id", "team_name", "team_type", "daily_cost", "status"],
    )
    team_ids = [t.name for t in teams]

    # Projects assigned to this IM — match any known identifier value
    projects = frappe.get_all(
        "Project Control Center",
        filters={"implementation_manager": ["in", im_identifiers]},
        fields=["name", "project_code", "project_name", "project_status",
                "completion_percentage", "budget_amount", "customer"],
        order_by="modified desc",
        limit=50,
    )
    # Normalize field names for frontend
    for p in projects:
        p["completion_pct"] = p.pop("completion_percentage", 0) or 0
        p["budget"] = p.pop("budget_amount", 0) or 0
        p["status"] = p.pop("project_status", "Active") or "Active"

    debug_info = {
        "im_resolved": im_resolved,
        "im_identifiers": im_identifiers,
        "teams_found": len(team_ids),
        "projects_found": len(projects),
        "action_items": action_items,
    }

    if not team_ids:
        return {
            "im": im_resolved,
            "teams": [],
            "projects": projects,
            "kpi": {"team_count": 0, "revenue": 0, "cost": 0, "profit": 0},
            "action_items": action_items,
            "message": (
                f"No active INET Teams found for this IM "
                f"(searched: {', '.join(im_identifiers)}). "
                f"In INET Team master set 'Implementation Manager' to one of those values."
            ),
            "debug": debug_info,
            "last_updated": _iso_now(),
            "etag": current_etag,
        }

    placeholders = ", ".join(["%s"] * len(team_ids))

    # Revenue this month
    revenue_rows = frappe.db.sql(
        f"""
        SELECT COALESCE(SUM(wd.revenue_sar), 0) AS revenue,
               COALESCE(SUM(wd.total_cost_sar), 0) AS cost
        FROM `tabWork Done` wd
        JOIN `tabDaily Execution` exe ON exe.name = wd.execution
        JOIN `tabRollout Plan` rp ON rp.name = exe.rollout_plan
        WHERE rp.team IN ({placeholders})
        AND exe.execution_date BETWEEN %s AND %s
        """,
        tuple(team_ids) + (first_day, last_day),
        as_dict=True,
    )
    revenue = flt(revenue_rows[0].revenue if revenue_rows else 0)
    cost = flt(revenue_rows[0].cost if revenue_rows else 0)

    # Monthly target — projects where this IM is implementation_manager
    im_ph = ", ".join(["%s"] * len(im_identifiers))
    monthly_target_rows = frappe.db.sql(
        f"""
        SELECT COALESCE(SUM(pcc.monthly_target), 0) AS total
        FROM `tabProject Control Center` pcc
        WHERE pcc.implementation_manager IN ({im_ph})
        AND pcc.active_flag = 'Yes'
        """,
        tuple(im_identifiers),
        as_dict=True,
    )
    monthly_target = flt(monthly_target_rows[0].total if monthly_target_rows else 0)
    target_today = (
        (monthly_target * day_of_month) / days_in_month if days_in_month else 0.0
    )
    gap_today = target_today - revenue

    # Active teams today
    active_today_rows = frappe.db.sql(
        f"""
        SELECT COUNT(DISTINCT team) AS cnt FROM `tabDaily Execution`
        WHERE team IN ({placeholders})
        AND execution_date = %s
        """,
        tuple(team_ids) + (today_str,),
        as_dict=True,
    )
    active_today = cint(active_today_rows[0].cnt if active_today_rows else 0)

    # Planned activities — match the Planning page: every Rollout Plan whose
    # PO Dispatch belongs to this IM (or whose rp.im column points at this IM)
    # and is in `Planned` status. Counting by team missed plans that weren't
    # team-assigned yet or that were on teams outside the active list, which
    # made the dashboard KPI disagree with the Planning table.
    im_ph = ", ".join(["%s"] * len(im_identifiers))
    rp_im_clause = ""
    rp_im_params = []
    if frappe.db.has_column("Rollout Plan", "im"):
        rp_im_clause = f" OR IFNULL(rp.im,'') IN ({im_ph})"
        rp_im_params = list(im_identifiers)
    planned_rows = frappe.db.sql(
        f"""
        SELECT COUNT(*) AS cnt
        FROM `tabRollout Plan` rp
        LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        WHERE rp.plan_status = 'Planned'
        AND (IFNULL(pd.im,'') IN ({im_ph}){rp_im_clause})
        """,
        tuple(im_identifiers) + tuple(rp_im_params),
        as_dict=True,
    )
    planned_activities = cint(planned_rows[0].cnt if planned_rows else 0)

    # ── Site-centric KPIs (rollout plans = sites in this context) ─────────
    site_kpi_rows = frappe.db.sql(
        f"""
        SELECT
            COUNT(*) AS total_assigned,
            SUM(CASE WHEN rp.plan_status = 'Completed' THEN 1 ELSE 0 END) AS completed_total,
            SUM(CASE WHEN rp.plan_status = 'In Execution' THEN 1 ELSE 0 END) AS in_progress_cnt,
            SUM(CASE WHEN rp.plan_status = 'Planning with Issue' THEN 1 ELSE 0 END) AS delayed_cnt,
            SUM(CASE WHEN rp.plan_date = %s THEN 1 ELSE 0 END) AS today_target_cnt,
            SUM(CASE WHEN rp.plan_status = 'Completed' AND rp.modified >= %s THEN 1 ELSE 0 END) AS today_completed_cnt
        FROM `tabRollout Plan` rp
        LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        WHERE (IFNULL(pd.im,'') IN ({im_ph}){rp_im_clause})
        """,
        (today_str, f"{today_str} 00:00:00") + tuple(im_identifiers) + tuple(rp_im_params),
        as_dict=True,
    )
    skr = (site_kpi_rows[0] if site_kpi_rows else {}) or {}
    site_kpi = {
        "total_assigned": cint(skr.get("total_assigned") or 0),
        "completed_total": cint(skr.get("completed_total") or 0),
        "in_progress": cint(skr.get("in_progress_cnt") or 0),
        "delayed": cint(skr.get("delayed_cnt") or 0),
        "today_target": cint(skr.get("today_target_cnt") or 0),
        "today_completed": cint(skr.get("today_completed_cnt") or 0),
    }

    # ── Team performance: completed sites per team in the date window ────
    team_perf = []
    if team_ids:
        team_perf_rows = frappe.db.sql(
            f"""
            SELECT t.team_id AS team_id,
                   COALESCE(t.team_name, t.team_id) AS team_name,
                   COUNT(rp.name) AS sites_done
            FROM `tabINET Team` t
            LEFT JOIN `tabRollout Plan` rp
                ON rp.team = t.name
                AND rp.plan_status = 'Completed'
                AND rp.plan_date BETWEEN %s AND %s
            WHERE t.name IN ({placeholders})
            GROUP BY t.name, t.team_id, t.team_name
            ORDER BY sites_done DESC, team_name ASC
            """,
            (first_day, last_day) + tuple(team_ids),
            as_dict=True,
        )
        team_perf = [
            {"team_id": r.team_id, "team_name": r.team_name, "sites_done": cint(r.sites_done or 0)}
            for r in (team_perf_rows or [])
        ]

    # ── Project progress: completion % per project for this IM ───────────
    project_progress = []
    if projects:
        project_codes = [p["project_code"] for p in projects if p.get("project_code")]
        if project_codes:
            ph_pr = ", ".join(["%s"] * len(project_codes))
            prog_rows = frappe.db.sql(
                f"""
                SELECT pd.project_code AS project_code,
                       COUNT(rp.name) AS total,
                       SUM(CASE WHEN rp.plan_status = 'Completed' THEN 1 ELSE 0 END) AS done
                FROM `tabPO Dispatch` pd
                LEFT JOIN `tabRollout Plan` rp ON rp.po_dispatch = pd.name
                WHERE pd.project_code IN ({ph_pr})
                AND IFNULL(pd.im,'') IN ({im_ph})
                GROUP BY pd.project_code
                """,
                tuple(project_codes) + tuple(im_identifiers),
                as_dict=True,
            )
            prog_map = {r.project_code: r for r in (prog_rows or [])}
            for p in projects[:6]:
                pc = p.get("project_code")
                row = prog_map.get(pc)
                total = cint(row.total) if row else 0
                done = cint(row.done) if row else 0
                pct = int(round((done / total) * 100)) if total > 0 else 0
                project_progress.append({
                    "project_code": pc,
                    "project_name": p.get("project_name") or pc,
                    "total": total,
                    "done": done,
                    "pct": pct,
                })

    # ── Site status table: latest rollout plans for this IM ──────────────
    # Site ID = DUID code from PO Dispatch (joins to DUID Master).
    # Location = Area Master.area_name via DUID Master.area. No fallback to
    # center_area / site_name — if the DUID has no Area set, the cell stays
    # empty so it's obvious where the master data is missing.
    site_status = frappe.db.sql(
        f"""
        SELECT pd.site_code AS site_id,
               am.area_name AS location,
               pd.project_code AS project,
               rp.plan_status AS status,
               rp.modified AS last_update,
               rp.issue_category AS issue
        FROM `tabRollout Plan` rp
        LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        LEFT JOIN `tabDUID Master` dm ON dm.name = pd.site_code
        LEFT JOIN `tabArea Master` am ON am.name = dm.area
        WHERE (IFNULL(pd.im,'') IN ({im_ph}){rp_im_clause})
        ORDER BY rp.modified DESC
        LIMIT 8
        """,
        tuple(im_identifiers) + tuple(rp_im_params),
        as_dict=True,
    )
    site_status = [
        {
            "site_id": r.site_id or "—",
            "location": r.location or "—",
            "project": r.project or "—",
            "status": r.status or "—",
            "last_update": str(r.last_update) if r.last_update else "",
            "issue": r.issue or "",
        }
        for r in (site_status or [])
    ]

    # ── Material shortage: rollout plans / executions tagged with that issue
    material_categories = ("Material Shortage", "Spare Parts Pending")
    mat_ph = ", ".join(["%s"] * len(material_categories))
    mat_rows = frappe.db.sql(
        f"""
        SELECT COUNT(*) AS cnt FROM (
            SELECT rp.name FROM `tabRollout Plan` rp
            LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
            WHERE rp.issue_category IN ({mat_ph})
            AND (IFNULL(pd.im,'') IN ({im_ph}){rp_im_clause})
            UNION
            SELECT de.name FROM `tabDaily Execution` de
            LEFT JOIN `tabRollout Plan` rp ON rp.name = de.rollout_plan
            LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
            WHERE de.issue_category IN ({mat_ph})
            AND (IFNULL(pd.im,'') IN ({im_ph}){rp_im_clause})
        ) AS u
        """,
        tuple(material_categories) + tuple(im_identifiers) + tuple(rp_im_params)
        + tuple(material_categories) + tuple(im_identifiers) + tuple(rp_im_params),
        as_dict=True,
    )
    material_shortage = cint(mat_rows[0].cnt if mat_rows else 0)

    # ── Site locations: lat/lon from DUID Master joined via PO Dispatch ──
    site_locations = []
    has_lat = frappe.db.has_column("DUID Master", "latitude")
    has_lon = frappe.db.has_column("DUID Master", "longitude")
    if has_lat and has_lon:
        site_loc_rows = frappe.db.sql(
            f"""
            SELECT
                pd.site_code AS site_code,
                COALESCE(pd.site_name, pd.site_code) AS site_name,
                pd.project_code AS project_code,
                d.latitude AS lat,
                d.longitude AS lon,
                MAX(rp.modified) AS last_update,
                /* Status: latest rollout plan's status if any, else 'Planned' */
                SUBSTRING_INDEX(GROUP_CONCAT(rp.plan_status ORDER BY rp.modified DESC), ',', 1) AS status
            FROM `tabPO Dispatch` pd
            INNER JOIN `tabDUID Master` d ON d.name = pd.site_code
            LEFT JOIN `tabRollout Plan` rp ON rp.po_dispatch = pd.name
            WHERE IFNULL(pd.im,'') IN ({im_ph})
              AND IFNULL(d.latitude, 0) <> 0
              AND IFNULL(d.longitude, 0) <> 0
            GROUP BY pd.site_code, pd.site_name, pd.project_code, d.latitude, d.longitude
            ORDER BY last_update DESC
            LIMIT 200
            """,
            tuple(im_identifiers),
            as_dict=True,
        )
        for r in (site_loc_rows or []):
            try:
                lat = float(r.lat)
                lon = float(r.lon)
            except (TypeError, ValueError):
                continue
            site_locations.append({
                "site_code": r.site_code,
                "site_name": r.site_name or r.site_code,
                "project_code": r.project_code,
                "lat": lat,
                "lon": lon,
                "status": r.status or "Planned",
                "last_update": str(r.last_update) if r.last_update else "",
            })

    # ── Activity timeline: recent Daily Execution events ─────────────────
    timeline_rows = frappe.db.sql(
        f"""
        SELECT de.name AS exec_id,
               de.execution_date,
               de.modified AS ts,
               de.execution_status,
               de.rollout_plan,
               COALESCE(pd.site_code, pd.site_name) AS site
        FROM `tabDaily Execution` de
        LEFT JOIN `tabRollout Plan` rp ON rp.name = de.rollout_plan
        LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch
        WHERE (IFNULL(pd.im,'') IN ({im_ph}){rp_im_clause})
        ORDER BY de.modified DESC
        LIMIT 8
        """,
        tuple(im_identifiers) + tuple(rp_im_params),
        as_dict=True,
    )
    timeline = []
    for r in (timeline_rows or []):
        verb = (
            "completed" if (r.execution_status or "").lower() == "completed"
            else "delay reported" if (r.execution_status or "").lower() in ("hold", "cancelled", "postponed")
            else "work started" if (r.execution_status or "").lower() == "in progress"
            else "updated"
        )
        site = r.site or r.rollout_plan or "—"
        timeline.append({
            "ts": str(r.ts) if r.ts else "",
            "label": f"Site {site} {verb}",
            "exec_id": r.exec_id,
        })

    return {
        "im": im_resolved,
        "teams": teams,
        "projects": projects,
        "action_items": action_items,
        "site_kpi": site_kpi,
        "team_perf": team_perf,
        "project_progress": project_progress,
        "site_status": site_status,
        "timeline": timeline,
        "material_shortage": material_shortage,
        "site_locations": site_locations,
        "kpi": {
            "monthly_target": monthly_target,
            "target_today": target_today,
            "revenue": revenue,
            "cost": cost,
            "profit": revenue - cost,
            "gap_today": gap_today,
            "active_teams_today": active_today,
            "total_teams": len(teams),
            "team_count": len(teams),
            "planned_activities": planned_activities,
        },
        "debug": debug_info,
        "last_updated": _iso_now(),
        "etag": current_etag,
    }


@frappe.whitelist()
def get_im_reports():
    """
    IM-scoped report bundle for the portal (single round-trip).
    Same IM resolution as get_im_dashboard; no hardcoded placeholders.
    """
    im_resolved, im_identifiers = _require_inet_im_session()
    first_day, last_day, _today_str = _month_bounds()
    im_ph = ", ".join(["%s"] * len(im_identifiers))

    disp_rows = frappe.db.sql(
        f"""
        SELECT dispatch_status, COALESCE(project_code, '') AS project_code,
               COALESCE(line_amount, 0) AS line_amount, dispatch_mode
        FROM `tabPO Dispatch`
        WHERE im IN ({im_ph})
        LIMIT 2000
        """,
        tuple(im_identifiers),
        as_dict=True,
    ) or []
    total_lines = len(disp_rows)
    total_amount = sum(flt(r.get("line_amount")) for r in disp_rows)
    by_status = {}
    by_project = {}
    by_mode = {}
    for r in disp_rows:
        st = (r.get("dispatch_status") or "").strip() or "(No status)"
        by_status[st] = by_status.get(st, 0) + 1
        pk = (r.get("project_code") or "").strip() or "(No project)"
        by_project[pk] = by_project.get(pk, 0) + flt(r.get("line_amount"))
        md = (r.get("dispatch_mode") or "").strip() or "(No mode)"
        by_mode[md] = by_mode.get(md, 0) + 1

    teams = frappe.get_all(
        "INET Team",
        filters={"im": ["in", im_identifiers], "status": "Active"},
        fields=["name", "team_id", "team_name"],
    )
    team_ids = [t.name for t in teams]

    rollout_status_counts = []
    rollouts_recent = []
    exec_status_counts = []
    executions_recent = []
    wd_mtd = {"count": 0, "revenue_sar": 0.0, "by_billing": {}}

    if team_ids:
        tph = ", ".join(["%s"] * len(team_ids))
        rollout_status_counts = (
            frappe.db.sql(
                f"""
                SELECT rp.plan_status AS status_key, COUNT(*) AS cnt
                FROM `tabRollout Plan` rp
                WHERE rp.team IN ({tph})
                GROUP BY rp.plan_status
                ORDER BY cnt DESC
                """,
                tuple(team_ids),
                as_dict=True,
            )
            or []
        )

        rollouts_recent = (
            frappe.db.sql(
                f"""
                SELECT rp.name, rp.plan_date, rp.plan_status, rp.team, rp.po_dispatch,
                       rp.visit_type, rp.target_amount
                FROM `tabRollout Plan` rp
                WHERE rp.team IN ({tph})
                ORDER BY rp.modified DESC
                LIMIT 80
                """,
                tuple(team_ids),
                as_dict=True,
            )
            or []
        )
        for r in rollouts_recent:
            tid = r.get("team")
            r["team_name"] = frappe.db.get_value("INET Team", tid, "team_name") or tid

        exec_status_counts = (
            frappe.db.sql(
                f"""
                SELECT de.execution_status AS status_key, COUNT(*) AS cnt
                FROM `tabDaily Execution` de
                WHERE de.team IN ({tph})
                AND de.execution_date BETWEEN %s AND %s
                GROUP BY de.execution_status
                ORDER BY cnt DESC
                """,
                tuple(team_ids) + (first_day, last_day),
                as_dict=True,
            )
            or []
        )

        executions_recent = (
            frappe.db.sql(
                f"""
                SELECT de.name, de.rollout_plan, de.execution_date, de.execution_status,
                       de.qc_status, de.team, rp.po_dispatch
                FROM `tabDaily Execution` de
                LEFT JOIN `tabRollout Plan` rp ON rp.name = de.rollout_plan
                WHERE de.team IN ({tph})
                AND de.execution_date BETWEEN %s AND %s
                ORDER BY de.modified DESC
                LIMIT 60
                """,
                tuple(team_ids) + (first_day, last_day),
                as_dict=True,
            )
            or []
        )

        wd_rows = (
            frappe.db.sql(
                f"""
                SELECT wd.billing_status AS billing_status,
                       COUNT(*) AS c,
                       COALESCE(SUM(wd.revenue_sar), 0) AS rev
                FROM `tabWork Done` wd
                INNER JOIN `tabDaily Execution` exe ON exe.name = wd.execution
                WHERE exe.team IN ({tph})
                AND exe.execution_date BETWEEN %s AND %s
                GROUP BY wd.billing_status
                """,
                tuple(team_ids) + (first_day, last_day),
                as_dict=True,
            )
            or []
        )
        by_billing = {}
        total_rev = 0.0
        total_cnt = 0
        for w in wd_rows:
            bs = (w.get("billing_status") or "").strip() or "Pending"
            c = cint(w.get("c"))
            rev = flt(w.get("rev"))
            by_billing[bs] = {"count": c, "revenue_sar": rev}
            total_cnt += c
            total_rev += rev
        wd_mtd = {"count": total_cnt, "revenue_sar": total_rev, "by_billing": by_billing}

    projects = frappe.get_all(
        "Project Control Center",
        filters={"implementation_manager": ["in", im_identifiers]},
        fields=[
            "name",
            "project_code",
            "project_name",
            "project_status",
            "completion_percentage",
            "budget_amount",
            "actual_cost",
        ],
        order_by="modified desc",
        limit_page_length=50,
    )
    for p in projects:
        p["completion_pct"] = p.get("completion_percentage") or 0
        p["budget"] = flt(p.get("budget_amount"))
        p["status"] = p.get("project_status") or "Active"

    return {
        "im": im_resolved,
        "last_updated": _iso_now(),
        "period": {"from": first_day, "to": last_day},
        "dispatch_summary": {
            "total_lines": total_lines,
            "total_amount": total_amount,
            "by_status": by_status,
            "by_project": by_project,
            "by_dispatch_mode": by_mode,
        },
        "rollout_status_counts": rollout_status_counts,
        "rollouts_recent": rollouts_recent,
        "execution_status_counts": exec_status_counts,
        "executions_recent": executions_recent,
        "work_done_mtd": wd_mtd,
        "projects": projects,
        "teams": teams,
    }


@frappe.whitelist()
def get_field_team_dashboard(team_id=None):
    """
    Return today's planned work for a specific team.
    Enriches rollout plan data with dispatch info (item_code, item_description,
    project_code, site_name).

    Returns
    -------
    {
        "team": team_id,
        "today": "YYYY-MM-DD",
        "planned": [...],
        "last_updated": ...,
    }
    """
    today_str = nowdate()

    if not team_id:
        return {"team": None, "today": today_str, "planned": [], "last_updated": _iso_now()}

    # Rollout Plans for this team today
    plans = frappe.db.sql(
        """
        SELECT rp.name, rp.po_dispatch, rp.plan_date, rp.visit_type,
               rp.visit_number, rp.visit_multiplier, rp.target_amount,
               rp.achieved_amount, rp.completion_pct, rp.plan_status
        FROM `tabRollout Plan` rp
        WHERE rp.team = %s
        AND rp.plan_date = %s
        AND rp.plan_status IN ('Planned', 'Ready for Execution', 'In Execution', 'Planning with Issue')
        ORDER BY rp.name
        """,
        (team_id, today_str),
        as_dict=True,
    )

    if not plans:
        return {
            "team": team_id,
            "today": today_str,
            "planned": [],
            "last_updated": _iso_now(),
        }

    dispatch_names = [p.po_dispatch for p in plans if p.po_dispatch]
    plan_names = [p.name for p in plans]
    dispatch_map = {}

    if dispatch_names:
        placeholders = ", ".join(["%s"] * len(dispatch_names))
        dispatch_rows = frappe.db.sql(
            f"""
            SELECT name, poid, item_code, item_description, qty, project_code,
                   customer, site_name, site_code, center_area, region_type
            FROM `tabPO Dispatch`
            WHERE name IN ({placeholders})
            """,
            tuple(dispatch_names),
            as_dict=True,
        )
        dispatch_map = {d.name: d for d in dispatch_rows}

    # Latest Daily Execution per plan: powers the IM-confirmation badge and
    # the field team's own tl_status indicator on the card. One pass.
    exec_map = {}
    if plan_names:
        ex_placeholders = ", ".join(["%s"] * len(plan_names))
        exec_rows = frappe.db.sql(
            f"""
            SELECT de.rollout_plan, de.execution_status, de.tl_status,
                   de.qc_status, de.ciag_status, de.achieved_qty, de.achieved_amount
            FROM `tabDaily Execution` de
            INNER JOIN (
              SELECT rollout_plan, MAX(modified) AS m
              FROM `tabDaily Execution`
              WHERE rollout_plan IN ({ex_placeholders})
              GROUP BY rollout_plan
            ) t ON t.rollout_plan = de.rollout_plan AND t.m = de.modified
            """,
            tuple(plan_names),
            as_dict=True,
        )
        exec_map = {r["rollout_plan"]: r for r in exec_rows}

    # Customer Activity Type lookup, keyed by (customer, item_code).
    cim_map = _batch_customer_activity_types(
        [{"customer": (d or {}).get("customer"), "item_code": (d or {}).get("item_code")} for d in dispatch_map.values()]
    )

    # Merge dispatch data into plan records
    enriched = []
    for plan in plans:
        dispatch_info = dispatch_map.get(plan.po_dispatch) or {}
        ex = exec_map.get(plan.name) or {}
        enriched.append(
            {
                "name": plan.name,
                "po_dispatch": plan.po_dispatch,
                "plan_date": plan.plan_date,
                "visit_type": plan.visit_type,
                "visit_number": plan.visit_number,
                "visit_multiplier": plan.visit_multiplier,
                "target_amount": plan.target_amount,
                "achieved_amount": plan.achieved_amount,
                "completion_pct": plan.completion_pct,
                "plan_status": plan.plan_status,
                "poid": dispatch_info.get("poid"),
                "item_code": dispatch_info.get("item_code"),
                "item_description": dispatch_info.get("item_description"),
                "qty": dispatch_info.get("qty"),
                "project_code": dispatch_info.get("project_code"),
                "site_name": dispatch_info.get("site_name"),
                "site_code": dispatch_info.get("site_code"),
                "center_area": dispatch_info.get("center_area"),
                "region_type": dispatch_info.get("region_type")
                or region_type_from_center_area(dispatch_info.get("center_area")),
                "customer_activity_type": cim_map.get(
                    (dispatch_info.get("customer") or "", dispatch_info.get("item_code") or "")
                ),
                # IM-confirmed and field-side execution flags so the card
                # can show "IM ✓" or "Awaiting IM" without a second fetch.
                "execution_status": ex.get("execution_status"),
                "tl_status": ex.get("tl_status"),
                "qc_status": ex.get("qc_status"),
                "ciag_status": ex.get("ciag_status"),
            }
        )

    return {
        "team": team_id,
        "today": today_str,
        "planned": enriched,
        "last_updated": _iso_now(),
    }


# ---------------------------------------------------------------------------
# Project Detail — Full project summary with related records
# ---------------------------------------------------------------------------

def _normalize_duid_site(site_code):
    """Group key for site / DUID; empty becomes a single bucket."""
    s = (site_code or "").strip()
    return s if s else "(No DUID)"


def _build_rollout_by_duid_groups(dispatches, plans, executions, work_done):
    """
    Rollout view grouped by DUID (PO Dispatch.site_code).
    Planned activity = Rollout Plan with visit_type Work Done (default path).
    Additional = Extra Visit, Re-Visit.
    Expenses = Work Done cost lines for executions under those plans.
    Time logs = Execution Time Log rows linked to rollout plans in the group.
    """
    dispatch_by_name = {d.name: d for d in dispatches}

    def new_group(duid_label):
        return {
            "duid_label": duid_label,
            "site_code": "",
            "site_name": "",
            "po_lines": [],
            "planned_activities": [],
            "additional_activities": [],
            "expenses": [],
            "time_logs": [],
        }

    groups = {}

    def ensure_group(duid_label):
        if duid_label not in groups:
            groups[duid_label] = new_group(duid_label)
        return groups[duid_label]

    plan_to_dispatch = {p.name: p.get("po_dispatch") for p in plans}
    exec_to_plan = {e.name: e.get("rollout_plan") for e in executions}

    def duid_for_dispatch_name(dname):
        row = dispatch_by_name.get(dname)
        if not row:
            return "(Unknown dispatch)"
        return _normalize_duid_site(row.get("site_code"))

    def duid_for_plan_name(pname):
        pd = plan_to_dispatch.get(pname)
        if not pd:
            return "(Unknown dispatch)"
        return duid_for_dispatch_name(pd)

    for d in dispatches:
        key = _normalize_duid_site(d.get("site_code"))
        g = ensure_group(key)
        if not g.get("site_code") and d.get("site_code"):
            g["site_code"] = (d.get("site_code") or "").strip()
        if not g.get("site_name") and d.get("site_name"):
            g["site_name"] = (d.get("site_name") or "").strip()
        g["po_lines"].append(d)

    additional_visit_types = frozenset({"Extra Visit", "Re-Visit"})

    for p in plans:
        key = duid_for_plan_name(p.name)
        g = ensure_group(key)
        vt = (p.get("visit_type") or "Work Done").strip()
        if vt in additional_visit_types:
            g["additional_activities"].append(p)
        else:
            g["planned_activities"].append(p)

    for wd in work_done:
        ex = wd.get("execution")
        if not ex:
            continue
        rplan = exec_to_plan.get(ex)
        if not rplan:
            continue
        key = duid_for_plan_name(rplan)
        ensure_group(key)["expenses"].append(wd)

    plan_names = [p.name for p in plans]
    time_logs = []
    if plan_names and frappe.db.exists("DocType", "Execution Time Log"):
        time_logs = frappe.get_all(
            "Execution Time Log",
            filters={"rollout_plan": ["in", plan_names]},
            fields=["*"],
            order_by="start_time desc",
            limit_page_length=2000,
        )
    for tl in time_logs:
        rp = tl.get("rollout_plan")
        if not rp:
            continue
        key = duid_for_plan_name(rp)
        ensure_group(key)["time_logs"].append(tl)

    def sort_key(k):
        if k == "(No DUID)":
            return (2, k)
        if k == "(Unknown dispatch)":
            return (1, k)
        return (0, k)

    ordered = [groups[k] for k in sorted(groups.keys(), key=sort_key)]
    for g in ordered:
        poids = sorted({r.get("name") for r in g["po_lines"] if r.get("name")})
        g["poid_list"] = poids
    return ordered


@frappe.whitelist()
def get_project_summary(project_code):
    """Get complete project summary with all related records."""
    if not project_code:
        frappe.throw("project_code is required")

    project = frappe.get_doc("Project Control Center", project_code).as_dict()

    # PO Dispatches for this project (full rows for detail / rollout grouping)
    dispatches = frappe.get_all(
        "PO Dispatch",
        filters={"project_code": project_code},
        fields=["*"],
        order_by="site_code asc, modified desc",
        limit_page_length=500,
    )

    # Get dispatch names for downstream queries
    dispatch_names = [d.name for d in dispatches]

    # Rollout Plans
    plans = []
    if dispatch_names:
        plans = frappe.get_all(
            "Rollout Plan",
            filters={"po_dispatch": ["in", dispatch_names]},
            fields=["*"],
            order_by="plan_date desc",
            limit_page_length=500,
        )

    # Daily Executions
    plan_names = [p.name for p in plans]
    executions = []
    if plan_names:
        executions = frappe.get_all(
            "Daily Execution",
            filters={"rollout_plan": ["in", plan_names]},
            fields=["*"],
            order_by="execution_date desc",
            limit_page_length=500,
        )
        plan_poid_map = {p.name: p.po_dispatch for p in plans}
        for e in executions:
            e["system_id"] = plan_poid_map.get(e.rollout_plan)

    # Work Done
    execution_names = [e.name for e in executions]
    work_done = []
    if execution_names:
        work_done = frappe.get_all(
            "Work Done",
            filters={"execution": ["in", execution_names]},
            fields=["*"],
            limit_page_length=500,
        )
        exec_poid_map = {e.name: e.get("system_id") for e in executions}
        for wd in work_done:
            wd["system_id"] = exec_poid_map.get(wd.execution)

    # Teams involved — from rollout plans + Team Assignment doctype
    dispatch_teams = set(p.team for p in plans if p.get("team"))

    # Team Assignments for this project
    team_assignments = frappe.get_all("Team Assignment",
        filters={"project": project_code},
        fields=["name", "team_id", "assignment_date", "end_date", "role_in_project",
                "daily_cost", "utilization_percentage", "status"],
        order_by="assignment_date desc",
        limit_page_length=200)
    assigned_teams = set(ta.team_id for ta in team_assignments if ta.team_id)

    all_team_ids = list(dispatch_teams | assigned_teams)
    team_details = []
    if all_team_ids:
        team_details = frappe.get_all("INET Team",
            filters={"team_id": ["in", all_team_ids]},
            fields=["team_id", "team_name", "im", "team_type", "status", "daily_cost"])
        # Enrich with assignment info
        assignment_map = {ta.team_id: ta for ta in team_assignments}
        for t in team_details:
            ta = assignment_map.get(t.team_id)
            if ta:
                t["assignment_name"] = ta.name
                t["role_in_project"] = ta.role_in_project
                t["assignment_date"] = str(ta.assignment_date) if ta.assignment_date else None
                t["end_date"] = str(ta.end_date) if ta.end_date else None
                t["assignment_status"] = ta.status
                t["utilization_percentage"] = ta.utilization_percentage

    # Financial summary
    total_po_value = sum(flt(d.line_amount) for d in dispatches)
    total_revenue = sum(flt(w.revenue_sar) for w in work_done)
    total_cost = sum(flt(w.total_cost_sar) for w in work_done)
    total_margin = sum(flt(w.margin_sar) for w in work_done)

    rollout_by_duid = _build_rollout_by_duid_groups(dispatches, plans, executions, work_done)

    return {
        "project": project,
        "dispatches": dispatches,
        "plans": plans,
        "executions": executions,
        "work_done": work_done,
        "teams": team_details,
        "team_assignments": team_assignments,
        "rollout_by_duid": rollout_by_duid,
        "financial_summary": {
            "total_po_value": total_po_value,
            "total_revenue": total_revenue,
            "total_cost": total_cost,
            "total_margin": total_margin,
            "dispatch_count": len(dispatches),
            "plan_count": len(plans),
            "execution_count": len(executions),
            "work_done_count": len(work_done),
        },
    }


# ---------------------------------------------------------------------------
# Activity Cost Master — List API
# ---------------------------------------------------------------------------

@frappe.whitelist()
def list_activity_costs():
    """Return all active Activity Cost Master records."""
    return frappe.get_all(
        "Activity Cost Master",
        filters={"active_flag": 1},
        fields=["name", "activity_code", "standard_activity", "category", "base_cost_sar"],
        order_by="activity_code asc",
        limit_page_length=500,
    )


# ---------------------------------------------------------------------------
# Execution Time Log — field time on rollout (Next PMS–style; not ERPNext Timesheet)
# ---------------------------------------------------------------------------


def _frappe_dt_to_epoch_ms(value):
    """
    Convert Frappe DB datetime (wall time in site timezone) to UTC epoch milliseconds.
    Used so portal timers stay correct for users in any region (vs parsing strings in local JS).
    """
    if value is None:
        return None
    dt = get_datetime(value)
    if not dt:
        return None
    try:
        from zoneinfo import ZoneInfo

        tzname = frappe.db.get_single_value("System Settings", "time_zone")
        if tzname and getattr(dt, "tzinfo", None) is None:
            dt = dt.replace(tzinfo=ZoneInfo(tzname))
            return int(dt.timestamp() * 1000)
    except Exception:
        pass
    return int(get_datetime(value).timestamp() * 1000)


def _session_inet_field_team_id():
    """INET Team.team_id for the logged-in field user, if any."""
    user = frappe.session.user
    if not user or user == "Guest":
        return None
    ft = frappe.get_all(
        "INET Team",
        filters={"field_user": user, "status": "Active"},
        fields=["team_id"],
        limit=1,
    )
    if ft and ft[0].team_id:
        return ft[0].team_id
    first = (frappe.db.get_value("User", user, "full_name") or "").split()[0]
    if first:
        ft2 = frappe.get_all(
            "INET Team",
            filters={"team_name": ["like", f"%{first}%"], "status": "Active"},
            fields=["team_id"],
            limit=1,
        )
        if ft2 and ft2[0].team_id:
            return ft2[0].team_id
    return None


def _assert_rollout_plan_access_field_team(team_id, rollout_plan):
    if not team_id or not rollout_plan:
        frappe.throw("Team or plan missing", frappe.PermissionError)
    if not frappe.db.exists("Rollout Plan", rollout_plan):
        frappe.throw("Rollout Plan not found")
    plan_team = frappe.db.get_value("Rollout Plan", rollout_plan, "team")
    if plan_team != team_id:
        frappe.throw("Not permitted", frappe.PermissionError)


def _im_team_ids_for_filter(im_filter=None):
    """team_id values for teams managed by the resolved IM."""
    _im_r, im_ids, _meta = resolve_im_for_session(im_filter)
    if not im_ids:
        return []
    teams = frappe.get_all(
        "INET Team",
        filters={"im": ["in", im_ids], "status": "Active"},
        fields=["team_id"],
        limit_page_length=500,
    )
    return [t.team_id for t in teams if t.team_id]


@frappe.whitelist()
def start_execution_timer(rollout_plan):
    """
    Start a timer for the given Rollout Plan (field user). One running log per user.
    """
    rollout_plan = (rollout_plan or "").strip()
    if not rollout_plan:
        frappe.throw("rollout_plan is required")

    user = frappe.session.user
    roles = set(frappe.get_roles(user))
    if "INET Field Team" not in roles and "Administrator" not in roles:
        frappe.throw("Only field team users can start execution timers", frappe.PermissionError)

    team_id = _session_inet_field_team_id()
    if not team_id:
        frappe.throw("No active field team linked to your user")

    _assert_rollout_plan_access_field_team(team_id, rollout_plan)

    plan_status = frappe.db.get_value("Rollout Plan", rollout_plan, "plan_status")
    if plan_status in ("Completed", "Cancelled"):
        frappe.throw("Cannot start a timer on a completed or cancelled plan.")

    existing = frappe.get_all(
        "Execution Time Log",
        filters={"user": user, "is_running": 1},
        fields=["name", "rollout_plan"],
        limit=1,
        ignore_permissions=True,
    )
    if existing:
        frappe.throw(
            f"You already have a running timer ({existing[0].name}). Stop it first."
        )

    log = frappe.new_doc("Execution Time Log")
    log.rollout_plan = rollout_plan
    log.team_id = team_id
    log.user = user
    log.start_time = now_datetime()
    log.is_running = 1
    log.insert(ignore_permissions=True)
    frappe.db.commit()

    server_now = now_datetime()
    return {
        "log_name": log.name,
        "rollout_plan": rollout_plan,
        "start_time": str(log.start_time),
        "server_time": str(server_now),
        "start_time_ms": _frappe_dt_to_epoch_ms(log.start_time),
        "server_time_ms": _frappe_dt_to_epoch_ms(server_now),
        "user": user,
    }


@frappe.whitelist()
def stop_execution_timer(log_name):
    """Stop a running execution timer and persist duration."""
    log_name = (log_name or "").strip()
    if not log_name:
        frappe.throw("log_name is required")

    log = frappe.get_doc("Execution Time Log", log_name)
    if not log.is_running:
        frappe.throw("This timer is not running.")

    user = frappe.session.user
    roles = set(frappe.get_roles(user))
    can_stop = log.user == user or "Administrator" in roles or "System Manager" in roles or "INET Admin" in roles
    if not can_stop:
        frappe.throw("You can only stop your own timers.", frappe.PermissionError)

    log.end_time = now_datetime()
    log.is_running = 0
    start = get_datetime(log.start_time)
    end = get_datetime(log.end_time)
    diff_seconds = time_diff_in_seconds(end, start)
    log.duration_minutes = int(diff_seconds / 60)
    log.duration_hours = round(diff_seconds / 3600, 2)
    log.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "log_name": log.name,
        "rollout_plan": log.rollout_plan,
        "duration_minutes": log.duration_minutes,
        "duration_hours": log.duration_hours,
    }


@frappe.whitelist()
def get_running_execution_timer():
    """Current user's running Execution Time Log, if any."""
    user = frappe.session.user
    if not user or user == "Guest":
        return None

    running = frappe.get_all(
        "Execution Time Log",
        filters={"user": user, "is_running": 1},
        fields=["name", "rollout_plan", "start_time", "team_id"],
        limit=1,
        ignore_permissions=True,
    )
    if not running:
        return None
    r = running[0]
    desc = frappe.db.get_value(
        "Rollout Plan",
        r.rollout_plan,
        ["po_dispatch"],
        as_dict=True,
    )
    item_hint = ""
    if desc and desc.po_dispatch:
        item_hint = frappe.db.get_value("PO Dispatch", desc.po_dispatch, "item_description") or ""

    server_now = now_datetime()
    elapsed_seconds = None
    try:
        now_ms = _frappe_dt_to_epoch_ms(server_now)
        st_ms = _frappe_dt_to_epoch_ms(r.start_time)
        if now_ms is not None and st_ms is not None:
            elapsed_seconds = int(max(0, (now_ms - st_ms) // 1000))
    except Exception:
        elapsed_seconds = None
    return {
        "log_name": r.name,
        "rollout_plan": r.rollout_plan,
        "start_time": str(r.start_time) if r.start_time else None,
        "server_time": str(server_now),
        "start_time_ms": _frappe_dt_to_epoch_ms(r.start_time),
        "server_time_ms": _frappe_dt_to_epoch_ms(server_now),
        "elapsed_seconds": elapsed_seconds,
        "team_id": r.team_id,
        "item_description": item_hint,
    }


@frappe.whitelist()
def save_execution_time_log_manual(rollout_plan, start_time, end_time, notes=None):
    """
    Create a completed time log without using the live timer (field user).
    start_time / end_time: ISO or Frappe datetime strings.
    """
    rollout_plan = (rollout_plan or "").strip()
    if not rollout_plan or not start_time or not end_time:
        frappe.throw("rollout_plan, start_time, and end_time are required")

    user = frappe.session.user
    roles = set(frappe.get_roles(user))
    if "INET Field Team" not in roles and "Administrator" not in roles:
        frappe.throw("Not permitted", frappe.PermissionError)

    team_id = _session_inet_field_team_id()
    if not team_id:
        frappe.throw("No active field team linked to your user")

    _assert_rollout_plan_access_field_team(team_id, rollout_plan)

    st = get_datetime(start_time)
    et = get_datetime(end_time)
    if et < st:
        frappe.throw("End time cannot be before start time")

    diff_seconds = time_diff_in_seconds(et, st)
    log = frappe.new_doc("Execution Time Log")
    log.rollout_plan = rollout_plan
    log.team_id = team_id
    log.user = user
    log.start_time = st
    log.end_time = et
    log.is_running = 0
    log.duration_minutes = int(diff_seconds / 60)
    log.duration_hours = round(diff_seconds / 3600, 2)
    log.notes = notes or ""
    log.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": log.name, "duration_hours": log.duration_hours}


@frappe.whitelist()
def list_execution_time_logs(filters=None, limit=100, offset=0):
    """
    List execution time logs with role-based scoping.
    filters (JSON): team_id, im, user, rollout_plan, from_date, to_date, is_running,
    search (or q) — free-text match across log, user, rollout, PO dispatch fields;
    applied server-side on the full dataset before limit/offset.
    """
    if isinstance(filters, str):
        filters = frappe.parse_json(filters or "{}")
    if not filters:
        filters = {}

    user = frappe.session.user
    roles = set(frappe.get_roles(user))
    is_desk_admin = (
        "Administrator" in roles or "System Manager" in roles or "INET Admin" in roles
    )
    is_im = "INET IM" in roles
    is_field = "INET Field Team" in roles

    db_filters = {}

    if filters.get("rollout_plan"):
        db_filters["rollout_plan"] = filters["rollout_plan"]
    if filters.get("is_running") is not None and filters.get("is_running") != "":
        db_filters["is_running"] = cint(filters["is_running"])

    # Role scoping
    if is_desk_admin:
        if filters.get("team_id"):
            db_filters["team_id"] = filters["team_id"]
        if filters.get("user"):
            db_filters["user"] = filters["user"]
    elif is_im:
        team_ids = _im_team_ids_for_filter(filters.get("im"))
        if filters.get("team_id"):
            if filters["team_id"] not in team_ids:
                return {"logs": [], "total": 0}
            db_filters["team_id"] = filters["team_id"]
        else:
            if not team_ids:
                return {"logs": [], "total": 0}
            db_filters["team_id"] = ["in", team_ids]
        if filters.get("user"):
            db_filters["user"] = filters["user"]
    elif is_field:
        db_filters["user"] = user
        ft_team = _session_inet_field_team_id()
        if ft_team:
            db_filters["team_id"] = ft_team
    else:
        frappe.throw("Not permitted", frappe.PermissionError)

    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    if from_date and to_date:
        db_filters["start_time"] = ["between", [f"{from_date} 00:00:00", f"{to_date} 23:59:59"]]
    elif from_date:
        db_filters["start_time"] = [">=", f"{from_date} 00:00:00"]
    elif to_date:
        db_filters["start_time"] = ["<=", f"{to_date} 23:59:59"]

    search_term_etl = filters.get("search") or filters.get("q") or ""
    like_tokens_etl = _sql_like_tokens(search_term_etl)
    lim_etl = _portal_row_limit(limit, 100)
    off_etl = cint(offset)

    if like_tokens_etl:
        wheres = ["1=1"]
        params = []
        joins = (
            "LEFT JOIN `tabRollout Plan` rp ON rp.name = etl.rollout_plan "
            "LEFT JOIN `tabPO Dispatch` pd ON pd.name = rp.po_dispatch "
            "LEFT JOIN `tabUser` u ON u.name = etl.user"
        )
        if filters.get("rollout_plan"):
            wheres.append("etl.rollout_plan = %s")
            params.append(filters["rollout_plan"])
        if filters.get("is_running") is not None and filters.get("is_running") != "":
            wheres.append("etl.is_running = %s")
            params.append(cint(filters["is_running"]))

        if is_desk_admin:
            if filters.get("team_id"):
                wheres.append("etl.team_id = %s")
                params.append(filters["team_id"])
            if filters.get("user"):
                wheres.append("etl.user = %s")
                params.append(filters["user"])
        elif is_im:
            team_ids_sq = _im_team_ids_for_filter(filters.get("im"))
            if filters.get("team_id"):
                if filters["team_id"] not in team_ids_sq:
                    return {"logs": [], "total": 0}
                wheres.append("etl.team_id = %s")
                params.append(filters["team_id"])
            else:
                if not team_ids_sq:
                    return {"logs": [], "total": 0}
                ph_sq = ", ".join(["%s"] * len(team_ids_sq))
                wheres.append(f"etl.team_id IN ({ph_sq})")
                params.extend(team_ids_sq)
            if filters.get("user"):
                wheres.append("etl.user = %s")
                params.append(filters["user"])
        elif is_field:
            wheres.append("etl.user = %s")
            params.append(user)
            ft_team_sq = _session_inet_field_team_id()
            if ft_team_sq:
                wheres.append("etl.team_id = %s")
                params.append(ft_team_sq)

        if from_date and to_date:
            wheres.append("etl.start_time BETWEEN %s AND %s")
            params.extend([f"{from_date} 00:00:00", f"{to_date} 23:59:59"])
        elif from_date:
            wheres.append("etl.start_time >= %s")
            params.append(f"{from_date} 00:00:00")
        elif to_date:
            wheres.append("etl.start_time <= %s")
            params.append(f"{to_date} 23:59:59")

        concat_etl = (
            "CONCAT_WS(' ', IFNULL(etl.name,''), IFNULL(etl.rollout_plan,''), IFNULL(etl.team_id,''), "
            "IFNULL(etl.user,''), IFNULL(u.full_name,''), IFNULL(etl.notes,''), IFNULL(pd.project_code,''), "
            "IFNULL(pd.item_description,''), IFNULL(pd.site_name,''), IFNULL(pd.po_no,''))"
        )
        ors_etl = " OR ".join([f"{concat_etl} LIKE %s"] * len(like_tokens_etl))
        wheres.append(f"({ors_etl})")
        params.extend(like_tokens_etl)

        wc = " AND ".join(wheres)
        total = int(
            frappe.db.sql(
                f"SELECT COUNT(*) FROM `tabExecution Time Log` etl {joins} WHERE {wc}",
                tuple(params),
            )[0][0]
        )
        if lim_etl:
            etl_page_sql = f" LIMIT {int(lim_etl)} OFFSET {int(off_etl)}"
        elif off_etl:
            etl_page_sql = f" LIMIT 18446744073709551615 OFFSET {int(off_etl)}"
        else:
            etl_page_sql = ""
        logs = frappe.db.sql(
            f"""
            SELECT etl.name, etl.rollout_plan, etl.team_id, etl.user, etl.start_time, etl.end_time,
                   etl.duration_hours, etl.duration_minutes, etl.is_running, etl.notes
            FROM `tabExecution Time Log` etl
            {joins}
            WHERE {wc}
            ORDER BY etl.start_time DESC
            {etl_page_sql}
            """,
            tuple(params),
            as_dict=True,
        ) or []
    else:
        total = frappe.db.count("Execution Time Log", db_filters) or 0

        ga_etl = dict(
            filters=db_filters,
            fields=[
                "name",
                "rollout_plan",
                "team_id",
                "user",
                "start_time",
                "end_time",
                "duration_hours",
                "duration_minutes",
                "is_running",
                "notes",
            ],
            order_by="start_time desc",
            limit_start=off_etl,
            ignore_permissions=True,
        )
        if lim_etl:
            ga_etl["limit_page_length"] = lim_etl
        logs = frappe.get_all("Execution Time Log", **ga_etl)

    # Batch-enrich: two bulk fetches replace 2 queries per row (plan + dispatch).
    plan_ids = list({r.get("rollout_plan") for r in logs if r.get("rollout_plan")})
    plan_map = {}
    if plan_ids:
        for p in frappe.get_all(
            "Rollout Plan",
            filters={"name": ["in", plan_ids]},
            fields=["name", "po_dispatch", "plan_date", "visit_type", "plan_status"],
            limit_page_length=len(plan_ids) + 1,
        ):
            plan_map[p.name] = p
    dispatch_ids = list({p.po_dispatch for p in plan_map.values() if p.po_dispatch})
    dispatch_map = {}
    if dispatch_ids:
        d_fields = ["name", "item_description", "project_code", "site_name"]
        if frappe.db.has_column("PO Dispatch", "poid"):
            d_fields.append("poid")
        for d in frappe.get_all(
            "PO Dispatch",
            filters={"name": ["in", dispatch_ids]},
            fields=d_fields,
            limit_page_length=len(dispatch_ids) + 1,
        ):
            dispatch_map[d.name] = d

    for row in logs:
        raw_start = row.get("start_time")
        raw_end = row.get("end_time")
        if row.get("is_running") and raw_start:
            now_ms = _frappe_dt_to_epoch_ms(now_datetime())
            st_ms = _frappe_dt_to_epoch_ms(raw_start)
            if now_ms is not None and st_ms is not None:
                row["elapsed_seconds"] = int(max(0, (now_ms - st_ms) // 1000))
        row["start_time"] = str(raw_start) if raw_start else None
        row["end_time"] = str(raw_end) if raw_end else None
        # get_cached_value uses the in-process Frappe cache, so this is cheap
        row["user_full_name"] = (
            frappe.get_cached_value("User", row.get("user"), "full_name") or row.get("user")
        )
        pd = plan_map.get(row.get("rollout_plan"))
        if pd:
            row["plan_date"] = str(pd.plan_date) if pd.plan_date else None
            row["visit_type"] = pd.visit_type
            row["plan_status"] = pd.plan_status
            if pd.po_dispatch:
                disp = dispatch_map.get(pd.po_dispatch)
                if disp:
                    row["item_description"] = disp.item_description
                    row["project_code"] = disp.project_code
                    row["site_name"] = disp.site_name
                    row["poid"] = (disp.get("poid") or disp.name) if disp else None
                    row["system_id"] = pd.po_dispatch

    return {"logs": logs, "total": total}


# ---------------------------------------------------------------------------
# Timesheet APIs — ERPNext Timesheet (legacy; portal uses Execution Time Log)
# ---------------------------------------------------------------------------

@frappe.whitelist()
def create_timesheet(payload):
    """
    Create an ERPNext Timesheet with time_logs.

    payload = {
        "employee": "HR-EMP-00001" or team name,
        "team": "Team-01",
        "time_logs": [
            {
                "activity_type": "...",
                "from_time": "2026-04-05 08:00:00",
                "to_time": "2026-04-05 17:00:00",
                "hours": 9,
                "project": "PRJ-001",
                "description": "..."
            }
        ]
    }
    """
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)

    team = payload.get("team")
    time_logs = payload.get("time_logs") or []

    if not time_logs:
        frappe.throw("At least one time log is required.")

    doc = frappe.new_doc("Timesheet")
    doc.company = frappe.defaults.get_global_default("company") or "INET"

    # Try to resolve employee from team
    employee = payload.get("employee")
    if not employee and frappe.db.exists("DocType", "Employee"):
        employee = frappe.db.get_value("Employee", {"user_id": frappe.session.user}, "name")
    if not employee and team:
        # Look up employee linked to this team (if any)
        emp = frappe.db.get_value("Employee", {"employee_name": ["like", f"%{team}%"]}, "name")
        if emp:
            employee = emp

    if employee:
        doc.employee = employee

    # Custom field to track which INET team submitted
    if hasattr(doc, "custom_inet_team"):
        doc.custom_inet_team = team

    for log in time_logs:
        doc.append("time_logs", {
            "activity_type": log.get("activity_type", "Execution"),
            "from_time": log.get("from_time"),
            "to_time": log.get("to_time"),
            "hours": flt(log.get("hours", 0)),
            "project": log.get("project"),
            "description": log.get("description", ""),
        })

    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    return {"name": doc.name, "status": doc.status}


@frappe.whitelist()
def list_timesheets(filters=None):
    """
    List timesheets with optional filters.

    filters = {
        "team": "Team-01",
        "im": "Ajmal",
        "from_date": "2026-04-01",
        "to_date": "2026-04-30",
        "status": "Draft",
    }
    """
    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
    if not filters:
        filters = {}

    db_filters = {}
    if filters.get("status"):
        db_filters["docstatus"] = 1 if filters["status"] == "Submitted" else 0

    include_bounds = filters.pop("include_log_bounds", None) or filters.pop("include_log_times", None)

    timesheets = frappe.get_all(
        "Timesheet",
        filters=db_filters,
        fields=[
            "name", "employee", "employee_name", "company",
            "total_hours", "total_billable_hours", "total_billed_hours",
            "status", "start_date", "end_date", "creation", "modified",
        ],
        order_by="modified desc",
        limit_page_length=200,
    )

    if include_bounds and timesheets and frappe.db.exists("DocType", "Timesheet Detail"):
        names = [t.name for t in timesheets]
        try:
            ph = ", ".join(["%s"] * len(names))
            bounds = frappe.db.sql(
                f"""
                SELECT parent, MIN(from_time) AS log_start, MAX(to_time) AS log_end
                FROM `tabTimesheet Detail`
                WHERE parent IN ({ph}) AND parenttype = 'Timesheet'
                GROUP BY parent
                """,
                tuple(names),
                as_dict=True,
            )
            bmap = {b.parent: b for b in (bounds or [])}
            for ts in timesheets:
                b = bmap.get(ts.name)
                if b:
                    ts["log_start"] = b.log_start
                    ts["log_end"] = b.log_end
        except Exception:
            pass

    # Filter by date range if provided
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")

    if from_date or to_date:
        filtered = []
        for ts in timesheets:
            sd = str(ts.start_date) if ts.start_date else ""
            if from_date and sd < from_date:
                continue
            if to_date and sd > to_date:
                continue
            filtered.append(ts)
        timesheets = filtered

    # If team filter, fetch time_logs for each and check project/team link
    team_filter = filters.get("team")
    im_filter = filters.get("im")

    if team_filter or im_filter:
        team_ids = set()
        if im_filter:
            _im_r, im_ids, _ = resolve_im_for_session(im_filter)
            lookup = im_ids or [im_filter]
            im_teams = frappe.get_all(
                "INET Team",
                filters={"im": ["in", lookup], "status": "Active"},
                fields=["team_id", "team_name"],
            )
            for t in im_teams:
                if t.get("team_id"):
                    team_ids.add(t["team_id"])
                if t.get("team_name"):
                    team_ids.add(t["team_name"])
        if team_filter:
            team_ids.add(team_filter)

        if im_filter and not team_ids:
            return []

        if team_ids:
            filtered = []
            for ts in timesheets:
                emp_name = (ts.employee_name or "").lower()
                if any(str(tid).lower() in emp_name for tid in team_ids if tid):
                    filtered.append(ts)
            timesheets = filtered

    return timesheets


@frappe.whitelist()
def approve_timesheet(name):
    """Submit/approve a timesheet."""
    doc = frappe.get_doc("Timesheet", name)
    if doc.docstatus == 0:
        doc.submit()
        frappe.db.commit()
    return {"name": doc.name, "status": "Submitted"}


@frappe.whitelist()
def get_timesheet_detail(name):
    """Get full timesheet with time_logs."""
    doc = frappe.get_doc("Timesheet", name)
    return doc.as_dict()


# Fields the IM is allowed to edit on their own INET Team. team_id / im /
# team_type / subcontractor / field_user / isdp_account / daily_cost are
# admin-only — IMs only manage display name, status, members and notes.
_IM_TEAM_EDITABLE_FIELDS = (
    "team_name",
    "status",
    "note",
)


@frappe.whitelist()
def update_im_team(name, payload=None):
    """Update an INET Team owned by the logged-in IM.

    Only the IM whose identifiers match the team's `im` field may edit, and
    only the safelist of fields above. ``team_members`` (when provided)
    replaces the child table; expects ``[{employee, designation, is_team_lead}, ...]``.
    """
    if not name:
        frappe.throw("name is required")
    if isinstance(payload, str):
        payload = frappe.parse_json(payload) if payload else {}
    payload = payload or {}

    _im_resolved, im_identifiers = _require_inet_im_session()

    team = frappe.db.get_value("INET Team", name, ["name", "im"], as_dict=True)
    if not team:
        frappe.throw(f"INET Team not found: {name}")
    if (team.im or "") not in im_identifiers:
        frappe.throw("Not permitted: this team is not assigned to you", frappe.PermissionError)

    members_raw = payload.get("team_members")
    members_changed = members_raw is not None
    clean = []

    updates = {}
    for fname in _IM_TEAM_EDITABLE_FIELDS:
        if fname not in payload:
            continue
        val = payload.get(fname)
        if fname == "daily_cost":
            try:
                val = flt(val)
            except Exception:
                continue
        elif fname == "daily_cost_applies":
            val = 1 if val in (1, "1", True, "true", "yes", "on") else 0
        elif isinstance(val, str):
            val = val.strip()
        updates[fname] = val

    if updates:
        frappe.db.set_value("INET Team", name, updates, update_modified=True)

    if members_changed:
        seen_emp = set()
        team_leads = 0
        for m in (members_raw or []):
            emp = (m or {}).get("employee")
            if not emp:
                continue
            emp = str(emp).strip()
            if not emp or emp in seen_emp:
                continue
            seen_emp.add(emp)
            is_lead = 1 if (m.get("is_team_lead") in (1, "1", True, "true", "yes", "on")) else 0
            if is_lead:
                team_leads += 1
            desig = (m.get("designation") or "").strip() or None
            clean.append({
                "employee": emp,
                "designation": desig,
                "is_team_lead": is_lead,
            })
        if team_leads > 1:
            frappe.throw("Only one team member can be marked as Team Lead.")

        # Replace the child table via the document API so Frappe handles
        # name allocation, validate hooks (e.g. lead-syncs-field_user) and
        # parent linkage automatically.
        doc = frappe.get_doc("INET Team", name)
        doc.set("team_members", [])
        for row in clean:
            doc.append("team_members", row)
        doc.save(ignore_permissions=True)

    frappe.db.commit()
    return {
        "name": name,
        "updated": len(updates),
        "fields": list(updates.keys()),
        "members_replaced": members_changed,
        "member_count": len(clean) if members_changed else None,
    }


@frappe.whitelist()
def get_im_team_detail(name):
    """Return one INET Team's data including members for the edit modal.
    Restricted to teams owned by the logged-in IM."""
    if not name:
        frappe.throw("name is required")
    _im_resolved, im_identifiers = _require_inet_im_session()
    team = frappe.db.get_value("INET Team", name, ["name", "im"], as_dict=True)
    if not team or (team.im or "") not in im_identifiers:
        frappe.throw("Not permitted: this team is not assigned to you", frappe.PermissionError)
    doc = frappe.get_doc("INET Team", name).as_dict()
    members = []
    for m in (doc.get("team_members") or []):
        emp_full = ""
        if m.get("employee"):
            emp_full = frappe.db.get_value("Employee", m.get("employee"), "employee_name") or ""
        members.append({
            "employee": m.get("employee"),
            "employee_name": emp_full,
            "designation": m.get("designation") or "",
            "is_team_lead": 1 if m.get("is_team_lead") else 0,
        })
    return {
        "name": doc.get("name"),
        "team_id": doc.get("team_id"),
        "team_name": doc.get("team_name"),
        "team_type": doc.get("team_type"),
        "im": doc.get("im"),
        "subcontractor": doc.get("subcontractor"),
        "isdp_account": doc.get("isdp_account"),
        "field_user": doc.get("field_user"),
        "status": doc.get("status"),
        "daily_cost": doc.get("daily_cost"),
        "daily_cost_applies": doc.get("daily_cost_applies"),
        "note": doc.get("note"),
        "team_members": members,
    }


@frappe.whitelist()
def list_employees_for_picker(search=None, limit=50):
    """Lightweight Employee picker for the IM team-edit modal."""
    _require_inet_im_session()
    if search:
        s = f"%{(search or '').strip()}%"
        emps = frappe.db.sql(
            """
            SELECT name, employee_name, designation
            FROM `tabEmployee`
            WHERE status = 'Active'
              AND (name LIKE %s OR IFNULL(employee_name,'') LIKE %s)
            ORDER BY employee_name ASC
            LIMIT %s
            """,
            (s, s, int(limit)),
            as_dict=True,
        )
    else:
        emps = frappe.get_all(
            "Employee",
            filters={"status": "Active"},
            fields=["name", "employee_name", "designation"],
            order_by="employee_name asc",
            limit_page_length=int(limit),
        )
    return emps or []


# ──────────────────────────────────────────────────────────────────────
# POID-level remarks (general / manager / team_lead) — role-scoped
# ──────────────────────────────────────────────────────────────────────
_REMARK_TYPES = {"general", "manager", "team_lead"}


def _resolve_dispatch_for_remarks(po_dispatch):
    """Find the PO Dispatch row for a POID or system_id and return its name."""
    name = (po_dispatch or "").strip()
    if not name:
        frappe.throw("po_dispatch is required")
    if frappe.db.exists("PO Dispatch", name):
        return name
    hit = frappe.db.get_value("PO Dispatch", {"poid": name}, "name")
    if hit:
        return hit
    frappe.throw(f"PO Dispatch not found: {name}")


def _user_role_class():
    """Coarse role bucket for remark visibility/edit rules."""
    roles = set(frappe.get_roles(frappe.session.user))
    if "Administrator" in roles or "System Manager" in roles or "INET Admin" in roles:
        return "pm"
    if "INET IM" in roles:
        return "im"
    if "INET Field Team" in roles:
        return "field"
    return None


@frappe.whitelist()
def get_po_remarks(po_dispatch):
    """Return the remark trio for a POID, masking fields the caller can't read.

    Visibility:
      - PM / admin: sees all three; edits any of the three
      - IM:        sees all three; edits manager + team_lead
      - Field:     sees only team_lead; edits team_lead
    """
    role = _user_role_class()
    if role is None:
        frappe.throw("Not permitted", frappe.PermissionError)
    name = _resolve_dispatch_for_remarks(po_dispatch)
    remark_cols = [c for c in ("general_remark", "manager_remark", "team_lead_remark")
                   if frappe.db.has_column("PO Dispatch", c)]
    row = frappe.db.get_value(
        "PO Dispatch", name,
        ["name", "poid", *remark_cols],
        as_dict=True,
    ) or {}

    visible = {
        "po_dispatch": row.get("name"),
        "poid": row.get("poid") or row.get("name"),
        "role": role,
    }
    if role in ("pm", "im"):
        visible["general_remark"] = row.get("general_remark") or ""
        visible["manager_remark"] = row.get("manager_remark") or ""
        visible["team_lead_remark"] = row.get("team_lead_remark") or ""
    elif role == "field":
        visible["team_lead_remark"] = row.get("team_lead_remark") or ""
    visible["editable"] = {
        # general:   PM only
        # manager:   PM + IM
        # team_lead: PM + IM + Field
        "general_remark": role == "pm",
        "manager_remark": role in ("pm", "im"),
        "team_lead_remark": role in ("pm", "im", "field"),
    }
    return visible


@frappe.whitelist()
def update_po_remark(po_dispatch, remark_type, value):
    """Set one remark on a POID. Permissions:
        - general    → PM
        - manager    → PM + IM
        - team_lead  → PM + IM + Field
    """
    role = _user_role_class()
    if role is None:
        frappe.throw("Not permitted", frappe.PermissionError)
    name = _resolve_dispatch_for_remarks(po_dispatch)
    rtype = (remark_type or "").strip().lower()
    if rtype not in _REMARK_TYPES:
        frappe.throw("Invalid remark_type")
    field = f"{rtype}_remark"

    allowed = (
        (rtype == "general" and role == "pm")
        or (rtype == "manager" and role in ("pm", "im"))
        or (rtype == "team_lead" and role in ("pm", "im", "field"))
    )
    if not allowed:
        frappe.throw("Not permitted to edit this remark", frappe.PermissionError)

    text = str(value or "")
    if len(text) > 8000:
        text = text[:8000]
    if not frappe.db.has_column("PO Dispatch", field):
        frappe.throw(
            f"Field {field} is not yet available on this site — run "
            "`bench --site <site> migrate` to add the role-scoped remark columns.",
            frappe.ValidationError,
        )
    frappe.db.set_value("PO Dispatch", name, field, text, update_modified=True)
    frappe.db.commit()
    return {"po_dispatch": name, "remark_type": rtype, "value": text}


# ──────────────────────────────────────────────────────────────────────
# Sub-Contract flow — IM-driven, lives outside the rollout chain
# ──────────────────────────────────────────────────────────────────────
def _can_subcon_dispatch(role, im_identifiers, pd):
    if role == "pm":
        return True
    if role != "im":
        return False
    pd_im = (pd.get("im") or "")
    if pd_im and pd_im not in (im_identifiers or []):
        return False
    return True


@frappe.whitelist()
def get_my_subcon_capability(im=None):
    """Return whether the current session can sub-contract POIDs.

    PM/admin: always True. IM: True only if `IM Master.can_subcon = 1`.
    Field: never.
    """
    role = _user_role_class()
    if role == "pm":
        return {"role": role, "can_subcon": True, "im": None}
    if role != "im":
        return {"role": role, "can_subcon": False, "im": None}
    im_resolved, im_identifiers, _ = resolve_im_for_session(im)
    if not im_identifiers:
        return {"role": role, "can_subcon": False, "im": None}
    # `im_resolved` is the IM Master name; `im_identifiers` is an unordered
    # set-derived list (may contain email / full_name / etc.). Probe both.
    target = im_resolved if im_resolved and frappe.db.exists("IM Master", im_resolved) else None
    if not target:
        for ident in im_identifiers:
            if frappe.db.exists("IM Master", ident):
                target = ident
                break
    flag = 0
    if target:
        flag = cint(frappe.db.get_value("IM Master", target, "can_subcon") or 0)
    return {"role": role, "can_subcon": bool(flag), "im": target}


@frappe.whitelist()
def list_subcon_teams_for_picker(search=None, limit=200):
    """Sub-Contract Team picker — only teams with category 'Sub-Contract Team'."""
    role = _user_role_class()
    if role not in ("pm", "im"):
        frappe.throw("Not permitted", frappe.PermissionError)
    s = (search or "").strip()
    where = ["IFNULL(team_category,'Field Team') = 'Sub-Contract Team'", "IFNULL(status,'Active') = 'Active'"]
    params = []
    if s:
        where.append("(name LIKE %s OR IFNULL(team_id,'') LIKE %s OR IFNULL(team_name,'') LIKE %s)")
        like = f"%{s}%"
        params.extend([like, like, like])
    sql = f"""
        SELECT name, team_id, team_name, subcontractor, department
        FROM `tabINET Team`
        WHERE {' AND '.join(where)}
        ORDER BY IFNULL(team_name, team_id) ASC
        LIMIT {int(limit)}
    """
    return frappe.db.sql(sql, tuple(params), as_dict=True)


def _assign_subcon_one(role, im_identifiers, name, team, remark):
    """Stamp subcon fields on a single PO Dispatch. Returns (ok, info_or_error)."""
    pd = frappe.db.get_value(
        "PO Dispatch", name,
        ["name", "im", "dispatch_status", "subcon_status", "subcon_team", "is_dummy_po", "poid"],
        as_dict=True,
    ) or {}
    if not pd.get("name"):
        return False, {"po_dispatch": name, "error": "PO Dispatch not found"}

    cur_status = (pd.get("dispatch_status") or "")
    if cur_status == "Planned":
        return False, {"po_dispatch": name, "poid": pd.get("poid") or name, "error": "Already in Rollout Planning — cancel the plan first."}
    if cur_status == "Completed":
        return False, {"po_dispatch": name, "poid": pd.get("poid") or name, "error": "Already completed."}
    if cur_status == "Cancelled":
        return False, {"po_dispatch": name, "poid": pd.get("poid") or name, "error": "Cancelled — cannot sub-contract."}
    if (pd.get("subcon_status") or "") == "Pending":
        return False, {"po_dispatch": name, "poid": pd.get("poid") or name, "error": "Already sub-contracted (pending)."}
    if (pd.get("is_dummy_po") or 0) and not (pd.get("poid") or "").strip():
        return False, {"po_dispatch": name, "poid": pd.get("poid") or name, "error": "Map the dummy PO to a real PO line before sub-contracting."}

    if role == "im":
        if not _can_subcon_dispatch(role, im_identifiers, pd):
            return False, {"po_dispatch": name, "poid": pd.get("poid") or name, "error": "Not assigned to you."}

    updates = {
        "subcon_team": team["name"],
        "subcon_status": "Pending",
        "subcon_completed_on": None,
        "dispatch_status": "Sub-Contracted",
    }
    if remark is not None and str(remark or "").strip():
        updates["subcon_remark"] = str(remark or "")[:8000]
    frappe.db.set_value("PO Dispatch", name, updates, update_modified=True)
    return True, {
        "po_dispatch": name,
        "poid": pd.get("poid") or name,
        "subcon_team": team["name"],
        "subcon_team_name": team.get("team_name") or team.get("team_id"),
        "subcon_status": "Pending",
        "dispatch_status": "Sub-Contracted",
    }


@frappe.whitelist()
def assign_subcon(po_dispatch=None, po_dispatches=None, subcon_team=None, remark=None):
    """Sub-contract one or many PO Dispatches to a non-field team.

    Accepts either ``po_dispatch`` (single name) or ``po_dispatches`` (list / JSON
    array). The list form is preferred for bulk actions from the UI.

    Side-effects per dispatch:
        subcon_team           = <team>
        subcon_status         = 'Pending'
        subcon_completed_on   = NULL
        dispatch_status       = 'Sub-Contracted'
        subcon_remark         = <remark> (optional)

    NOTE: subcon dispatches are intentionally kept OUT of the rollout chain — no
    Rollout Plan, Daily Execution or Work Done rows are created. Reporting that
    needs them must read the PO Dispatch row directly.
    """
    role = _user_role_class()
    if role not in ("pm", "im"):
        frappe.throw("Not permitted", frappe.PermissionError)
    if not subcon_team:
        frappe.throw("subcon_team is required")

    raw = po_dispatches if po_dispatches not in (None, "", []) else po_dispatch
    if isinstance(raw, str):
        try:
            parsed = frappe.parse_json(raw)
            if isinstance(parsed, (list, tuple)):
                raw = parsed
        except Exception:
            pass
    if isinstance(raw, (list, tuple)):
        candidates = [str(x).strip() for x in raw if str(x or "").strip()]
    elif raw:
        candidates = [str(raw).strip()]
    else:
        candidates = []
    if not candidates:
        frappe.throw("po_dispatch is required")

    team = frappe.db.get_value(
        "INET Team", subcon_team,
        ["name", "team_id", "team_name", "team_category", "status"],
        as_dict=True,
    )
    if not team:
        frappe.throw(f"INET Team not found: {subcon_team}")
    if (team.get("team_category") or "Field Team") != "Sub-Contract Team":
        frappe.throw("Sub-contracting is only allowed to teams with category 'Sub-Contract Team'.")
    if (team.get("status") or "Active") != "Active":
        frappe.throw("Selected team is not Active.")

    im_identifiers = []
    if role == "im":
        im_resolved, im_identifiers, _ = resolve_im_for_session()
        # Gate by IM Master.can_subcon. resolve_im_for_session returns
        # `im_resolved` = IM Master name; `im_identifiers` is an unordered
        # set-derived list that can include email / full_name / etc., so
        # always look up the IM Master row by the resolved name (or by any
        # identifier that maps to one).
        my_im = im_resolved if im_resolved and frappe.db.exists("IM Master", im_resolved) else None
        if not my_im:
            for ident in im_identifiers or []:
                if frappe.db.exists("IM Master", ident):
                    my_im = ident
                    break
        if not my_im:
            frappe.throw("Could not resolve your IM Master record.", frappe.PermissionError)
        if not cint(frappe.db.get_value("IM Master", my_im, "can_subcon") or 0):
            frappe.throw("Sub-contracting is not enabled for your IM profile. Ask the admin to enable 'Can Sub-Contract' on your IM Master.", frappe.PermissionError)

    # Resolve POIDs / system_ids to real names
    resolved = []
    for token in candidates:
        try:
            resolved.append(_resolve_dispatch_for_remarks(token))
        except Exception:
            resolved.append(None)

    updated = []
    errors = []
    for i, name in enumerate(resolved):
        if not name:
            errors.append({"po_dispatch": candidates[i], "error": "PO Dispatch not found"})
            continue
        ok, info = _assign_subcon_one(role, im_identifiers, name, team, remark)
        (updated if ok else errors).append(info)

    frappe.db.commit()
    return {
        "updated": updated,
        "errors": errors,
        "summary": {
            "total": len(candidates),
            "updated_count": len(updated),
            "error_count": len(errors),
            "subcon_team": team["name"],
            "subcon_team_name": team.get("team_name") or team.get("team_id"),
        },
    }


def _mark_subcon_done_one(role, im_identifiers, name, completed, remark):
    pd = frappe.db.get_value(
        "PO Dispatch", name,
        ["name", "im", "dispatch_status", "subcon_status", "subcon_team", "subcon_remark", "poid"],
        as_dict=True,
    ) or {}
    if not pd.get("name"):
        return False, {"po_dispatch": name, "error": "PO Dispatch not found"}
    if (pd.get("subcon_status") or "") != "Pending":
        return False, {"po_dispatch": name, "poid": pd.get("poid") or name, "error": "Not in 'Sub-Contract Pending' state."}
    if role == "im" and not _can_subcon_dispatch(role, im_identifiers, pd):
        return False, {"po_dispatch": name, "poid": pd.get("poid") or name, "error": "Not assigned to you."}

    updates = {
        "subcon_status": "Work Done",
        "subcon_completed_on": completed,
        # Keep dispatch_status='Sub-Contracted' through the Work Done step. The
        # subcon flow lives outside the rollout chain — same way regular rows
        # keep dispatch_status='Planned' even after Daily Execution completes.
        # Completion is conveyed by subcon_status='Work Done' and the execution
        # status surfaced in the Work Done feed.
    }
    addition = (str(remark or "")).strip() if remark is not None else ""
    if addition:
        existing = (pd.get("subcon_remark") or "").strip()
        if existing:
            combined = existing + "\n— Work Done note —\n" + addition
        else:
            combined = addition
        updates["subcon_remark"] = combined[:8000]
    frappe.db.set_value("PO Dispatch", name, updates, update_modified=True)
    return True, {
        "po_dispatch": name,
        "poid": pd.get("poid") or name,
        "subcon_status": "Work Done",
        "subcon_completed_on": str(completed),
        "dispatch_status": "Sub-Contracted",
    }


@frappe.whitelist()
def mark_subcon_work_done(po_dispatch=None, po_dispatches=None, completed_on=None, remark=None):
    """Mark one or many sub-contracted POIDs as Work Done.

    Accepts either ``po_dispatch`` (single name) or ``po_dispatches`` (list / JSON).
    """
    role = _user_role_class()
    if role not in ("pm", "im"):
        frappe.throw("Not permitted", frappe.PermissionError)

    raw = po_dispatches if po_dispatches not in (None, "", []) else po_dispatch
    if isinstance(raw, str):
        try:
            parsed = frappe.parse_json(raw)
            if isinstance(parsed, (list, tuple)):
                raw = parsed
        except Exception:
            pass
    if isinstance(raw, (list, tuple)):
        candidates = [str(x).strip() for x in raw if str(x or "").strip()]
    elif raw:
        candidates = [str(raw).strip()]
    else:
        candidates = []
    if not candidates:
        frappe.throw("po_dispatch is required")

    completed = completed_on or nowdate()
    try:
        completed = getdate(completed)
    except Exception:
        completed = nowdate()

    im_identifiers = []
    if role == "im":
        _, im_identifiers, _ = resolve_im_for_session()

    resolved = []
    for token in candidates:
        try:
            resolved.append(_resolve_dispatch_for_remarks(token))
        except Exception:
            resolved.append(None)

    updated = []
    errors = []
    for i, name in enumerate(resolved):
        if not name:
            errors.append({"po_dispatch": candidates[i], "error": "PO Dispatch not found"})
            continue
        ok, info = _mark_subcon_done_one(role, im_identifiers, name, completed, remark)
        (updated if ok else errors).append(info)

    frappe.db.commit()
    return {
        "updated": updated,
        "errors": errors,
        "summary": {
            "total": len(candidates),
            "updated_count": len(updated),
            "error_count": len(errors),
            "completed_on": str(completed),
        },
    }


@frappe.whitelist()
def list_subcon_dispatches(
    im=None, search=None, status="all", limit=300,
    project_code=None, site_code=None, subcon_team=None,
):
    """Sub-Contract list feed.

    status: "all" | "pending" | "done"
    Visible to PM (all IMs) and IM (own POIDs only).
    Optional filters ``project_code``, ``site_code`` and ``subcon_team`` accept
    a single value or a JSON-array / comma-separated list.
    """
    role = _user_role_class()
    if role not in ("pm", "im"):
        frappe.throw("Not permitted", frappe.PermissionError)

    where = ["IFNULL(pd.subcon_status,'') != ''"]
    params = []

    if role == "im":
        _, im_identifiers, _ = resolve_im_for_session(im)
        if not im_identifiers:
            return []
        ph = ", ".join(["%s"] * len(im_identifiers))
        where.append(f"IFNULL(pd.im,'') IN ({ph})")
        params.extend(list(im_identifiers))
    elif role == "pm" and im:
        _, im_identifiers, _ = resolve_im_for_session(im)
        if im_identifiers:
            ph = ", ".join(["%s"] * len(im_identifiers))
            where.append(f"IFNULL(pd.im,'') IN ({ph})")
            params.extend(list(im_identifiers))

    s = (status or "all").lower()
    if s == "pending":
        where.append("pd.subcon_status = 'Pending'")
    elif s in ("done", "work_done"):
        where.append("pd.subcon_status = 'Work Done'")

    for col, raw in (("project_code", project_code), ("site_code", site_code), ("subcon_team", subcon_team)):
        clause, in_params = _sql_in_or_eq(f"pd.{col}", raw)
        if clause:
            where.append(clause)
            params.extend(in_params)

    if search:
        clause, like_params = _sql_search_clause(
            "CONCAT_WS(' ', pd.poid, pd.po_no, pd.item_code, pd.item_description, pd.site_name, pd.site_code, pd.project_code, IFNULL(t.team_name,''), IFNULL(t.team_id,''))",
            search,
        )
        if clause:
            where.append(clause)
            params.extend(like_params)

    limit_int = int(limit) if limit else 300
    sql = f"""
        SELECT pd.name AS po_dispatch,
               pd.poid AS poid,
               pd.po_no, pd.po_line_no,
               pd.item_code, pd.item_description,
               pd.qty, pd.line_amount,
               pd.project_code, pd.customer, pd.im,
               pd.site_code, pd.site_name, pd.center_area, pd.region_type,
               pd.dispatch_status, pd.target_month,
               pd.subcon_team, pd.subcon_status, pd.subcon_completed_on, pd.subcon_remark,
               t.team_id   AS subcon_team_id,
               t.team_name AS subcon_team_name,
               pd.modified
        FROM `tabPO Dispatch` pd
        LEFT JOIN `tabINET Team` t ON t.name = pd.subcon_team
        WHERE {' AND '.join(where)}
        ORDER BY pd.modified DESC
        LIMIT {limit_int}
    """
    return frappe.db.sql(sql, tuple(params), as_dict=True)


# ──────────────────────────────────────────────────────────────────────
# PO Archive Import — closed/cancelled lines are imported with terminal
# status and stay out of the active workflow. Used to backfill historical
# POs from spreadsheets so the PM can dump them later.
# ──────────────────────────────────────────────────────────────────────
_PO_ARCHIVE_ALIAS = {
    "ID": "source_id",
    "Contract": "contract",
    "Project Domain": "project_domain",
    "POID": "poid",
    "PO Status": "po_status",
    "PO STATUS": "po_status",
    "Status": "po_status",
    "PO status": "po_status",
    "PO NO.": "po_no",
    "PO No.": "po_no",
    "PO No": "po_no",
    "PO Line NO.": "po_line_no",
    "PO Line No.": "po_line_no",
    "PO Line No": "po_line_no",
    "Shipment NO.": "shipment_no",
    "Shipment No.": "shipment_no",
    "Site Name": "site_name",
    "Site Code": "site_code",
    "Item Code": "item_code",
    "Item Description": "item_description",
    "Unit": "unit",
    "Requested Qty": "qty",
    "Due Qty": "due_qty",
    "Billed Quantity": "billed_quantity",
    "Quantity Cancel": "quantity_cancel",
    "Start Date": "start_date",
    "End Date": "end_date",
    "Sub Contract NO.": "sub_contract_no",
    "Currency": "currency",
    "Unit Price": "rate",
    "Line Amount": "line_amount",
    "Tax Rate": "tax_rate",
    "Payment Terms": "payment_terms",
    "Project Code": "project_code",
    "Project Name": "project_name",
    "Center Area": "center_area",
    "Publish Date": "publish_date",
    "Huawei Owner": "huawei_owner",
    "Inet Owner": "inet_owner",
    "Inet IM Remarks": "manager_remark",
    # PIC / invoice tracking columns
    "SQC Status": "sqc_status",
    "PAT Status": "pat_status",
    "IM's Remarks on Rejection": "im_rejection_remark",
    "PIC Remarks": "pic_status",
    "ISDP / I-Buy Owner": "isdp_ibuy_owner",
    "MS1": "ms1_amount",
    "MS2": "ms2_amount",
    "PIC Remarks (2nd Milestone)": "pic_status_ms2",
    "ISD / I-Buy Owner (2nd Milestone)": "isdp_owner_ms2",
    "Remaining Milestone": "remaining_milestone_pct",
    "1st Payment PO Amount": "ms1_amount",
    "1st Payment Invoiced": "ms1_invoiced",
    "Subcon Per% MS1": "subcon_pct_ms1",
    "Inet Per% MS1": "inet_pct_ms1",
    "2nd Payment PO Amount": "ms2_amount",
    "2nd Payment Invoiced": "ms2_invoiced",
    "Subcon Per% MS2": "subcon_pct_ms2",
    "Inet Per% MS2": "inet_pct_ms2",
    "Invoicing Month (First Payment Milestone)": "ms1_invoice_month",
    "Invoicing Month (Second Payment Milestone)": "ms2_invoice_month",
    # Duplicate-named cols ("Applied Date", "IBUY / INV date", "Payment Received
    # Date", "Detail Remarks/Dependency") are disambiguated positionally by
    # ``_resolve_archive_header_keys`` using the most recent "PIC Remarks"
    # anchor as the milestone marker.
}


def _resolve_archive_header_keys(headers):
    """Return ``{col_index: standard_key}`` for a Master-Tracker header row.

    The Master Tracker has several columns whose **header text repeats** —
    ``Detail Remarks/Dependency`` appears 3× (IM, MS1 PIC, MS2 PIC),
    ``Applied Date`` 2× (MS1, MS2), ``IBUY / INV date`` 2× and
    ``Payment Received Date`` 2×. A plain dict keyed by header name would
    collapse them onto the same field.

    We walk the header row left-to-right, tracking the current milestone
    context: after seeing ``PIC Remarks`` we're in MS1; after
    ``PIC Remarks (2nd Milestone)`` we're in MS2; before either anchor the
    duplicate columns are treated as IM-side and skipped.
    """
    out = {}
    milestone = 0  # 0 = pre-PIC, 1 = MS1, 2 = MS2

    DUP_BY_MS = {
        ("Detail Remarks/Dependency", 1): "pic_detail_remark",
        ("Detail Remarks/Dependency", 2): "pic_detail_remark_ms2",
        ("Applied Date", 1): "ms1_applied_date",
        ("Applied Date", 2): "ms2_applied_date",
        ("IBUY / INV date", 1): "ms1_ibuy_inv_date",
        ("IBUY / INV date", 2): "ms2_ibuy_inv_date",
        ("Payment Received Date", 1): "ms1_payment_received_date",
        ("Payment Received Date", 2): "ms2_payment_received_date",
    }
    DUP_NAMES = {h for (h, _) in DUP_BY_MS.keys()}

    for i, h in enumerate(headers or []):
        h = str(h or "").strip() if h is not None else ""
        if not h:
            continue
        if h == "PIC Remarks":
            milestone = 1
            out[i] = "pic_status"
            continue
        if h == "PIC Remarks (2nd Milestone)":
            milestone = 2
            out[i] = "pic_status_ms2"
            continue
        if h in DUP_NAMES:
            key = DUP_BY_MS.get((h, milestone))
            if key:
                out[i] = key
            # Pre-PIC duplicates (e.g. col 30 IM detail) are left unmapped.
            continue
        alias = _PO_ARCHIVE_ALIAS.get(h)
        if alias:
            out[i] = alias
    return out


def _stamp_archive_pic_fields(dispatch_name, src_line):
    """Copy PIC / invoice tracking fields from an archive row onto a PO Dispatch.

    Bypasses ``doc.save()`` (which would be expensive for 12k+ rows) and writes
    via ``frappe.db.set_value`` directly. Computes ``ms1_unbilled`` /
    ``ms2_unbilled`` here so they don't drift when validate hasn't run.
    """
    PIC_KEYS = (
        "sqc_status", "pat_status", "im_rejection_remark",
        "pic_status", "pic_status_ms2",
        "isdp_ibuy_owner", "isdp_owner_ms2",
        "pic_detail_remark", "pic_detail_remark_ms2",
        "ms1_applied_date", "ms2_applied_date",
        "ms1_amount", "ms2_amount",
        "ms1_invoiced", "ms2_invoiced",
        "subcon_pct_ms1", "inet_pct_ms1",
        "subcon_pct_ms2", "inet_pct_ms2",
        "remaining_milestone_pct",
        "ms1_invoice_month", "ms2_invoice_month",
        "ms1_ibuy_inv_date", "ms2_ibuy_inv_date",
        "ms1_payment_received_date", "ms2_payment_received_date",
        "manager_remark",
    )
    NUMERIC = {
        "ms1_amount", "ms2_amount", "ms1_invoiced", "ms2_invoiced",
        "subcon_pct_ms1", "inet_pct_ms1", "subcon_pct_ms2", "inet_pct_ms2",
        "remaining_milestone_pct",
    }
    updates = {}
    for k in PIC_KEYS:
        v = src_line.get(k)
        if v is None or v == "":
            continue
        if k in NUMERIC:
            try:
                updates[k] = flt(v)
            except Exception:
                continue
        else:
            updates[k] = str(v).strip()[:8000] if isinstance(v, str) else v
    # Always normalize ms1/ms2_unbilled when we touched the matching amount or
    # invoiced value — keeps the archive data consistent without running validate.
    if "ms1_amount" in updates or "ms1_invoiced" in updates:
        amt = updates.get("ms1_amount")
        if amt is None:
            amt = flt(frappe.db.get_value("PO Dispatch", dispatch_name, "ms1_amount") or 0)
        inv = flt(updates.get("ms1_invoiced", 0))
        updates["ms1_unbilled"] = round(flt(amt) - inv, 4)
    if "ms2_amount" in updates or "ms2_invoiced" in updates:
        amt = updates.get("ms2_amount")
        if amt is None:
            amt = flt(frappe.db.get_value("PO Dispatch", dispatch_name, "ms2_amount") or 0)
        inv = flt(updates.get("ms2_invoiced", 0))
        updates["ms2_unbilled"] = round(flt(amt) - inv, 4)
    if updates:
        frappe.db.set_value("PO Dispatch", dispatch_name, updates, update_modified=False)


def _archive_date_keys():
    """All archive-row keys whose source values are Excel date serials in xlsb."""
    return (
        "start_date",
        "end_date",
        "publish_date",
        "ms1_applied_date",
        "ms2_applied_date",
        "ms1_invoice_month",
        "ms2_invoice_month",
        "ms1_ibuy_inv_date",
        "ms2_ibuy_inv_date",
        "ms1_payment_received_date",
        "ms2_payment_received_date",
    )


def _excel_serial_to_date_str(value):
    """Excel stores dates as serials; pyxlsb returns the float. Pass-through strings."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        try:
            from datetime import datetime as _dt, timedelta as _td
            base = _dt(1899, 12, 30)
            return (base + _td(days=float(value))).strftime("%Y-%m-%d")
        except Exception:
            return None
    return str(value)


def _po_archive_resolve_local_path(file_url):
    if file_url and file_url.startswith("/private/files/"):
        return frappe.get_site_path("private", "files", file_url[len("/private/files/"):])
    if file_url and file_url.startswith("/files/"):
        return frappe.get_site_path("public", "files", file_url[len("/files/"):])
    try:
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        return file_doc.get_full_path()
    except Exception:
        frappe.throw(f"Cannot resolve file path for: {file_url}")


def _detect_archive_header_row(rows_iter):
    """Find the header row (the first row containing 'PO NO.' or 'PO Status').

    Master Tracker exports often prepend a totals row before the real headers.
    """
    headers = None
    skipped = 0
    while True:
        try:
            row = next(rows_iter)
        except StopIteration:
            break
        vals = [(str(v).strip() if v is not None else "") for v in row]
        if any(h in ("PO NO.", "PO No.", "PO No", "PO Status", "POID") for h in vals):
            headers = vals
            break
        skipped += 1
        if skipped > 8:
            break
    return headers, rows_iter


def _yield_po_archive_rows(file_path, sheet_hint="PO Tracker"):
    """Yield raw row dicts (aliased keys) from xlsb / xlsx / xlsm / csv."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsb":
        try:
            from pyxlsb import open_workbook
        except ImportError:
            frappe.throw("pyxlsb is not installed. Run `bench pip install pyxlsb`.")
        with open_workbook(file_path) as wb:
            target = sheet_hint if sheet_hint in wb.sheets else wb.sheets[0]
            with wb.get_sheet(target) as sheet:
                py_iter = ([c.v for c in r] for r in sheet.rows())
                headers, py_iter = _detect_archive_header_row(py_iter)
                if not headers:
                    return
                col_map = _resolve_archive_header_keys(headers)
                date_keys = _archive_date_keys()
                for vals in py_iter:
                    if not any(v not in (None, "") for v in vals):
                        continue
                    row_dict = {}
                    for col_idx, cell_val in enumerate(vals):
                        std_key = col_map.get(col_idx)
                        if std_key:
                            row_dict[std_key] = cell_val
                    if row_dict:
                        for k in date_keys:
                            if k in row_dict:
                                row_dict[k] = _excel_serial_to_date_str(row_dict[k])
                        # Whole-number ints come back as floats from xlsb; clean
                        # them so POIDs don't end up with `-1-1.0`.
                        for k in ("po_line_no", "shipment_no"):
                            if k in row_dict:
                                row_dict[k] = _normalize_int_token(row_dict[k])
                        yield row_dict
        return

    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            frappe.throw("openpyxl is not installed. Run `bench pip install openpyxl`.")
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_hint] if sheet_hint in wb.sheetnames else wb.active
        py_iter = (list(r) for r in ws.iter_rows(values_only=True))
        headers, py_iter = _detect_archive_header_row(py_iter)
        if not headers:
            return
        col_map = _resolve_archive_header_keys(headers)
        for vals in py_iter:
            if not any(v not in (None, "") for v in vals):
                continue
            row_dict = {}
            for col_idx, cell_val in enumerate(vals):
                std_key = col_map.get(col_idx)
                if std_key:
                    row_dict[std_key] = cell_val
            if row_dict:
                yield row_dict
        return

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        py_iter = iter(rows)
        headers, py_iter = _detect_archive_header_row(py_iter)
        if not headers:
            return
        col_map = _resolve_archive_header_keys(headers)
        for vals in py_iter:
            if not any(v not in (None, "") for v in vals):
                continue
            row_dict = {}
            for col_idx, cell_val in enumerate(vals):
                std_key = col_map.get(col_idx)
                if std_key:
                    row_dict[std_key] = cell_val
            if row_dict:
                yield row_dict
        return

    frappe.throw(f"Unsupported file type: {ext}")


def _archive_status_for(po_status):
    """Map normalized po_status → PO Dispatch dispatch_status; None for non-archive rows."""
    s = str(po_status or "").strip().upper()
    if s in ("CLOSED", "CLOSE", "COMPLETED", "DONE"):
        return "Closed"
    if s in ("CANCELLED", "CANCELED", "CANCEL"):
        return "Cancelled"
    return None


@frappe.whitelist()
def preview_po_archive_file(file_url):
    """Pre-flight scan: counts, missing UOMs / Items / Projects, and a small sample.

    The customer is *not* required at this stage — it's resolved per-line from
    the row's ``project_code`` (Project Control Center.customer) at import time.
    """
    file_path = _po_archive_resolve_local_path(file_url)
    counts = {"CLOSED": 0, "CANCELLED": 0, "OPEN": 0, "OTHER": 0}
    sample = []
    total = 0
    uoms = set()
    items = set()
    projects = set()
    projects_no_customer = set()
    rows_missing_project = 0
    for row_dict in _yield_po_archive_rows(file_path):
        total += 1
        target = _archive_status_for(row_dict.get("po_status"))
        if target == "Closed":
            counts["CLOSED"] += 1
        elif target == "Cancelled":
            counts["CANCELLED"] += 1
        else:
            ps = str(row_dict.get("po_status") or "").strip().upper()
            counts["OPEN" if ps == "OPEN" else "OTHER"] += 1
        if not target:
            continue
        if (row_dict.get("unit") or "").strip():
            uoms.add(str(row_dict["unit"]).strip())
        if (row_dict.get("item_code") or "").strip():
            items.add(str(row_dict["item_code"]).strip())
        if (row_dict.get("project_code") or "").strip():
            projects.add(str(row_dict["project_code"]).strip())
        else:
            rows_missing_project += 1
        if len(sample) < 10:
            sample.append({
                "poid": row_dict.get("poid"),
                "po_status": row_dict.get("po_status"),
                "po_no": row_dict.get("po_no"),
                "po_line_no": row_dict.get("po_line_no"),
                "item_code": row_dict.get("item_code"),
                "site_code": row_dict.get("site_code"),
                "site_name": row_dict.get("site_name"),
                "project_code": row_dict.get("project_code"),
                "project_name": row_dict.get("project_name"),
                "qty": row_dict.get("qty"),
                "rate": row_dict.get("rate"),
                "line_amount": row_dict.get("line_amount"),
                "unit": row_dict.get("unit"),
            })

    # UOM check: anything not present in `tabUOM` is missing. Compare
    # case-insensitively because the DB collation accepts `Set`/`SET` as the
    # same row, but `_bulk_existing_set` returns the canonical casing.
    missing_uoms = []
    if uoms:
        existing_uom = _bulk_existing_set("UOM", list(uoms))
        existing_uom_ci = {str(x).strip().lower() for x in existing_uom}
        missing_uoms = sorted([u for u in uoms if u.strip().lower() not in existing_uom_ci])
    # Item check (case-insensitive: same reason).
    missing_items = []
    if items:
        existing_items = _bulk_existing_set("Item", list(items))
        existing_items_ci = {str(x).strip().lower() for x in existing_items}
        missing_items = sorted([i for i in items if i.strip().lower() not in existing_items_ci])
    # Project check: which projects are missing & which exist but have no customer.
    missing_projects = []
    if projects:
        existing_projects = _bulk_existing_set("Project Control Center", list(projects))
        missing_projects = sorted([p for p in projects if p not in existing_projects])
        if existing_projects:
            ph = ", ".join(["%s"] * len(existing_projects))
            no_cust_rows = frappe.db.sql(
                f"SELECT name FROM `tabProject Control Center` "
                f"WHERE name IN ({ph}) AND IFNULL(customer,'') = ''",
                tuple(existing_projects),
            )
            projects_no_customer = sorted([r[0] for r in no_cust_rows])

    return {
        "total_rows": total,
        "counts": counts,
        "to_import": counts["CLOSED"] + counts["CANCELLED"],
        "sample": sample,
        "missing_uoms": missing_uoms,
        "missing_items": missing_items,
        "missing_projects": missing_projects,
        "projects_without_customer": list(projects_no_customer),
        "rows_missing_project_code": rows_missing_project,
        "unique_projects": len(projects),
        "unique_uoms": len(uoms),
        "unique_items": len(items),
    }


def _archive_log_update(log_name, **fields):
    if not log_name or not frappe.db.exists("PO Upload Log", log_name):
        return
    fields = {k: v for k, v in fields.items() if v is not None}
    if not fields:
        return
    frappe.db.set_value("PO Upload Log", log_name, fields, update_modified=True)
    frappe.db.commit()


def _bulk_existing_set(doctype, names):
    if not names:
        return set()
    names = list({str(n).strip() for n in names if str(n or "").strip()})
    if not names:
        return set()
    ph = ", ".join(["%s"] * len(names))
    rows = frappe.db.sql(
        f"SELECT name FROM `tab{doctype}` WHERE name IN ({ph})",
        tuple(names),
    )
    return {r[0] for r in rows}


def _run_po_archive_import(file_url, customer, log_name, chunk_size=200):
    """Background worker: import CLOSED/CANCELLED rows from an archive file.
    Updates the PO Upload Log row with progress; final status Completed/Partial/Failed.
    No auto-dispatch — these rows stay out of the workflow.

    ``customer`` is optional. When empty, each row's customer is resolved from
    its ``project_code`` → ``Project Control Center.customer``. Rows whose
    project has no customer are recorded as failed in the per-PO breakdown.
    """
    try:
        _archive_log_update(log_name, status="Running")
        # Customer is now optional at the upload level — resolved per project.
        fallback_cust = _resolve_customer_link_name(customer) if customer else None

        file_path = _po_archive_resolve_local_path(file_url)
        groups = {}  # keyed by (po_no, customer)
        total_seen = 0
        total_archive = 0
        rows_missing_customer = 0
        unresolved_projects = set()

        # Build a project → customer map up-front
        first_pass_projects = set()
        all_rows = []
        for row in _yield_po_archive_rows(file_path):
            total_seen += 1
            target = _archive_status_for(row.get("po_status"))
            if not target:
                continue
            row["_target_status"] = target
            all_rows.append(row)
            if (row.get("project_code") or "").strip():
                first_pass_projects.add(str(row["project_code"]).strip())
        proj_cust_map = {}
        if first_pass_projects:
            ph = ", ".join(["%s"] * len(first_pass_projects))
            for r in frappe.db.sql(
                f"SELECT name, customer FROM `tabProject Control Center` WHERE name IN ({ph})",
                tuple(first_pass_projects), as_dict=True,
            ) or []:
                if r.customer:
                    proj_cust_map[r.name] = r.customer

        for row in all_rows:
            po_no = str(row.get("po_no") or "").strip()
            if not po_no:
                continue
            proj = (row.get("project_code") or "").strip()
            cust = proj_cust_map.get(proj) or fallback_cust
            if not cust:
                rows_missing_customer += 1
                if proj:
                    unresolved_projects.add(proj)
                # Skip — will surface in summary
                continue
            groups.setdefault((po_no, cust), []).append(row)
            total_archive += 1

        if total_archive == 0 and rows_missing_customer == 0:
            _archive_log_update(
                log_name, status="Completed",
                total_rows=total_seen, lines_imported=0, lines_skipped=0,
                po_created=0, po_updated=0,
                notes="No CLOSED or CANCELLED rows found in file.",
            )
            return
        if total_archive == 0 and rows_missing_customer > 0:
            _archive_log_update(
                log_name, status="Failed",
                total_rows=total_seen, lines_imported=0,
                lines_skipped=rows_missing_customer,
                po_created=0, po_updated=0,
                notes=(
                    f"All {rows_missing_customer} archive rows lack a resolvable customer "
                    f"(missing on Project Control Center). Set customer on these projects first: "
                    f"{', '.join(sorted(unresolved_projects)[:30])}"
                ),
            )
            return

        # Pre-resolve master existence so we don't run per-row checks.
        all_projects = {str(l.get("project_code") or "").strip() for ls in groups.values() for l in ls if l.get("project_code")}
        all_items = {str(l.get("item_code") or "").strip() for ls in groups.values() for l in ls if l.get("item_code")}
        all_duids = {str(l.get("site_code") or "").strip() for ls in groups.values() for l in ls if l.get("site_code")}
        all_uoms = {str(l.get("unit") or "").strip() for ls in groups.values() for l in ls if l.get("unit")}
        existing_projects = _bulk_existing_set("Project Control Center", list(all_projects))
        existing_items = _bulk_existing_set("Item", list(all_items))
        existing_duids = _bulk_existing_set("DUID Master", list(all_duids))
        existing_uoms = _bulk_existing_set("UOM", list(all_uoms))
        # Case-insensitive lookup: file may say `Set`, DB says `SET`.
        existing_uoms_ci = {str(x).strip().lower(): str(x) for x in existing_uoms}
        cim_done = set()

        def _ensure_project(line, resolved_cust):
            code = str(line.get("project_code") or "").strip()
            if not code:
                return None
            if code in existing_projects:
                return code
            try:
                ok, _ = ensure_project_control_center(
                    code, resolved_cust, line.get("project_name"), line.get("center_area"),
                )
                if ok:
                    existing_projects.add(code)
                    return code
            except Exception:
                pass
            return None

        def _ensure_item(line):
            code = str(line.get("item_code") or "").strip()
            if not code or code in existing_items:
                return
            try:
                ok, _ = ensure_item_master(code, line.get("item_description"))
                if ok:
                    existing_items.add(code)
            except Exception:
                pass

        def _ensure_duid(line):
            code = str(line.get("site_code") or "").strip()
            if not code or code in existing_duids:
                return
            try:
                ok, _ = ensure_duid_master(code, line.get("site_name"), line.get("center_area"))
                if ok:
                    existing_duids.add(code)
            except Exception:
                pass

        po_created = 0
        po_updated = 0
        lines_imported = 0
        lines_skipped = 0
        lines_failed = 0
        per_po = []

        group_keys = list(groups.keys())
        for chunk_idx in range(0, len(group_keys), chunk_size):
            chunk = group_keys[chunk_idx: chunk_idx + chunk_size]
            for key in chunk:
                po_no, group_customer = key
                lines = groups[key]
                statuses = {l.get("_target_status") for l in lines}
                hdr_status = "CANCELLED" if statuses == {"Cancelled"} else "CLOSED"

                existing_name = frappe.db.get_value("PO Intake", {"po_no": po_no}, "name")
                existing_poids = set()
                if existing_name:
                    rows_existing = frappe.db.sql(
                        "SELECT po_line_no, shipment_number, poid FROM `tabPO Intake Line` WHERE parent=%s",
                        existing_name, as_dict=True,
                    ) or []
                    for r in rows_existing:
                        pid = (r.poid or "").strip() or _make_poid(po_no, r.po_line_no, r.shipment_number)
                        existing_poids.add(pid)

                new_entries = []
                po_skipped = 0
                for line in lines:
                    poid = _poid_for_upload_line(po_no, line)
                    if poid in existing_poids:
                        po_skipped += 1
                        continue

                    # Item code is mandatory on PO Intake Line for the standard
                    # workflow, but archive imports may include legacy rows where
                    # it's blank or set to "NA". Fall back to the description so
                    # ``ensure_item_master`` can create something usable; if even
                    # the description is empty we rely on ``flags.ignore_mandatory``
                    # below to let the row through.
                    item_code, used_desc = _resolve_item_code_with_fallback(line)
                    if used_desc and item_code:
                        line["item_code"] = item_code
                        if not str(line.get("item_description") or "").strip():
                            line["item_description"] = item_code

                    proj = _ensure_project(line, group_customer)
                    _ensure_item(line)
                    _ensure_duid(line)

                    if item_code:
                        ck = (group_customer, item_code)
                        if ck not in cim_done:
                            try:
                                ensure_customer_item_master(group_customer, item_code)
                            except Exception:
                                pass
                            cim_done.add(ck)

                    # UOM may not exist; setting an unknown link will fail
                    # validation. Probe case-insensitively and use the canonical
                    # DB name when found.
                    raw_uom = (str(line.get("unit") or "").strip() or None)
                    uom_value = existing_uoms_ci.get(raw_uom.lower()) if raw_uom else None

                    line_center_area = line.get("center_area")
                    line_region = region_type_from_center_area(line_center_area)
                    desc = line.get("item_description")
                    if isinstance(desc, str):
                        desc = desc[:8000]  # Small Text upper bound, generous

                    append_row = {
                        "source_id": line.get("source_id"),
                        "po_line_no": cint(line.get("po_line_no") or 0),
                        "shipment_number": _normalize_int_token(line.get("shipment_no")) or None,
                        "poid": poid,
                        "site_code": str(line.get("site_code") or "").strip() or None,
                        "site_name": line.get("site_name"),
                        "item_code": item_code,
                        "item_description": desc,
                        "uom": uom_value,
                        "qty": flt(line.get("qty", 0)),
                        "due_qty": flt(line.get("due_qty", 0)),
                        "billed_quantity": flt(line.get("billed_quantity", 0)),
                        "quantity_cancel": flt(line.get("quantity_cancel", 0)),
                        "start_date": line.get("start_date") or None,
                        "end_date": line.get("end_date") or None,
                        "sub_contract_no": line.get("sub_contract_no"),
                        "currency": line.get("currency"),
                        "rate": flt(line.get("rate", 0)),
                        "line_amount": flt(line.get("line_amount", 0)),
                        "tax_rate": line.get("tax_rate"),
                        "payment_terms": line.get("payment_terms"),
                        "project_code": proj,
                        "project_name": line.get("project_name"),
                        "center_area": line_center_area,
                        "region_type": line_region,
                        "publish_date": line.get("publish_date") or None,
                        "po_line_status": line.get("_target_status"),
                        "dispatch_mode": "Manual",
                    }
                    new_entries.append((line, append_row))
                    existing_poids.add(poid)

                if not new_entries:
                    lines_skipped += po_skipped
                    if po_skipped > 0:
                        per_po.append({
                            "po_no": po_no, "intake_name": existing_name,
                            "lines_added": 0, "lines_skipped": po_skipped, "is_new": False,
                        })
                    continue

                try:
                    if existing_name:
                        doc = frappe.get_doc("PO Intake", existing_name)
                        doc.status = _normalize_po_intake_status(hdr_status)
                        for _src, append_row in new_entries:
                            doc.append("po_lines", append_row)
                        # Archive: legacy lines may have blank item_code / qty.
                        doc.flags.ignore_mandatory = True
                        doc.save(ignore_permissions=True)
                        is_new_po = False
                        po_updated += 1
                    else:
                        doc = frappe.new_doc("PO Intake")
                        doc.po_no = po_no
                        doc.customer = group_customer
                        doc.status = _normalize_po_intake_status(hdr_status)
                        doc.publish_date = lines[0].get("publish_date") or None
                        doc.center_area = lines[0].get("center_area")
                        for _src, append_row in new_entries:
                            doc.append("po_lines", append_row)
                        doc.flags.ignore_mandatory = True
                        doc.insert(ignore_permissions=True)
                        is_new_po = True
                        po_created += 1
                except Exception as e:
                    frappe.db.rollback()
                    # Capture the FULL error text so the user can read it from the log,
                    # along with which row triggered it where determinable.
                    err_msg = frappe.utils.cstr(e)[:1000]
                    lines_failed += len(new_entries)
                    per_po.append({
                        "po_no": po_no, "intake_name": existing_name,
                        "lines_added": 0, "lines_skipped": len(new_entries),
                        "is_new": not bool(existing_name), "error": err_msg,
                    })
                    continue

                child_sql_rows = frappe.db.sql(
                    "SELECT name, po_line_no, shipment_number, poid FROM `tabPO Intake Line` WHERE parent=%s",
                    doc.name, as_dict=True,
                ) or []
                new_poid_map = {ar["poid"]: ar for _src, ar in new_entries}
                for r in child_sql_rows:
                    got_poid = (r.poid or "").strip() or _make_poid(po_no, r.po_line_no, r.shipment_number)
                    if got_poid not in new_poid_map:
                        continue
                    line_dict = dict(new_poid_map[got_poid])
                    line_dict["name"] = r.name
                    line_dict["parent"] = doc.name
                    target_status = next(
                        (l["_target_status"] for l in lines if _poid_for_upload_line(po_no, l) == got_poid),
                        "Closed",
                    )
                    try:
                        dispatch_name = _upsert_po_dispatch_for_line(
                            doc.name, po_no, line_dict,
                            customer=group_customer,
                            dispatch_status=target_status,
                            dispatch_mode="Manual",
                        )
                        # Stamp the rich PIC / invoice-tracking fields from
                        # the source archive row onto the just-upserted dispatch.
                        src_line = next(
                            (l for l in lines if _poid_for_upload_line(po_no, l) == got_poid),
                            None,
                        )
                        if dispatch_name and src_line:
                            _stamp_archive_pic_fields(dispatch_name, src_line)
                    except Exception as e:
                        per_po.append({
                            "po_no": po_no, "intake_name": doc.name,
                            "lines_added": 0, "lines_skipped": 1,
                            "is_new": False, "error": f"dispatch error: {frappe.utils.cstr(e)[:800]}",
                        })

                lines_imported += len(new_entries)
                lines_skipped += po_skipped
                per_po.append({
                    "po_no": po_no, "intake_name": doc.name,
                    "lines_added": len(new_entries),
                    "lines_skipped": po_skipped,
                    "is_new": is_new_po,
                })
                frappe.db.commit()

            _archive_log_update(
                log_name,
                total_rows=total_archive,
                lines_imported=lines_imported,
                lines_skipped=lines_skipped,
                po_created=po_created,
                po_updated=po_updated,
                notes=f"Processed {min(chunk_idx + chunk_size, len(group_keys))}/{len(group_keys)} POs",
            )

        try:
            log_doc = frappe.get_doc("PO Upload Log", log_name)
            log_doc.set("po_details", [])
            for d in per_po[:5000]:
                if d.get("error"):
                    detail_status = "Failed"
                elif d.get("lines_added"):
                    detail_status = "New" if d.get("is_new") else "Appended"
                else:
                    detail_status = "Duplicate"
                log_doc.append("po_details", {
                    "po_no": d.get("po_no"),
                    "intake_name": d.get("intake_name"),
                    "status": detail_status,
                    "lines_added": d.get("lines_added") or 0,
                    "lines_skipped": d.get("lines_skipped") or 0,
                    "error": d.get("error") or "",
                })
            log_doc.status = "Completed" if lines_failed == 0 else "Partial"
            log_doc.total_rows = total_archive
            log_doc.lines_imported = lines_imported
            log_doc.lines_skipped = lines_skipped + lines_failed
            log_doc.po_created = po_created
            log_doc.po_updated = po_updated
            log_doc.auto_dispatched = 0
            log_doc.notes = (
                f"Archive import: {total_archive} archive rows; "
                f"{lines_imported} imported, {lines_skipped} skipped (duplicate), "
                f"{lines_failed} failed."
            )
            log_doc.save(ignore_permissions=True)
            frappe.db.commit()
        except Exception:
            _archive_log_update(
                log_name, status="Partial",
                total_rows=total_archive,
                lines_imported=lines_imported,
                lines_skipped=lines_skipped + lines_failed,
                po_created=po_created, po_updated=po_updated,
                notes="Imported but failed to write per-PO breakdown.",
            )
    except Exception as e:
        frappe.log_error(title="PO archive import failed", message=frappe.get_traceback())
        _archive_log_update(log_name, status="Failed", notes=f"Error: {str(e)[:240]}")


@frappe.whitelist()
def start_po_archive_import(file_url, customer=None):
    """Kick off an archive import in the background. Returns a log_name to poll.

    ``customer`` is optional — if not given, each row's customer is resolved
    from its ``project_code``. The file is attached to the PO Upload Log doc
    so admins can re-download what was imported.
    """
    if not file_url:
        frappe.throw("file_url is required")
    file_path = _po_archive_resolve_local_path(file_url)
    if not os.path.exists(file_path):
        frappe.throw(f"File not found on server: {file_url}")
    resolved_cust = None
    if customer:
        resolved_cust = _resolve_customer_link_name(customer)
        if not resolved_cust:
            frappe.throw(f"Customer not found: {customer}")

    log = frappe.new_doc("PO Upload Log")
    log.uploaded_by = frappe.session.user
    log.uploaded_at = frappe.utils.now_datetime()
    log.customer = resolved_cust
    log.file_name = os.path.basename(file_path)
    log.file_url = file_url
    log.upload_mode = "Archive"
    log.status = "Queued"
    log.total_rows = 0
    log.lines_imported = 0
    log.lines_skipped = 0
    log.po_created = 0
    log.po_updated = 0
    log.auto_dispatched = 0
    log.insert(ignore_permissions=True)

    # Attach the source file to the log so it stays linked.
    try:
        attach = frappe.get_doc({
            "doctype": "File",
            "file_url": file_url,
            "file_name": os.path.basename(file_path),
            "attached_to_doctype": "PO Upload Log",
            "attached_to_name": log.name,
            "is_private": 1 if "/private/" in (file_url or "") else 0,
        })
        attach.insert(ignore_permissions=True, ignore_links=True)
    except Exception:
        # Non-fatal: the import still runs; the file_url is on the log doc.
        pass

    frappe.db.commit()

    frappe.enqueue(
        "inet_app.api.command_center._run_po_archive_import",
        queue="long", timeout=3600,
        file_url=file_url, customer=resolved_cust, log_name=log.name,
    )
    return {"log_name": log.name, "status": log.status}


@frappe.whitelist()
def get_po_archive_import_status(log_name):
    """Lightweight status snapshot for the FE poller."""
    if not frappe.db.exists("PO Upload Log", log_name):
        frappe.throw(f"PO Upload Log not found: {log_name}")
    row = frappe.db.get_value(
        "PO Upload Log", log_name,
        ["status", "total_rows", "lines_imported", "lines_skipped",
         "po_created", "po_updated", "notes", "uploaded_at", "file_name"],
        as_dict=True,
    ) or {}
    row["log_name"] = log_name
    return row
