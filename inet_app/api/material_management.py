"""Material Management API — Huawei Outbound Import, Material Requests, Stock Entries."""
import os
import re

import frappe
from frappe.utils import flt, now, nowdate


# Regex: "JDL Outbound Plan of HUAWEI CWH For 10th_May_2026.xlsx"
_FILENAME_DATE_RE = re.compile(
    r"For\s+(\d{1,2})(?:st|nd|rd|th)?[_\s]+(\w+)[_\s]+(\d{4})",
    re.IGNORECASE,
)

_MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


def _parse_outbound_date(filename):
    """Extract outbound date from filename like '...For 10th_May_2026.xlsx'."""
    m = _FILENAME_DATE_RE.search(filename)
    if not m:
        return nowdate()
    day = int(m.group(1))
    month_name = m.group(2).lower()
    year = int(m.group(3))
    month = _MONTH_MAP.get(month_name, 1)
    return f"{year}-{month:02d}-{day:02d}"


def _resolve_file_path(file_url):
    """Resolve a Frappe file_url to an absolute filesystem path."""
    from frappe.utils.file_manager import get_file_path
    try:
        path = get_file_path(file_url)
        if path and os.path.exists(path):
            return path
    except Exception:
        pass
    if file_url.startswith("/private/files/"):
        return os.path.join(frappe.get_site_path("private", "files"), file_url[len("/private/files/"):])
    elif file_url.startswith("/files/"):
        return os.path.join(frappe.get_site_path("public", "files"), file_url[len("/files/"):])
    return file_url


@frappe.whitelist()
def start_huawei_outbound_import(name):
    """Start processing a Huawei Outbound Import record (called from form button)."""
    frappe.only_for(["System Manager", "Stock Manager"])
    doc = frappe.get_doc("Huawei Outbound Import", name)
    if doc.status not in ("Draft", "Failed"):
        frappe.throw(f"Cannot start import in status '{doc.status}'.")

    frappe.db.set_value("Huawei Outbound Import", name, "status", "Processing")
    frappe.db.commit()

    try:
        file_path = _resolve_file_path(doc.file)
        result = import_huawei_outbound_from_doc(file_path, doc.outbound_date)

        inet_bills = result.get("inet_bills", [])
        frappe.db.set_value("Huawei Outbound Import", name, {
            "status": "Completed",
            "total_rows": result.get("total_rows", 0),
            "new_rows": result.get("new_rows", 0),
            "inet_count": result.get("inet_count", 0),
            "duplicates_skipped": result.get("duplicates_skipped", 0),
        })
        frappe.db.commit()

        summary = f"Total: {result.get('total_rows', 0)} | New: {result.get('new_rows', 0)} | INET: {result.get('inet_count', 0)} | Dups: {result.get('duplicates_skipped', 0)}"
        summary += f"\nProjects: {result.get('project_matched', 0)} matched, {result.get('project_missing', 0)} not found"
        frappe.msgprint(summary, title="Import Summary", indicator="green")

        if inet_bills:
            lines = "\n".join(f"{b['bill_no']} | {b['du_id']} | Vol: {b['total_volume']}" for b in inet_bills[:50])
            frappe.msgprint(f"<pre>INET Bills ({len(inet_bills)}):\n{lines}</pre>", title="INET Bills in This Import")

        return {"status": "Completed", "inet_count": len(inet_bills)}

    except Exception as e:
        frappe.db.set_value("Huawei Outbound Import", name, {
            "status": "Failed",
            "error_message": str(e)[:5000],
        })
        frappe.db.commit()
        frappe.log_error(frappe.get_traceback(), "Huawei Outbound Import failed")
        raise


def import_huawei_outbound_from_doc(file_path=None, outbound_date=None):
    """Import from a local file path (called from Huawei Outbound Import doctype).

    Returns dict with counts — does NOT throw on duplicate rows, just skips them.
    """
    # Resolve file path: Frappe stores file_url like "/private/files/x.xlsx"
    if file_path and (file_path.startswith("/private/files/") or file_path.startswith("/files/")):
        if file_path.startswith("/private/files/"):
            file_path = os.path.join(frappe.get_site_path("private", "files"), file_path[len("/private/files/"):])
        else:
            file_path = os.path.join(frappe.get_site_path("public", "files"), file_path[len("/files/"):])

    if not file_path or not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    filename = os.path.basename(file_path)
    if not outbound_date:
        outbound_date = _parse_outbound_date(filename)

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "orderQuery" not in wb.sheetnames:
        raise ValueError(f"Sheet 'orderQuery' not found. Available: {', '.join(wb.sheetnames)}")

    ws = wb["orderQuery"]
    if ws.max_row < 2:
        raise ValueError("File has no data rows")

    header = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(1, col).value or "").strip()
        if val:
            header[val] = col

    required = ["Bill No.", "Request No.", "Subcon"]
    for req in required:
        if req not in header:
            raise ValueError(f"Required column '{req}' not found in Excel. Found: {list(header.keys())}")

    # Warn if optional columns are missing
    optional = ["Project Name", "DU ID", "Customer Site ID", "Delivery Purpose", "Status"]
    missing_optional = [c for c in optional if c not in header]
    if missing_optional:
        frappe.msgprint(f"Optional columns not found in Excel: {', '.join(missing_optional)}. These fields will be empty.", title="Missing Columns", indicator="blue")

    def _cell(row, name):
        c = header.get(name)
        if c is None:
            return ""
        v = ws.cell(row, c).value
        return str(v).strip() if v is not None else ""

    import_batch = f"{filename} ({frappe.utils.nowdate()})"
    total_rows = 0
    new_rows = 0
    duplicates = 0
    inet_count = 0
    inet_bills = []
    project_matched = 0
    project_missing = 0

    for row in range(2, ws.max_row + 1):
        bill_no = _cell(row, "Bill No.")
        if not bill_no:
            continue
        total_rows += 1

        if frappe.db.exists("Huawei Outbound Plan", bill_no):
            duplicates += 1
            continue

        subcon_name = _cell(row, "Subcon")
        du_id = _cell(row, "DU ID")
        project = _cell(row, "Project Name")
        row_status = _cell(row, "Status") or "Prepared"

        # Ensure Subcon Master
        if subcon_name and not frappe.db.exists("Huawei Subcon Master", subcon_name):
            try:
                frappe.get_doc({
                    "doctype": "Huawei Subcon Master",
                    "subcon_name": subcon_name,
                    "status": "Active",
                }).insert(ignore_permissions=True)
            except Exception:
                pass

        # Ensure DUID Master
        if du_id and not frappe.db.exists("DUID Master", du_id):
            try:
                frappe.get_doc({
                    "doctype": "DUID Master",
                    "duid": du_id,
                }).insert(ignore_permissions=True)
            except Exception:
                pass

        # Look up Project Control Center — first by name (project_code), then by project_name field
        project_link = None
        if project:
            if frappe.db.exists("Project Control Center", project):
                project_link = project
                project_matched += 1
            elif frappe.db.exists("Project Control Center", {"project_name": project}):
                project_link = frappe.db.get_value("Project Control Center", {"project_name": project}, "name")
                project_matched += 1
            else:
                project_missing += 1
        else:
            project_missing += 1

        # Look up DUID Master
        duid_link = None
        if du_id and frappe.db.exists("DUID Master", du_id):
            duid_link = du_id

        is_inet = subcon_name.strip().upper() == "INET"

        frappe.get_doc({
            "doctype": "Huawei Outbound Plan",
            "bill_no": bill_no,
            "request_no": _cell(row, "Request No."),
            "outbound_date": outbound_date,
            "project_name": project or "",
            "project": project_link or None,
            "subcon": subcon_name if subcon_name else None,
            "outbound_status": row_status,
            "du_id": du_id or "",
            "duid_master": duid_link or None,
            "customer_site_id": _cell(row, "Customer Site ID"),
            "delivery_purpose": _cell(row, "Delivery Purpose"),
            "total_volume": flt(_cell(row, "Total Volume")),
            "import_batch": import_batch,
        }).insert(ignore_permissions=True)
        new_rows += 1
        if is_inet:
            inet_count += 1
            inet_bills.append({
                "bill_no": bill_no,
                "request_no": _cell(row, "Request No."),
                "du_id": du_id,
                "total_volume": _cell(row, "Total Volume"),
            })

    frappe.db.commit()
    return {
        "total_rows": total_rows,
        "new_rows": new_rows,
        "duplicates_skipped": duplicates,
        "inet_count": inet_count,
        "inet_bills": inet_bills,
        "project_matched": project_matched,
        "project_missing": project_missing,
        "outbound_date": str(outbound_date),
        "filename": filename,
    }


@frappe.whitelist()
def create_material_receipt_from_outbound(bill_no):
    """Open a new Stock Entry (Material Receipt) linked to a Huawei Outbound Plan.

    Returns a redirect URL to the new Stock Entry form. Items and qty are entered
    manually by the user. The Outbound Plan link is pre-filled on the Stock Entry.
    """
    roles = set(frappe.get_roles(frappe.session.user))
    if not roles & {"Administrator", "System Manager", "Stock Manager"}:
        frappe.throw("Only Stock Manager or Administrator can create material receipts.", frappe.PermissionError)

    if not frappe.db.exists("Huawei Outbound Plan", bill_no):
        frappe.throw(f"Huawei Outbound Plan {bill_no} not found.")

    plan = frappe.get_doc("Huawei Outbound Plan", bill_no)
    if (plan.subcon or "").strip().upper() != "INET":
        frappe.throw("Material Receipt can only be created for INET items.")

    if plan.material_receipt:
        frappe.throw(f"Material Receipt already exists: {plan.material_receipt}")

    # Prefer the Link field (duid_master) over the Data field (du_id); the Link
    # field is always the canonical DUID Master record name. Fall back to du_id
    # for plans imported before duid_master was populated.
    du_id = (plan.duid_master or plan.du_id or "").strip()

    new_se_url = (
        f"/app/stock-entry/new-stock-entry"
        f"?stock_entry_type=Material Receipt"
        f"&huawei_outbound_plan={bill_no}"
        f"&duid={du_id}"
    )
    return {
        "redirect_url": new_se_url,
        "du_id": du_id,
    }


def before_stock_entry_submit(doc, method=None):
    """Fill DUID inventory dimension on Stock Entry items following the
    correct direction per entry type:

    - Material Receipt  → to_duid only   (stock arriving at a location)
    - Material Transfer → duid + to_duid (stock moving between locations)
    - Material Issue    → duid only      (stock leaving a location)

    Source is the linked Material Request's duid field, falling back to
    the PO Dispatch site_code via mr.poid.
    """
    se_type = doc.stock_entry_type
    if se_type not in ("Material Receipt", "Material Transfer", "Material Issue"):
        return

    set_source = se_type in ("Material Transfer", "Material Issue")
    set_target = se_type in ("Material Receipt", "Material Transfer")

    # Batch-resolve MR → DUID to avoid N+1 queries
    mr_names = list({item.get("material_request") for item in doc.items if item.get("material_request")})
    mr_duid_map = {}
    if mr_names:
        for mr_row in frappe.db.get_all(
            "Material Request",
            filters={"name": ["in", mr_names]},
            fields=["name", "duid", "poid"],
            ignore_permissions=True,
        ):
            duid = (mr_row.get("duid") or "").strip()
            if not duid:
                poid = mr_row.get("poid")
                if poid:
                    duid = (frappe.db.get_value("PO Dispatch", poid, "site_code") or "").strip()
            mr_duid_map[mr_row["name"]] = duid

    # For Material Receipt from Huawei Outbound Plan, use plan's DUID
    plan_duid = ""
    if se_type == "Material Receipt" and doc.get("huawei_outbound_plan"):
        plan = frappe.db.get_value(
            "Huawei Outbound Plan",
            doc.huawei_outbound_plan,
            ["duid_master", "du_id"],
            as_dict=True,
        ) or {}
        plan_duid = (plan.get("duid_master") or plan.get("du_id") or "").strip()

    for item in doc.items:
        mr_name = item.get("material_request")
        duid = mr_duid_map.get(mr_name) or plan_duid or ""
        if not duid:
            continue
        if set_source and not item.get("duid"):
            item.duid = duid
        if set_target and not item.get("to_duid"):
            item.to_duid = duid


def on_stock_entry_submit(doc, method=None):
    """When a Material Receipt is submitted, link the Huawei Outbound Plan
    and set its status to Received.

    Primary: uses huawei_outbound_plan field on the Stock Entry header.
    Fallback: matches by duid field on items for receipts without a header link.
    """
    if doc.stock_entry_type != "Material Receipt":
        return

    plan_name = (doc.get("huawei_outbound_plan") or "").strip()
    if plan_name and frappe.db.exists("Huawei Outbound Plan", plan_name):
        subcon = frappe.db.get_value("Huawei Outbound Plan", plan_name, "subcon")
        if (subcon or "").strip().upper() == "INET":
            frappe.db.set_value("Huawei Outbound Plan", plan_name, {
                "material_receipt": doc.name,
                "outbound_status": "Received",
            })
            frappe.db.commit()
        return

    # Fallback: match unlinked INET plans by duid on items
    du_ids = {(item.duid or "").strip() for item in doc.items if (item.duid or "").strip()}
    if not du_ids:
        return

    for du_id in du_ids:
        plans = frappe.db.get_all("Huawei Outbound Plan",
            filters={
                "du_id": du_id,
                "subcon": "INET",
                "material_receipt": ["in", ["", None]],
            },
            pluck="name",
        )
        for plan_name in plans:
            frappe.db.set_value("Huawei Outbound Plan", plan_name, {
                "material_receipt": doc.name,
                "outbound_status": "Received",
            })

    frappe.db.commit()


# ─── Phase 2: Material Request — uses standard ERPNext Material Request ───────

def _request_status(status, transfer_status):
    """Map ERPNext status + transfer_status to a portal-friendly label."""
    if status == "Cancelled":
        return "Rejected"
    if transfer_status == "Completed":
        return "Transferred"
    if status in ("Submitted", "Pending") and transfer_status in ("Not Started", "", None):
        return "Pending Approval"
    return status or "Draft"


@frappe.whitelist()
def list_material_requests(im=None, status=None, limit=50):
    """List Material Requests (type: Material Transfer) created via INET portal.

    IM users see only their own requests (filtered by im custom field).
    Stock Manager / Admin see all.
    """
    roles = set(frappe.get_roles(frappe.session.user))
    is_admin = bool(roles & {"Administrator", "System Manager", "Stock Manager", "INET Admin"})

    filters = {"material_request_type": "Material Transfer"}
    try:
        if frappe.db.has_column("Material Request", "is_return_request"):
            filters["is_return_request"] = ["!=", 1]
    except Exception:
        pass
    if not is_admin:
        im_name = frappe.db.get_value("IM Master", {"user": frappe.session.user}, "name")
        if im_name:
            filters["im"] = im_name
        else:
            return []

    if im:
        filters["im"] = im

    rows = frappe.db.get_all(
        "Material Request",
        filters=filters,
        fields=[
            "name", "transaction_date", "owner", "im", "poid", "duid",
            "status", "transfer_status", "set_warehouse", "set_from_warehouse",
        ],
        order_by="`tabMaterial Request`.transaction_date desc, `tabMaterial Request`.creation desc",
        limit=int(limit),
    )
    # Collect unique system-name poids and resolve to business POIDs in one batch
    poid_links = list({r["poid"] for r in rows if r.get("poid")})
    poid_map = {}
    if poid_links:
        for row in frappe.db.get_all(
            "PO Dispatch",
            filters={"name": ["in", poid_links]},
            fields=["name", "poid"],
        ):
            poid_map[row["name"]] = row["poid"]

    for r in rows:
        r["request_date"] = str(r.pop("transaction_date", "") or "")
        r["request_status"] = _request_status(r["status"], r["transfer_status"])
        r["team_warehouse"] = r.pop("set_warehouse", "")
        r["source_warehouse"] = r.pop("set_from_warehouse", "")
        if r.get("poid"):
            r["poid"] = poid_map.get(r["poid"], r["poid"])

    if status:
        rows = [r for r in rows if r["request_status"] == status]
    return rows


@frappe.whitelist()
def get_material_request(name):
    """Get a single Material Request with its items (permission-aware)."""
    roles = set(frappe.get_roles(frappe.session.user))
    doc = frappe.get_doc("Material Request", name)

    is_admin = bool(roles & {"Administrator", "System Manager", "Stock Manager", "INET Admin"})
    if not is_admin:
        im_name = frappe.db.get_value("IM Master", {"user": frappe.session.user}, "name")
        if doc.get("im") != im_name and doc.owner != frappe.session.user:
            frappe.throw("Not permitted.", frappe.PermissionError)

    # Find linked Stock Entries (transfer and issue)
    ses = frappe.db.get_all(
        "Stock Entry",
        filters={"material_request": name},
        fields=["name", "stock_entry_type"],
        order_by="`tabStock Entry`.creation asc",
    )
    transfer_se = next((s["name"] for s in ses if s["stock_entry_type"] == "Material Transfer"), None)
    issue_se = next((s["name"] for s in ses if s["stock_entry_type"] == "Material Issue"), None)

    # Resolve Link value (system name) → business POID
    poid_link = doc.get("poid") or ""
    poid_display = frappe.db.get_value("PO Dispatch", poid_link, "poid") if poid_link else ""

    return {
        "name": doc.name,
        "request_date": str(doc.transaction_date or ""),
        "owner": doc.owner,
        "im": doc.get("im"),
        "poid": poid_display or poid_link,
        "duid": doc.get("duid"),
        "request_status": _request_status(doc.status, doc.transfer_status),
        "status": doc.status,
        "transfer_status": doc.transfer_status,
        "team_warehouse": doc.set_warehouse,
        "source_warehouse": doc.set_from_warehouse,
        "rejection_reason": doc.get("rejection_reason"),
        "stock_entry_transfer": transfer_se,
        "stock_entry_issue": issue_se,
        "items": [
            {
                "item_code": i.item_code,
                "item_name": i.item_name,
                "qty": i.qty,
                "uom": i.uom or i.stock_uom,
                "duid": i.get("duid"),
                "poid": i.get("poid"),
            }
            for i in doc.items
        ],
    }


def _resolve_po_dispatch(poid_input):
    """Return the PO Dispatch system name given a business POID or system name.

    PO Dispatch uses system-generated names (SYS-…) while business users
    always reference the ``poid`` field (e.g. W-4178-ATN-01-REL-01).
    Try business poid first, fall back to document name.
    """
    name = frappe.db.get_value("PO Dispatch", {"poid": poid_input}, "name")
    if not name and frappe.db.exists("PO Dispatch", poid_input):
        name = poid_input
    return name


@frappe.whitelist()
def get_poid_details(poid):
    """Fetch PO Dispatch details for auto-fill in new request form.

    Accepts the business POID (poid field) or the system document name.
    Also returns the team from the latest Rollout Plan and the source
    warehouse from INET Settings.
    """
    name = _resolve_po_dispatch(poid)
    if not name:
        frappe.throw(f"PO Dispatch with POID '{poid}' not found.")
    row = frappe.db.get_value(
        "PO Dispatch", name,
        ["name", "poid", "site_code", "im", "project_code"],
        as_dict=True,
    )
    row["doc_name"] = row.pop("name")  # system name for Link field storage
    # Team from latest non-cancelled Rollout Plan for this POID
    plans = frappe.db.get_all(
        "Rollout Plan",
        filters={"po_dispatch": row["doc_name"], "plan_status": ["!=", "Cancelled"]},
        fields=["team"],
        order_by="creation desc",
        limit=1,
    )
    team = plans[0]["team"] if plans else ""

    # Fallback: any rollout plan (including cancelled) if nothing found above
    if not team:
        plans_any = frappe.db.get_all(
            "Rollout Plan",
            filters={"po_dispatch": row["doc_name"]},
            fields=["team"],
            order_by="creation desc",
            limit=1,
        )
        team = plans_any[0]["team"] if plans_any else ""

    row["team"] = team or ""
    row["team_warehouse"] = frappe.db.get_value("INET Team", team, "warehouse") or "" if team else ""
    # Source warehouse from INET Settings
    row["source_warehouse"] = frappe.db.get_single_value("INET Settings", "source_warehouse") or ""
    return row


@frappe.whitelist()
def get_source_warehouse():
    """Return the configured main store warehouse from INET Settings."""
    return frappe.db.get_single_value("INET Settings", "source_warehouse") or ""


@frappe.whitelist()
def get_duid_received_items(duid):
    """Return items received in the main warehouse for a DUID.

    Looks up Huawei Outbound Plans (INET, Received) for this DUID,
    then reads the Stock Entry Detail rows from each Material Receipt.
    These are customer-provided (Huawei) items — valuation_rate is 0.
    """
    if not duid:
        return []
    plans = frappe.db.get_all(
        "Huawei Outbound Plan",
        filters={"du_id": duid, "subcon": "INET", "material_receipt": ["is", "set"]},
        fields=["name", "material_receipt"],
    )
    if not plans:
        return []

    items = []
    seen = set()
    for plan in plans:
        rows = frappe.db.get_all(
            "Stock Entry Detail",
            filters={"parent": plan["material_receipt"]},
            fields=["item_code", "item_name", "qty", "uom", "t_warehouse"],
        )
        for r in rows:
            key = r["item_code"]
            if key in seen:
                continue
            # Only include items flagged as customer-provided on the Item master
            if not frappe.db.get_value("Item", r["item_code"], "is_customer_provided_item"):
                continue
            seen.add(key)
            items.append({
                "item_code": r["item_code"],
                "item_name": r["item_name"] or r["item_code"],
                "qty": r["qty"],
                "uom": r["uom"] or "Nos",
                "is_huawei": True,
                "material_receipt": plan["material_receipt"],
            })
    return items


@frappe.whitelist()
def search_items(query="", warehouse=None, limit=20):
    """Search Item master for company-owned items.

    If warehouse is given, also returns actual_qty from Bin for that warehouse.
    """
    q = (query or "").strip()
    filters = [
        ["disabled", "=", 0],
        ["is_stock_item", "=", 1],
        ["is_purchase_item", "=", 1],
        ["is_customer_provided_item", "=", 0],
    ]
    if q:
        filters.append(["item_code", "like", f"%{q}%"])

    items = frappe.db.get_all(
        "Item",
        filters=filters,
        fields=["item_code", "item_name", "stock_uom", "item_group"],
        order_by="item_code asc",
        limit=int(limit),
    )

    if not items and q:
        # fallback: search by item_name too
        name_filters = [
            ["disabled", "=", 0], ["is_stock_item", "=", 1],
            ["is_purchase_item", "=", 1], ["is_customer_provided_item", "=", 0],
            ["item_name", "like", f"%{q}%"],
        ]
        items = frappe.db.get_all("Item",
            filters=name_filters,
            fields=["item_code", "item_name", "stock_uom", "item_group"],
            order_by="item_code asc", limit=int(limit),
        )

    if warehouse and items:
        codes = [i["item_code"] for i in items]
        bins = {b["item_code"]: b["actual_qty"] for b in frappe.db.get_all(
            "Bin",
            filters={"item_code": ["in", codes], "warehouse": warehouse},
            fields=["item_code", "actual_qty"],
        )}
        for item in items:
            item["actual_qty"] = bins.get(item["item_code"], 0)

    return items


@frappe.whitelist()
def search_po_dispatches(query="", im=None, limit=20):
    """Search PO Dispatches by business POID field for the current IM."""
    conditions = [["docstatus", "!=", 2]]
    if im:
        conditions.append(["im", "=", im])
    if (query or "").strip():
        conditions.append(["poid", "like", f"%{query.strip()}%"])
    return frappe.db.get_all(
        "PO Dispatch",
        filters=conditions,
        fields=["name", "poid", "site_code", "project_code"],
        order_by="poid asc",
        limit=int(limit),
    )


@frappe.whitelist()
def get_im_teams(im=None):
    """Return teams for the current IM with their warehouse info.

    When called by an admin (System Manager / INET Admin) who has no IM record,
    returns all active teams so the material request popup can pre-select one.
    """
    if not im:
        im = frappe.db.get_value("IM Master", {"user": frappe.session.user}, "name")
    if not im:
        roles = set(frappe.get_roles(frappe.session.user))
        if roles & {"System Manager", "INET Admin", "Administrator"}:
            return frappe.db.get_all(
                "INET Team",
                filters={"status": "Active"},
                fields=["team_id", "team_name", "warehouse"],
                order_by="team_name asc",
            )
        return []
    return frappe.db.get_all(
        "INET Team",
        filters={"im": im, "status": "Active"},
        fields=["team_id", "team_name", "warehouse"],
        order_by="team_name asc",
    )


@frappe.whitelist()
def get_duid_stock_summary():
    """DUID-wise summary of INET Huawei Outbound materials in the main warehouse.

    Groups all INET Huawei Outbound Plan rows by DUID, showing
    how many shipments arrived (Received) vs. are expected (Prepared).
    DUIDs whose net stock in the source warehouse is zero (fully transferred out) are excluded.
    """
    plans = frappe.db.get_all(
        "Huawei Outbound Plan",
        filters={"subcon": "INET"},
        fields=["du_id", "bill_no", "outbound_status", "material_receipt",
                "total_volume", "outbound_date", "project_name"],
        order_by="outbound_date desc",
        limit=5000,
    )
    by_duid = {}
    for p in plans:
        duid = (p["du_id"] or "").strip()
        if not duid:
            continue
        if duid not in by_duid:
            by_duid[duid] = {
                "duid": duid,
                "total_volume": 0.0,
                "received_count": 0,
                "prepared_count": 0,
                "latest_date": str(p["outbound_date"] or ""),
                "project_name": p["project_name"] or "",
            }
        g = by_duid[duid]
        g["total_volume"] = round(g["total_volume"] + flt(p["total_volume"]), 4)
        if p["outbound_status"] == "Received":
            g["received_count"] += 1
        else:
            g["prepared_count"] += 1
        if str(p["outbound_date"] or "") > g["latest_date"]:
            g["latest_date"] = str(p["outbound_date"])

    # Build set of DUIDs that are fully transferred out of source warehouse.
    # Two checks — either is sufficient to mark a DUID as fully transferred:
    #
    #  1. SE balance: received_qty (to_duid OR duid on Receipt) - transferred_qty
    #     (duid on Transfer) - issued_qty (duid on Issue) <= 0
    #     Note: legacy SEs may have the wrong DUID column set (duid instead of
    #     to_duid on a Receipt), so we COALESCE both columns for receipts.
    #
    #  2. MR status fallback: all submitted MRs for this DUID are Transferred/
    #     Issued and none are Pending Approval — catches cases where the
    #     Transfer SE was created without the DUID dimension being set.
    fully_transferred = set()
    source_wh = frappe.db.get_single_value("INET Settings", "source_warehouse") or ""
    if source_wh:
        # Received: prefer to_duid; fall back to duid (legacy SEs had wrong direction)
        receipt_rows = frappe.db.sql("""
            SELECT COALESCE(NULLIF(sed.to_duid,''), NULLIF(sed.duid,'')) AS duid,
                   SUM(sed.qty) AS qty
            FROM `tabStock Entry Detail` sed
            JOIN `tabStock Entry` se ON se.name = sed.parent
            WHERE se.docstatus = 1
              AND se.stock_entry_type = 'Material Receipt'
              AND sed.t_warehouse = %s
              AND (NULLIF(sed.to_duid,'') IS NOT NULL OR NULLIF(sed.duid,'') IS NOT NULL)
            GROUP BY COALESCE(NULLIF(sed.to_duid,''), NULLIF(sed.duid,''))
        """, (source_wh,), as_dict=True)

        transfer_rows = frappe.db.sql("""
            SELECT sed.duid, SUM(sed.qty) AS qty
            FROM `tabStock Entry Detail` sed
            JOIN `tabStock Entry` se ON se.name = sed.parent
            WHERE se.docstatus = 1
              AND se.stock_entry_type = 'Material Transfer'
              AND sed.s_warehouse = %s
              AND sed.duid IS NOT NULL AND sed.duid != ''
            GROUP BY sed.duid
        """, (source_wh,), as_dict=True)

        issue_rows = frappe.db.sql("""
            SELECT sed.duid, SUM(sed.qty) AS qty
            FROM `tabStock Entry Detail` sed
            JOIN `tabStock Entry` se ON se.name = sed.parent
            WHERE se.docstatus = 1
              AND se.stock_entry_type = 'Material Issue'
              AND sed.s_warehouse = %s
              AND sed.duid IS NOT NULL AND sed.duid != ''
            GROUP BY sed.duid
        """, (source_wh,), as_dict=True)

        received   = {r.duid: flt(r.qty) for r in receipt_rows}
        transferred = {r.duid: flt(r.qty) for r in transfer_rows}
        issued     = {r.duid: flt(r.qty) for r in issue_rows}

        for duid, recv_qty in received.items():
            out_qty = transferred.get(duid, 0) + issued.get(duid, 0)
            if recv_qty > 0 and (recv_qty - out_qty) <= 0:
                fully_transferred.add(duid)

    # MR status fallback: DUIDs where all submitted MRs are done and none pending.
    # This catches SEs that were submitted without DUID dimension set.
    try:
        mr_rows = frappe.db.get_all(
            "Material Request",
            filters={"docstatus": 1},
            fields=["duid", "request_status"],
        )
        from collections import defaultdict
        mr_counts = defaultdict(lambda: {"done": 0, "pending": 0})
        for mr in mr_rows:
            d = (mr.get("duid") or "").strip()
            if not d:
                continue
            s = mr.get("request_status") or ""
            if s in ("Transferred", "Issued"):
                mr_counts[d]["done"] += 1
            elif s == "Pending Approval":
                mr_counts[d]["pending"] += 1
        for d, counts in mr_counts.items():
            if counts["done"] > 0 and counts["pending"] == 0:
                fully_transferred.add(d)
    except Exception:
        pass

    # Sort: received first, then by latest date; exclude fully-transferred DUIDs
    result = sorted(
        (v for v in by_duid.values() if v["duid"] not in fully_transferred),
        key=lambda x: (-x["received_count"], x["latest_date"]),
        reverse=False,
    )
    return result


@frappe.whitelist()
def create_material_request(payload):
    """Create and submit a Material Request (type: Material Transfer) from the portal."""
    import json
    roles = set(frappe.get_roles(frappe.session.user))
    if not roles & {"Administrator", "System Manager", "Stock Manager", "INET Admin", "INET IM"}:
        frappe.throw("Not permitted to create material requests.", frappe.PermissionError)

    data = json.loads(payload) if isinstance(payload, str) else payload

    items = data.get("items") or []
    if not items:
        frappe.throw("At least one item is required.")
    if not (data.get("team") or "").strip():
        frappe.throw("Please select a team.")

    im = data.get("im") or frappe.db.get_value("IM Master", {"user": frappe.session.user}, "name")
    req_date = data.get("request_date") or nowdate()
    duid = data.get("duid") or ""
    company = frappe.defaults.get_global_default("company")

    # Resolve business POID → system document name for the Link field
    poid_input = (data.get("poid") or "").strip()
    poid = _resolve_po_dispatch(poid_input) if poid_input else ""

    # Source warehouse always comes from INET Settings (not user input)
    source_wh = frappe.db.get_single_value("INET Settings", "source_warehouse") or ""

    # Team warehouse auto-filled from INET Team record
    team = (data.get("team") or "").strip()
    target_wh = ""
    if team:
        target_wh = frappe.db.get_value("INET Team", team, "warehouse") or ""
    if not target_wh:
        frappe.throw("Team Warehouse not configured. Please set a Warehouse on the selected team in INET Team.")

    # Prevent duplicate active MRs for the same POID + team warehouse
    if poid and target_wh:
        existing_mr = frappe.db.get_value(
            "Material Request",
            {"poid": poid, "set_warehouse": target_wh,
             "material_request_type": "Material Transfer", "docstatus": ["!=", 2]},
            "name",
        )
        if existing_mr:
            frappe.throw(
                f"Material Request {existing_mr} already exists for this POID and team. "
                "Cancel the existing request before creating a new one.",
                title="Duplicate Request",
            )

    doc = frappe.get_doc({
        "doctype": "Material Request",
        "material_request_type": "Material Transfer",
        "transaction_date": req_date,
        "schedule_date": req_date,
        "company": company,
        "set_warehouse": target_wh,
        "set_from_warehouse": source_wh,
        "im": im,
        "poid": poid,
        "duid": duid,
        "items": [
            {
                "item_code": i["item_code"],
                "qty": flt(i.get("qty", 0)),
                "uom": i.get("uom") or frappe.db.get_value("Item", i["item_code"], "stock_uom") or "",
                "warehouse": target_wh,
                "from_warehouse": source_wh,
                "schedule_date": req_date,
                "poid": poid,
                "duid": duid,
            }
            for i in items
        ],
    })
    doc.insert(ignore_permissions=True)
    doc.submit()
    frappe.db.commit()
    return {"name": doc.name, "status": "Pending Approval"}


@frappe.whitelist()
def approve_material_request(name):
    """Stock Manager approves: creates Material Transfer via ERPNext make_stock_entry,
    then sets duid + poid on the generated Stock Entry before submitting."""
    frappe.only_for(["System Manager", "Stock Manager"])

    mr = frappe.get_doc("Material Request", name)
    if mr.docstatus == 2:
        frappe.throw("This request has been cancelled.")
    if mr.transfer_status == "Completed":
        frappe.throw("Material Transfer already completed for this request.")
    if mr.docstatus != 1:
        frappe.throw(f"Cannot approve a request in status '{mr.status}'. Submit it first.")

    from erpnext.stock.doctype.material_request.material_request import make_stock_entry
    se = make_stock_entry(name)

    # Inject INET custom dimensions
    duid = mr.get("duid") or ""
    poid_val = mr.get("poid") or ""
    se.poid = poid_val
    for item in se.items:
        item.duid = duid

    se.insert(ignore_permissions=True)
    se.submit()
    frappe.db.commit()
    return {"name": name, "stock_entry": se.name, "status": "Approved"}


@frappe.whitelist()
def reject_material_request(name, reason=None):
    """Stock Manager rejects: stores reason then cancels the Material Request."""
    frappe.only_for(["System Manager", "Stock Manager"])

    mr = frappe.get_doc("Material Request", name)
    if mr.status not in ("Draft", "Submitted"):
        frappe.throw(f"Cannot reject a request in status '{mr.status}'.")

    if reason:
        frappe.db.set_value("Material Request", name, "rejection_reason", reason)

    if mr.docstatus == 1:
        mr.cancel()
    frappe.db.commit()
    return {"name": name, "status": "Rejected"}


def _create_material_issue_se(mr_name, qty_map=None):
    """Internal: create and submit a Material Issue SE for materials used in an execution.

    qty_map — dict {item_code: qty_used} from Daily Execution Material child table.
              If None/empty, issues the full transferred qty per item.
    Idempotent: returns existing SE if one already exists for this MR.
    Returns {"name": mr_name, "stock_entry": se.name, "status": ...}.
    """
    qty_map = qty_map or {}

    # Idempotency: return existing Issue SE if already created for this MR
    existing_issue = frappe.db.sql(
        """SELECT se.name FROM `tabStock Entry` se
           JOIN `tabStock Entry Detail` sed ON sed.parent = se.name
           WHERE se.docstatus = 1
             AND se.stock_entry_type = 'Material Issue'
             AND sed.material_request = %s
           LIMIT 1""",
        (mr_name,), as_list=True,
    )
    if existing_issue:
        return {"name": mr_name, "stock_entry": existing_issue[0][0], "status": "Already Issued"}

    # Verify at least one submitted Transfer SE exists for this MR
    has_transfer = frappe.db.sql(
        """SELECT 1 FROM `tabStock Entry` se
           JOIN `tabStock Entry Detail` sed ON sed.parent = se.name
           WHERE se.docstatus = 1
             AND se.stock_entry_type = 'Material Transfer'
             AND sed.material_request = %s
           LIMIT 1""",
        (mr_name,), as_list=True,
    )
    if not has_transfer:
        frappe.throw("No completed Material Transfer found for this request. Transfer the materials first.")

    # Get DUID per item from the Transfer SE Detail — this is the correct
    # inventory dimension value (Link to DUID Master) rather than the
    # plain string on the MR header.
    duid_by_item = {}
    transfer_item_rows = frappe.db.sql(
        """SELECT sed.item_code, sed.duid
           FROM `tabStock Entry` se
           JOIN `tabStock Entry Detail` sed ON sed.parent = se.name
           WHERE se.docstatus = 1
             AND se.stock_entry_type = 'Material Transfer'
             AND sed.material_request = %s""",
        (mr_name,), as_dict=True,
    )
    for row in transfer_item_rows:
        if row.item_code not in duid_by_item and row.duid:
            duid_by_item[row.item_code] = row.duid

    mr = frappe.get_doc("Material Request", mr_name)
    mr_duid = mr.get("duid") or ""  # fallback when SE item has no duid
    poid_val = mr.get("poid") or ""
    team_wh = mr.set_warehouse

    se_items = []
    for item in mr.items:
        used_qty = flt(qty_map.get(item.item_code, item.qty))
        if used_qty <= 0:
            continue
        se_items.append({
            "item_code": item.item_code,
            "qty": used_qty,
            "uom": item.uom or item.stock_uom,
            "s_warehouse": team_wh,
            "duid": duid_by_item.get(item.item_code) or mr_duid,
            "material_request": mr_name,
            "material_request_item": item.name,
        })

    if not se_items:
        frappe.throw("No items to issue (all quantities are zero).")

    se = frappe.get_doc({
        "doctype": "Stock Entry",
        "stock_entry_type": "Material Issue",
        "purpose": "Material Issue",
        "from_warehouse": team_wh,
        "poid": poid_val,
        "items": se_items,
    })
    se.insert(ignore_permissions=True)
    se.submit()
    frappe.db.commit()
    return {"name": mr_name, "stock_entry": se.name, "status": "Issued"}


@frappe.whitelist()
def issue_materials_for_work_done(name, qty_overrides=None):
    """Create Material Issue Stock Entry from team warehouse when work is done.

    qty_overrides — optional JSON dict mapping item_code → qty_used.
    When provided, only that qty is issued (0 = skip item).
    If not provided, checks Daily Execution material_usage for field team adjustments.
    """
    frappe.only_for(["System Manager", "Stock Manager"])

    # parse qty_overrides if JSON string
    if isinstance(qty_overrides, str):
        try:
            qty_overrides = frappe.parse_json(qty_overrides)
        except Exception:
            qty_overrides = None
    qty_map = qty_overrides or {}

    # If no overrides passed, check if field team saved usage on the DE child table
    if not qty_map:
        poid = frappe.db.get_value("Material Request", name, "poid")
        if poid:
            de_name = frappe.db.get_value(
                "Daily Execution", {"system_id": poid}, "name", order_by="modified desc"
            )
            if de_name and frappe.db.exists("DocType", "Daily Execution Material"):
                rows = frappe.db.get_all(
                    "Daily Execution Material",
                    filters={"parent": de_name},
                    fields=["item_code", "qty_used"],
                )
                qty_map = {r["item_code"]: flt(r["qty_used"]) for r in rows if r.get("item_code")}

    return _create_material_issue_se(name, qty_map)


# ─── Phase 3: Stock Balance & POID Material APIs ─────────────────────────────

@frappe.whitelist()
def get_team_material_stock(team_id=None):
    """Return current warehouse stock for one or more INET teams.

    - Field team user: their own team's stock (team_id auto-resolved from session).
    - IM: all active teams under their supervision.
    - Admin / System Manager: all active INET teams (or specific team_id if provided).

    Returns a list of:
      { team_id, team_name, warehouse, items: [{ item_code, item_name, qty, uom }] }
    """
    roles = set(frappe.get_roles(frappe.session.user))
    is_admin = bool(roles & {"System Manager", "INET Admin", "Administrator"})
    is_im = "INET IM" in roles

    # Determine which teams to fetch
    if team_id:
        teams = frappe.db.get_all(
            "INET Team", filters={"name": team_id, "status": "Active"},
            fields=["name as team_id", "team_name", "warehouse"],
            ignore_permissions=True,
        )
    elif is_admin:
        teams = frappe.db.get_all(
            "INET Team", filters={"status": "Active"},
            fields=["name as team_id", "team_name", "warehouse"],
            order_by="team_name asc", ignore_permissions=True,
        )
    elif is_im:
        im = frappe.db.get_value("IM Master", {"user": frappe.session.user}, "name")
        if not im:
            return []
        teams = frappe.db.get_all(
            "INET Team", filters={"im": im, "status": "Active"},
            fields=["name as team_id", "team_name", "warehouse"],
            order_by="team_name asc", ignore_permissions=True,
        )
    else:
        # Field user — resolve their team.
        # Primary: field_user field on INET Team matches the session user.
        # Fallback: find via team member employee → user_id link.
        resolved_team = frappe.db.get_value(
            "INET Team", {"field_user": frappe.session.user, "status": "Active"}, "name"
        )
        if not resolved_team:
            emp = frappe.db.get_value("Employee", {"user_id": frappe.session.user, "status": "Active"}, "name")
            if emp:
                member_parent = frappe.db.get_value("INET Team Member", {"employee": emp}, "parent")
                if member_parent:
                    team_status = frappe.db.get_value("INET Team", member_parent, "status")
                    if team_status == "Active":
                        resolved_team = member_parent
        if not resolved_team:
            return []
        teams = frappe.db.get_all(
            "INET Team", filters={"name": resolved_team},
            fields=["name as team_id", "team_name", "warehouse"],
            ignore_permissions=True,
        )

    out = []
    for team in teams:
        wh = team.get("warehouse") or ""
        items = []
        if wh:
            # ── Step 1: total qty per item from Bin (ground truth) ──
            bins = frappe.db.sql(
                """SELECT b.item_code,
                          IFNULL(i.item_name, b.item_code) AS item_name,
                          b.actual_qty                     AS qty,
                          IFNULL(i.stock_uom, '')          AS uom
                   FROM `tabBin` b
                   LEFT JOIN `tabItem` i ON i.name = b.item_code
                   WHERE b.warehouse = %s AND b.actual_qty > 0
                   ORDER BY i.item_name""",
                (wh,), as_dict=True,
            )

            # ── Step 2: per-DUID balance from SE Detail ──
            # SLE does not carry the duid inventory dimension column in this
            # installation. Reconstruct per-DUID balance from Stock Entry Detail:
            #   Transfer SE  → items arriving  (t_warehouse=team, to_duid set)
            #   Issue SE     → items consumed  (s_warehouse=team, duid set)
            duid_balance = {}   # (item_code, duid) → net qty
            if bins:
                ic_list = [r["item_code"] for r in bins]
                placeholders = ", ".join(["%s"] * len(ic_list))

                # Items transferred IN to this warehouse
                in_rows = frappe.db.sql(
                    f"""SELECT sed.item_code, sed.to_duid AS duid, SUM(sed.qty) AS qty
                        FROM `tabStock Entry Detail` sed
                        JOIN `tabStock Entry` se ON se.name = sed.parent
                        WHERE se.docstatus             = 1
                          AND se.stock_entry_type      = 'Material Transfer'
                          AND sed.t_warehouse          = %s
                          AND sed.to_duid IS NOT NULL AND sed.to_duid != ''
                          AND sed.item_code IN ({placeholders})
                        GROUP BY sed.item_code, sed.to_duid""",
                    (wh, *ic_list), as_dict=True,
                )
                for r in in_rows:
                    key = (r["item_code"], r["duid"])
                    duid_balance[key] = duid_balance.get(key, 0.0) + flt(r["qty"])

                # Items issued OUT from this warehouse
                out_rows = frappe.db.sql(
                    f"""SELECT sed.item_code, sed.duid, SUM(sed.qty) AS qty
                        FROM `tabStock Entry Detail` sed
                        JOIN `tabStock Entry` se ON se.name = sed.parent
                        WHERE se.docstatus         = 1
                          AND se.stock_entry_type  = 'Material Issue'
                          AND sed.s_warehouse      = %s
                          AND sed.duid IS NOT NULL AND sed.duid != ''
                          AND sed.item_code IN ({placeholders})
                        GROUP BY sed.item_code, sed.duid""",
                    (wh, *ic_list), as_dict=True,
                )
                for r in out_rows:
                    key = (r["item_code"], r["duid"])
                    duid_balance[key] = duid_balance.get(key, 0.0) - flt(r["qty"])

            # ── Step 3: build item_map with per-DUID sources ──
            item_map = {}
            for r in bins:
                ic = r["item_code"]
                uom = r["uom"] or ""
                sources = [
                    {"duid": duid, "qty": round(qty, 4), "uom": uom, "poid": "", "material_request": ""}
                    for (item_code, duid), qty in duid_balance.items()
                    if item_code == ic and qty > 0
                ]
                item_map[ic] = {
                    "item_code": ic,
                    "item_name": r["item_name"] or ic,
                    "uom": uom,
                    "qty": flt(r["qty"]),
                    "sources": sources,
                }

            # ── Step 3: enrich sources with POID from Material Request ──
            # Join MR → PO Dispatch to get the business POID for each DUID.
            if item_map:
                ic_list = list(item_map.keys())
                placeholders = ", ".join(["%s"] * len(ic_list))
                mr_rows = frappe.db.sql(
                    f"""SELECT mri.item_code,
                               IFNULL(pd.poid, '') AS poid,
                               IFNULL(mr.duid, '') AS duid,
                               mr.name             AS material_request
                        FROM `tabMaterial Request Item` mri
                        JOIN `tabMaterial Request` mr ON mr.name = mri.parent
                        LEFT JOIN `tabPO Dispatch` pd ON pd.name = mr.poid
                        WHERE mr.set_warehouse          = %s
                          AND mr.material_request_type  = 'Material Transfer'
                          AND mr.docstatus              = 1
                          AND mri.item_code IN ({placeholders})
                        ORDER BY mr.creation DESC""",
                    (wh, *ic_list), as_dict=True,
                )
                # (item_code, duid) → first matching MR with a POID
                mr_lookup = {}
                for r in mr_rows:
                    key = (r["item_code"], r["duid"])
                    if key not in mr_lookup and r["poid"]:
                        mr_lookup[key] = {"poid": r["poid"], "material_request": r["material_request"]}

                for ic, item in item_map.items():
                    for s in item["sources"]:
                        info = mr_lookup.get((ic, s["duid"])) or {}
                        s["poid"] = info.get("poid", "")
                        s["material_request"] = info.get("material_request", "")

            # ── Step 4: build final items list ──
            for item in item_map.values():
                has_customer = any(s["poid"] for s in item["sources"])
                item["item_type"] = "customer" if has_customer else "company"
                items.append(item)

        out.append({
            "team_id": team["team_id"],
            "team_name": team.get("team_name") or team["team_id"],
            "warehouse": wh,
            "items": items,
        })
    return out


@frappe.whitelist()
def get_available_stock(item_code, warehouse=None):
    """Return available qty for an item, optionally scoped to a warehouse."""
    if not item_code:
        return []
    if warehouse:
        qty = frappe.db.get_value("Bin", {"item_code": item_code, "warehouse": warehouse}, "actual_qty") or 0
        return [{"warehouse": warehouse, "qty": flt(qty)}]
    rows = frappe.db.get_all("Bin", filters={"item_code": item_code, "actual_qty": [">", 0]},
                              fields=["warehouse", "actual_qty"])
    return [{"warehouse": r.warehouse, "qty": flt(r.actual_qty)} for r in rows]


@frappe.whitelist()
def get_poid_materials(po_dispatch):
    """Return material items transferred for a POID (for field execution form)."""
    # resolve po_dispatch name
    pd_name = _resolve_po_dispatch(po_dispatch) if po_dispatch else None
    if not pd_name:
        return []
    # Get all submitted Material Requests for this POID
    mrs = frappe.get_all("Material Request",
        filters={"poid": pd_name, "material_request_type": "Material Transfer", "docstatus": 1},
        fields=["name", "set_warehouse"],
        ignore_permissions=True)
    if not mrs:
        return []

    out = []
    for mr in mrs:
        team_wh = mr.set_warehouse or ""

        # Sum actual transferred qty per item from Stock Entry Detail rows
        # that reference this MR. material_request lives on SED, not SE header.
        se_items = frappe.db.sql(
            """SELECT sed.item_code,
                      IFNULL(MAX(sed.item_name), sed.item_code) AS item_name,
                      SUM(sed.qty)                               AS qty_transferred,
                      IFNULL(MAX(sed.uom), MAX(sed.stock_uom))  AS uom
               FROM `tabStock Entry`        se
               JOIN `tabStock Entry Detail` sed ON sed.parent = se.name
               WHERE se.docstatus              = 1
                 AND se.stock_entry_type       = 'Material Transfer'
                 AND sed.material_request      = %s
               GROUP BY sed.item_code""",
            (mr.name,), as_dict=True,
        )

        if se_items:
            for row in se_items:
                out.append({
                    "material_request": mr.name,
                    "transferred": True,
                    "item_code": row.item_code,
                    "item_name": row.item_name or row.item_code,
                    "qty_transferred": flt(row.qty_transferred),
                    "qty_used": flt(row.qty_transferred),
                    "uom": row.uom or "Nos",
                    "team_warehouse": team_wh,
                })
        else:
            # SE not yet created — show MR items as Pending
            mr_items = frappe.get_all("Material Request Item",
                filters={"parent": mr.name},
                fields=["item_code", "item_name", "qty", "uom", "stock_uom"],
                ignore_permissions=True)
            for it in mr_items:
                out.append({
                    "material_request": mr.name,
                    "transferred": False,
                    "item_code": it.item_code,
                    "item_name": it.item_name or it.item_code,
                    "qty_transferred": flt(it.qty),
                    "qty_used": flt(it.qty),
                    "uom": it.uom or it.stock_uom or "Nos",
                    "team_warehouse": team_wh,
                })
    return out


# ─── Field Team Material Usage ────────────────────────────────────────────────

@frappe.whitelist()
def save_material_usage(po_dispatch, material_request, execution, usage_items):
    """Save field team's actual material usage for a POID execution."""
    if isinstance(usage_items, str):
        usage_items = frappe.parse_json(usage_items)

    existing = frappe.db.get_value("INET Material Usage",
        {"po_dispatch": po_dispatch, "material_request": material_request}, "name")
    if existing:
        frappe.db.set_value("INET Material Usage", existing, {
            "execution": execution or "",
            "usage_items": frappe.as_json(usage_items),
            "status": "Saved",
        })
        return {"name": existing, "action": "updated"}
    doc = frappe.get_doc({
        "doctype": "INET Material Usage",
        "po_dispatch": po_dispatch,
        "material_request": material_request,
        "execution": execution or "",
        "usage_items": frappe.as_json(usage_items),
        "created_by_team": frappe.session.user,
        "status": "Saved",
    })
    doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": doc.name, "action": "created"}


@frappe.whitelist()
def get_material_usage(po_dispatch, material_request=None):
    """Get saved material usage for a POID."""
    filters = {"po_dispatch": po_dispatch}
    if material_request:
        filters["material_request"] = material_request
    rows = frappe.db.get_all("INET Material Usage", filters=filters,
                              fields=["name", "material_request", "execution", "usage_items"],
                              order_by="creation desc", limit=1)
    if not rows:
        return None
    r = rows[0]
    try:
        r["usage_items"] = frappe.parse_json(r["usage_items"]) if r.get("usage_items") else []
    except Exception:
        r["usage_items"] = []
    return r


# ─── Material Return Flow ──────────────────────────────────────────────────────


def _get_team_duid_per_item(team_wh, item_codes):
    """Return {item_code: duid} for items in a team warehouse.

    Uses the Transfer SE Detail to_duid (the DUID items arrived under).
    Returns the DUID with the highest transferred qty per item, which is the
    primary source DUID to tag on the return SE.
    """
    if not item_codes or not team_wh:
        return {}
    placeholders = ", ".join(["%s"] * len(item_codes))
    rows = frappe.db.sql(
        f"""SELECT sed.item_code, sed.to_duid, SUM(sed.qty) AS total_qty
            FROM `tabStock Entry Detail` sed
            JOIN `tabStock Entry` se ON se.name = sed.parent
            WHERE se.docstatus = 1
              AND se.stock_entry_type = 'Material Transfer'
              AND sed.t_warehouse = %s
              AND sed.to_duid IS NOT NULL AND sed.to_duid != ''
              AND sed.item_code IN ({placeholders})
            GROUP BY sed.item_code, sed.to_duid
            ORDER BY total_qty DESC""",
        (team_wh, *item_codes), as_dict=True,
    )
    result = {}
    for r in rows:
        if r["item_code"] not in result:   # highest-qty DUID wins
            result[r["item_code"]] = r["to_duid"]
    return result


def _resolve_team_for_user():
    """Resolve INET Team name for the current field user (same logic as get_team_material_stock)."""
    resolved = frappe.db.get_value(
        "INET Team", {"field_user": frappe.session.user, "status": "Active"}, "name"
    )
    if not resolved:
        emp = frappe.db.get_value("Employee", {"user_id": frappe.session.user, "status": "Active"}, "name")
        if emp:
            parent = frappe.db.get_value("INET Team Member", {"employee": emp}, "parent")
            if parent and frappe.db.get_value("INET Team", parent, "status") == "Active":
                resolved = parent
    return resolved or ""


@frappe.whitelist()
def create_material_return_request(payload):
    """Field team (or IM on behalf of a team) submits a return request:
    team warehouse → source warehouse.  Creates a Material Request with
    is_return_request=1 that the IM then approves.
    """
    import json
    roles = set(frappe.get_roles(frappe.session.user))
    allowed = roles & {"Administrator", "System Manager", "Stock Manager", "INET Admin", "INET IM", "INET Field Team"}
    if not allowed:
        frappe.throw("Not permitted.", frappe.PermissionError)

    data = json.loads(payload) if isinstance(payload, str) else payload
    items = [i for i in (data.get("items") or []) if flt(i.get("qty", 0)) > 0]
    if not items:
        frappe.throw("At least one item with quantity > 0 is required.")

    # Resolve team
    team_id = (data.get("team_id") or "").strip()
    if not team_id:
        team_id = _resolve_team_for_user()
    if not team_id:
        frappe.throw("Team not found. Your account may not be linked to an active team.")

    team_wh = frappe.db.get_value("INET Team", team_id, "warehouse") or ""
    if not team_wh:
        frappe.throw("Team Warehouse not configured on the selected team.")

    source_wh = frappe.db.get_single_value("INET Settings", "source_warehouse") or ""
    if not source_wh:
        frappe.throw("Source Warehouse not configured in INET Settings.")

    company = frappe.defaults.get_global_default("company")
    req_date = nowdate()
    im = frappe.db.get_value("INET Team", team_id, "im") or ""
    reason = (data.get("reason") or "").strip()

    doc = frappe.get_doc({
        "doctype": "Material Request",
        "material_request_type": "Material Transfer",
        "transaction_date": req_date,
        "schedule_date": req_date,
        "company": company,
        "set_from_warehouse": team_wh,
        "set_warehouse": source_wh,
        "is_return_request": 1,
        "im": im,
        **({"return_reason": reason} if reason and frappe.db.has_column("Material Request", "return_reason") else {}),
        "items": [
            {
                "item_code": i["item_code"],
                "qty": flt(i["qty"]),
                "uom": i.get("uom") or frappe.db.get_value("Item", i["item_code"], "stock_uom") or "",
                "warehouse": source_wh,
                "from_warehouse": team_wh,
                "schedule_date": req_date,
            }
            for i in items
        ],
    })
    doc.insert(ignore_permissions=True)
    doc.submit()
    frappe.db.commit()
    return {"name": doc.name, "status": "Pending Approval"}


@frappe.whitelist()
def list_return_requests(team_id=None, status=None, limit=50):
    """List Material Return Requests.

    Field team: sees their team's requests.
    IM: sees all requests from teams under their supervision.
    Admin / Stock Manager: sees all.
    """
    roles = set(frappe.get_roles(frappe.session.user))
    is_admin = bool(roles & {"Administrator", "System Manager", "Stock Manager", "INET Admin"})
    is_im = "INET IM" in roles

    filters = {
        "material_request_type": "Material Transfer",
        "is_return_request": 1,
    }

    if is_admin:
        pass
    elif is_im:
        im = frappe.db.get_value("IM Master", {"user": frappe.session.user}, "name")
        if im:
            filters["im"] = im
        else:
            return []
    else:
        resolved = _resolve_team_for_user()
        if not resolved:
            return []
        team_wh = frappe.db.get_value("INET Team", resolved, "warehouse") or ""
        if team_wh:
            filters["set_from_warehouse"] = team_wh

    if team_id:
        team_wh_override = frappe.db.get_value("INET Team", team_id, "warehouse") or ""
        if team_wh_override:
            filters["set_from_warehouse"] = team_wh_override

    rows = frappe.db.get_all(
        "Material Request",
        filters=filters,
        fields=[
            "name", "transaction_date", "owner", "im",
            "status", "transfer_status", "set_from_warehouse", "set_warehouse",
            *( ["return_reason"] if frappe.db.has_column("Material Request", "return_reason") else [] ),
        ],
        order_by="`tabMaterial Request`.transaction_date desc, `tabMaterial Request`.creation desc",
        limit=int(limit),
    )

    # Resolve team names from warehouse reverse-lookup (batch)
    warehouses = list({r["set_from_warehouse"] for r in rows if r.get("set_from_warehouse")})
    wh_team_map = {}
    if warehouses:
        for t in frappe.db.get_all(
            "INET Team", filters={"warehouse": ["in", warehouses]},
            fields=["warehouse", "name", "team_name"],
        ):
            wh_team_map[t["warehouse"]] = {"team_id": t["name"], "team_name": t["team_name"]}

    for r in rows:
        r["request_date"] = str(r.pop("transaction_date", "") or "")
        r["request_status"] = _request_status(r["status"], r["transfer_status"])
        team_info = wh_team_map.get(r.get("set_from_warehouse") or "", {})
        r["team_id"] = team_info.get("team_id", "")
        r["team_name"] = team_info.get("team_name") or r.get("set_from_warehouse", "—")
        r["team_warehouse"] = r.get("set_from_warehouse", "")
        r["reason"] = r.pop("return_reason", "") or ""

    if status:
        rows = [r for r in rows if r["request_status"] == status]

    return rows


@frappe.whitelist()
def approve_material_return_request(name):
    """IM/Stock Manager approves a return request.

    Creates a Material Transfer SE: s_warehouse = team WH → t_warehouse = source WH.
    Sets duid = team DUID per item (inventory dimension for source-side tracking).
    """
    frappe.only_for(["System Manager", "Stock Manager"])

    mr = frappe.get_doc("Material Request", name)
    if not mr.get("is_return_request"):
        frappe.throw("This is not a return request. Use approve_material_request instead.")
    if mr.docstatus == 2:
        frappe.throw("This request has been cancelled.")
    if mr.transfer_status == "Completed":
        frappe.throw("Transfer already completed for this request.")
    if mr.docstatus != 1:
        frappe.throw(f"Cannot approve a request in status '{mr.status}'. Submit it first.")

    from erpnext.stock.doctype.material_request.material_request import make_stock_entry
    se = make_stock_entry(name)

    # Tag each item with the team DUID (inventory dimension — source side of transfer)
    team_wh = mr.set_from_warehouse
    item_codes = [i.item_code for i in mr.items]
    duid_by_item = _get_team_duid_per_item(team_wh, item_codes)

    for item in se.items:
        duid = duid_by_item.get(item.item_code, "")
        if duid:
            item.duid = duid   # before_stock_entry_submit respects pre-set values

    se.insert(ignore_permissions=True)
    se.submit()
    frappe.db.commit()
    return {"name": name, "stock_entry": se.name, "status": "Approved"}


@frappe.whitelist()
def create_direct_return_transfer(payload):
    """IM creates a Material Transfer SE directly (team WH → source WH) without MR.

    Used when the IM wants to return materials without a field-team request.
    """
    frappe.only_for(["System Manager", "Stock Manager"])
    import json

    data = json.loads(payload) if isinstance(payload, str) else payload
    team_id = (data.get("team_id") or "").strip()
    items = [i for i in (data.get("items") or []) if flt(i.get("qty", 0)) > 0]

    if not team_id:
        frappe.throw("Team is required.")
    if not items:
        frappe.throw("At least one item with quantity > 0 is required.")

    team_wh = frappe.db.get_value("INET Team", team_id, "warehouse") or ""
    if not team_wh:
        frappe.throw("Team Warehouse not configured on the selected team.")

    source_wh = frappe.db.get_single_value("INET Settings", "source_warehouse") or ""
    if not source_wh:
        frappe.throw("Source Warehouse not configured in INET Settings.")

    company = frappe.defaults.get_global_default("company")
    item_codes = [i["item_code"] for i in items]
    duid_by_item = _get_team_duid_per_item(team_wh, item_codes)

    se = frappe.get_doc({
        "doctype": "Stock Entry",
        "stock_entry_type": "Material Transfer",
        "company": company,
        "items": [
            {
                "item_code": i["item_code"],
                "qty": flt(i["qty"]),
                "uom": i.get("uom") or frappe.db.get_value("Item", i["item_code"], "stock_uom") or "",
                "s_warehouse": team_wh,
                "t_warehouse": source_wh,
                "duid": duid_by_item.get(i["item_code"], ""),
            }
            for i in items
        ],
    })
    se.insert(ignore_permissions=True)
    se.submit()
    frappe.db.commit()
    return {"stock_entry": se.name, "status": "Transferred"}
