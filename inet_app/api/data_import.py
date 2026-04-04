import frappe
from frappe.utils import flt, cint


@frappe.whitelist()
def import_control_center_xlsx(file_url):
	"""
	Read an uploaded CONTROL_CENTER.xlsx file and import data into master doctypes.

	Args:
		file_url (str): URL of the uploaded file (e.g. /private/files/CONTROL_CENTER.xlsx)

	Returns:
		dict: counts per entity type imported
	"""
	try:
		import openpyxl
	except ImportError:
		frappe.throw("openpyxl is required. Install it with: pip install openpyxl")

	# Resolve the file path from the URL
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	file_path = file_doc.get_full_path()

	wb = openpyxl.load_workbook(file_path, data_only=True)

	counts = {
		"Area Master": 0,
		"INET Team": 0,
		"Subcontractor Master": 0,
		"Project Control Center": 0,
		"Customer Item Master": 0,
		"Activity Cost Master": 0,
		"Subcontract Cost Master": 0,
		"Visit Multiplier Master": 0,
	}

	# ------------------------------------------------------------------ #
	# 04_AREA_MASTER: Row 3+, col 0=area_code, col 1=area_name            #
	# ------------------------------------------------------------------ #
	if "04_AREA_MASTER" in wb.sheetnames:
		ws = wb["04_AREA_MASTER"]
		for row in ws.iter_rows(min_row=3, values_only=True):
			area_code = row[0]
			if not area_code:
				continue
			area_code = str(area_code).strip()
			if not area_code:
				continue
			if not frappe.db.exists("Area Master", area_code):
				frappe.get_doc({
					"doctype": "Area Master",
					"area_code": area_code,
					"area_name": str(row[1]).strip() if row[1] else "",
				}).insert(ignore_permissions=True)
				counts["Area Master"] += 1

	# ------------------------------------------------------------------ #
	# 07_TEAM_MASTER: Row 3+                                               #
	# col 0=team_id, 1=team_name, 2=im, 3=team_type, 4=subcontractor,    #
	# 5=status, 6=daily_cost_applies, 7=daily_cost                        #
	# ------------------------------------------------------------------ #
	if "07_TEAM_MASTER" in wb.sheetnames:
		ws = wb["07_TEAM_MASTER"]
		for row in ws.iter_rows(min_row=3, values_only=True):
			team_id = row[0]
			if not team_id:
				continue
			team_id = str(team_id).strip()
			if not team_id:
				continue
			if not frappe.db.exists("INET Team", team_id):
				daily_cost_applies_raw = row[6]
				if isinstance(daily_cost_applies_raw, str):
					daily_cost_applies = 1 if daily_cost_applies_raw.strip().lower() == "yes" else 0
				else:
					daily_cost_applies = cint(daily_cost_applies_raw)

				frappe.get_doc({
					"doctype": "INET Team",
					"team_id": team_id,
					"team_name": str(row[1]).strip() if row[1] else "",
					"im": str(row[2]).strip() if row[2] else "",
					"team_type": str(row[3]).strip() if row[3] else "",
					"subcontractor": str(row[4]).strip() if row[4] else "",
					"status": str(row[5]).strip() if row[5] else "",
					"daily_cost_applies": daily_cost_applies,
					"daily_cost": flt(row[7]) if row[7] else 0,
				}).insert(ignore_permissions=True)
				counts["INET Team"] += 1

	# ------------------------------------------------------------------ #
	# 08_SUBCONTRACTOR_MASTER: Row 3+                                      #
	# col 0=subcontractor, 1=type, 2=inet_margin_pct, 3=sub_payout_pct,  #
	# 4=contract_model, 5=status                                           #
	# ------------------------------------------------------------------ #
	if "08_SUBCONTRACTOR_MASTER" in wb.sheetnames:
		ws = wb["08_SUBCONTRACTOR_MASTER"]
		for row in ws.iter_rows(min_row=3, values_only=True):
			subcontractor = row[0]
			if not subcontractor:
				continue
			subcontractor = str(subcontractor).strip()
			if not subcontractor:
				continue
			if not frappe.db.exists("Subcontractor Master", subcontractor):
				# Decimals in Excel (0.2 = 20%) → multiply by 100 for Percent fields
				inet_margin_pct = flt(row[2]) * 100 if row[2] else 0
				sub_payout_pct = flt(row[3]) * 100 if row[3] else 0

				frappe.get_doc({
					"doctype": "Subcontractor Master",
					"subcontractor": subcontractor,
					"type": str(row[1]).strip() if row[1] else "",
					"inet_margin_pct": inet_margin_pct,
					"sub_payout_pct": sub_payout_pct,
					"contract_model": str(row[4]).strip() if row[4] else "",
					"status": str(row[5]).strip() if row[5] else "",
				}).insert(ignore_permissions=True)
				counts["Subcontractor Master"] += 1

	# ------------------------------------------------------------------ #
	# 03_PROJECT_DOMAIN_MASTER: Row 3+                                     #
	# col 0=project_code, 1=project_name, 2=project_domain, 3=customer,  #
	# 4=huawei_im, 5=implementation_manager, 6=active_flag               #
	# ------------------------------------------------------------------ #
	if "03_PROJECT_DOMAIN_MASTER" in wb.sheetnames:
		ws = wb["03_PROJECT_DOMAIN_MASTER"]
		for row in ws.iter_rows(min_row=3, values_only=True):
			project_code = row[0]
			if not project_code:
				continue
			project_code = str(project_code).strip()
			if not project_code:
				continue
			if not frappe.db.exists("Project Control Center", project_code):
				active_flag_raw = row[6]
				if isinstance(active_flag_raw, str):
					active_flag = active_flag_raw.strip()
				elif active_flag_raw is None:
					active_flag = "Yes"
				else:
					active_flag = "Yes" if active_flag_raw else "No"

				frappe.get_doc({
					"doctype": "Project Control Center",
					"project_code": project_code,
					"project_name": str(row[1]).strip() if row[1] else "",
					"project_domain": str(row[2]).strip() if row[2] else "",
					"customer": str(row[3]).strip() if row[3] else "",
					"huawei_im": str(row[4]).strip() if row[4] else "",
					"implementation_manager": str(row[5]).strip() if row[5] else "",
					"active_flag": active_flag,
				}).insert(ignore_permissions=True)
				counts["Project Control Center"] += 1

	# ------------------------------------------------------------------ #
	# 12_CUSTOMER_ITEM_MASTER: Row 5+                                      #
	# col 0=customer, 1=item_code, 2=customer_activity_type, 3=skip,     #
	# 4=domain, 5=item_description, 6=unit_type, 7=standard_rate_sar,    #
	# 8=hard_rate_sar, 9=skip, 10=skip, 11=active_flag                   #
	# Uses autoname — no duplicate check on a single key                  #
	# ------------------------------------------------------------------ #
	if "12_CUSTOMER_ITEM_MASTER" in wb.sheetnames:
		ws = wb["12_CUSTOMER_ITEM_MASTER"]
		for row in ws.iter_rows(min_row=5, values_only=True):
			customer = row[0]
			if not customer:
				continue
			customer = str(customer).strip()
			if not customer:
				continue

			active_flag_raw = row[11]
			if isinstance(active_flag_raw, str):
				active_flag = active_flag_raw.strip()
			elif active_flag_raw is None:
				active_flag = "Yes"
			else:
				active_flag = "Yes" if active_flag_raw else "No"

			frappe.get_doc({
				"doctype": "Customer Item Master",
				"customer": customer,
				"item_code": str(row[1]).strip() if row[1] else "",
				"customer_activity_type": str(row[2]).strip() if row[2] else "",
				"domain": str(row[4]).strip() if row[4] else "",
				"item_description": str(row[5]).strip() if row[5] else "",
				"unit_type": str(row[6]).strip() if row[6] else "",
				"standard_rate_sar": flt(row[7]) if row[7] else 0,
				"hard_rate_sar": flt(row[8]) if row[8] else 0,
				"active_flag": active_flag,
			}).insert(ignore_permissions=True)
			counts["Customer Item Master"] += 1

	# ------------------------------------------------------------------ #
	# 15_ACTIVITY_COST_MASTER: Row 4+                                      #
	# col 0=activity_code, 1=standard_activity, 2=category,              #
	# 3=base_cost_sar, 4=cost_type, 5=skip, 6=billing_type,              #
	# 7=skip, 8=active_flag                                               #
	# ------------------------------------------------------------------ #
	if "15_ACTIVITY_COST_MASTER" in wb.sheetnames:
		ws = wb["15_ACTIVITY_COST_MASTER"]
		for row in ws.iter_rows(min_row=4, values_only=True):
			activity_code = row[0]
			if not activity_code:
				continue
			activity_code = str(activity_code).strip()
			if not activity_code:
				continue
			if not frappe.db.exists("Activity Cost Master", activity_code):
				active_flag_raw = row[8]
				if isinstance(active_flag_raw, str):
					active_flag = active_flag_raw.strip()
				elif active_flag_raw is None:
					active_flag = "Yes"
				else:
					active_flag = "Yes" if active_flag_raw else "No"

				frappe.get_doc({
					"doctype": "Activity Cost Master",
					"activity_code": activity_code,
					"standard_activity": str(row[1]).strip() if row[1] else "",
					"category": str(row[2]).strip() if row[2] else "",
					"base_cost_sar": flt(row[3]) if row[3] else 0,
					"cost_type": str(row[4]).strip() if row[4] else "",
					"billing_type": str(row[6]).strip() if row[6] else "",
					"active_flag": active_flag,
				}).insert(ignore_permissions=True)
				counts["Activity Cost Master"] += 1

	# ------------------------------------------------------------------ #
	# 16_SUBCONTRACT_COST_MASTER: Row 4+                                   #
	# col 0=subcontractor, 1=activity_code, 2=region_type,               #
	# 3=contract_type, 4=expected_cost_sar, 5=cost_type,                 #
	# 6=effective_from, 7=effective_to, 8=active_flag                    #
	# ------------------------------------------------------------------ #
	if "16_SUBCONTRACT_COST_MASTER" in wb.sheetnames:
		ws = wb["16_SUBCONTRACT_COST_MASTER"]
		for row in ws.iter_rows(min_row=4, values_only=True):
			subcontractor = row[0]
			if not subcontractor:
				continue
			subcontractor = str(subcontractor).strip()
			if not subcontractor:
				continue

			active_flag_raw = row[8]
			if isinstance(active_flag_raw, str):
				active_flag = active_flag_raw.strip()
			elif active_flag_raw is None:
				active_flag = "Yes"
			else:
				active_flag = "Yes" if active_flag_raw else "No"

			# effective_from and effective_to may be datetime objects from openpyxl
			effective_from = row[6]
			if hasattr(effective_from, "date"):
				effective_from = effective_from.date()
			elif effective_from:
				effective_from = str(effective_from).strip()

			effective_to = row[7]
			if hasattr(effective_to, "date"):
				effective_to = effective_to.date()
			elif effective_to:
				effective_to = str(effective_to).strip()

			frappe.get_doc({
				"doctype": "Subcontract Cost Master",
				"subcontractor": subcontractor,
				"activity_code": str(row[1]).strip() if row[1] else "",
				"region_type": str(row[2]).strip() if row[2] else "",
				"contract_type": str(row[3]).strip() if row[3] else "",
				"expected_cost_sar": flt(row[4]) if row[4] else 0,
				"cost_type": str(row[5]).strip() if row[5] else "",
				"effective_from": effective_from or None,
				"effective_to": effective_to or None,
				"active_flag": active_flag,
			}).insert(ignore_permissions=True)
			counts["Subcontract Cost Master"] += 1

	# ------------------------------------------------------------------ #
	# 17_VISIT_MULTIPLIER_MASTER: Row 4+                                   #
	# col 0=visit_type, 1=multiplier, 2=notes                             #
	# ------------------------------------------------------------------ #
	if "17_VISIT_MULTIPLIER_MASTER" in wb.sheetnames:
		ws = wb["17_VISIT_MULTIPLIER_MASTER"]
		for row in ws.iter_rows(min_row=4, values_only=True):
			visit_type = row[0]
			if not visit_type:
				continue
			visit_type = str(visit_type).strip()
			if not visit_type:
				continue
			if not frappe.db.exists("Visit Multiplier Master", visit_type):
				frappe.get_doc({
					"doctype": "Visit Multiplier Master",
					"visit_type": visit_type,
					"multiplier": flt(row[1]) if row[1] else 0,
					"notes": str(row[2]).strip() if row[2] else "",
				}).insert(ignore_permissions=True)
				counts["Visit Multiplier Master"] += 1

	frappe.db.commit()

	return counts
